#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_cronograma_premier_pet.py
Cronograma reestruturado — PremieRpet
Paleta: identidade Efcaz (#14B3CC teal + #153C5C navy + neutros)
Fonte: Arial | 2 abas | 64h total
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

OUTPUT_PATH = (
    r"c:\Users\gabriel.evangelista\Documents\ClaudeGL"
    r"\clientes\premier_pet\cronograma_premierpet_14-08-2026.xlsx"
)
LOGO_PATH = (
    r"c:\Users\gabriel.evangelista\Documents\ClaudeGL"
    r"\Documentos\logos_extraidas\slide1_img4.png"
)

# ── Paleta oficial Efcaz ──────────────────────────────────────────────────────
TEAL        = "14B3CC"   # primária Efcaz — botões, destaques
TEAL_DARK   = "0E8FA3"   # dark variant — acento esquerdo
NAVY        = "153C5C"   # secundária escura — fundos de cabeçalho

BRANCO      = "FFFFFF"
CINZA_CLARO = "F9F9F9"   # fundo alternado neutro
CINZA_BORDA = "EEEEEE"   # borda suave
TEXTO_PRINC = "333333"   # títulos
TEXTO_CORPO = "3F3F3F"   # corpo
TEXTO_SUB   = "696969"   # subtexto / notas

# ── Estrutura das Fases ───────────────────────────────────────────────────────
FASES = [
    {
        "num"    : 1,
        "label"  : "Fase 1 — Saneamento, Análise de Vencimento do Documento e Upload",
        "subtasks": [
            ("1.1", "Extração e inventário da base de documentos existente",
             "Inventário completo por fornecedor / categoria", "Impl.", "Usuário-chave", 5, "S1–S2"),
            ("1.2", "Análise de validade por categoria e linha de fornecimento",
             "Análise de status: vigente / a vencer / vencido", "Impl.", "—", 8, "S2–S3"),
            ("1.3", "Classificação de vencidos, próximos do vencimento e vigentes",
             "Ranking de criticidade documental", "Impl.", "—", 5, "S3"),
            ("1.4", "Geração de relatório de gaps documentais por fornecedor",
             "Relatório de gap por fornecedor (Excel / PDF)", "Impl. / CS", "Gestor de Compras", 7, "S3–S4"),
            ("1.5", "Repasse ao cliente: priorização e plano de ação documental",
             "Plano de regularização documental por fornecedor", "CS", "Gestor + Usuário-chave", 5, "S4"),
            ("1.6", "Upload das documentações na plataforma Efcaz ★ Resp. EFCAZ",
             "Documentos carregados na plataforma com log de execução", "EFCAZ", "—", 10, "S4–S6"),
        ],
    },
    {
        "num"    : 2,
        "label"  : "Fase 2 — Parametrização e Configuração da Base",
        "subtasks": [
            ("2.1", "Levantamento e mapeamento das linhas de fornecimento ativas",
             "Planilha de linhas de fornecimento mapeadas", "CS / Impl.", "Gestor de Compras", 2, "S6–S7"),
            ("2.2", "Configuração das linhas de fornecimento na plataforma Efcaz",
             "Linhas configuradas no sistema", "Impl.", "Gestor de Compras", 3, "S7"),
            ("2.3", "Montagem e parametrização da Matriz de Obrigatoriedade Documental",
             "Matriz validada e aplicada por linha de fornecimento", "CS / Impl.", "Gestor de Compras", 4, "S7–S8"),
            ("2.4", "Validação e ajuste fino da parametrização com equipe cliente",
             "Ata de validação e parametrização aprovada", "CS", "Gestor + Usuário-chave", 1, "S8"),
        ],
    },
    {
        "num"    : 3,
        "label"  : "Fase 3 — Certificação Cadastral e Enriquecimento",
        "subtasks": [
            ("3.1", "Preparação e limpeza da base de CNPJs para envio em lote",
             "Base de CNPJs limpa e validada", "Impl.", "Usuário-chave", 3, "S8–S9"),
            ("3.2", "Execução do envio de certificação cadastral em lote na plataforma",
             "Envio em lote realizado com log de execução", "Impl.", "—", 4, "S9"),
            ("3.3", "Monitoramento e validação dos retornos de certificação",
             "Relatório de retornos: aprovados / pendentes / erros", "Impl.", "Usuário-chave", 4, "S9–S10"),
            ("3.4", "Enriquecimento de dados e atualização de registros na plataforma",
             "Registros enriquecidos e atualizados", "Impl.", "—", 2, "S10"),
            ("3.5", "Relatório de cobertura pós-enriquecimento",
             "Relatório de cobertura cadastral final", "CS", "Gestor de Compras", 1, "S10"),
        ],
    },
]

TOTAL_HORAS = sum(t[5] for f in FASES for t in f["subtasks"])

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(c):
    return PatternFill("solid", fgColor=c)

def fnt(bold=False, color=TEXTO_CORPO, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def brd(color=CINZA_BORDA, left_color=None):
    """Borda fina uniforme; se left_color, acento medium na esquerda."""
    thin = Side(style="thin", color=color)
    left = Side(style="medium", color=left_color) if left_color else thin
    return Border(left=left, right=thin, top=thin, bottom=thin)

def rh(ws, row, h): ws.row_dimensions[row].height = h
def cw(ws, col, w): ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════════════════
#  ABA 1 — RESUMO EXECUTIVO
# ══════════════════════════════════════════════════════════════════════════════
def build_resumo(wb):
    ws = wb.create_sheet("Resumo Executivo")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDEFGH", [22, 4, 22, 4, 16, 4, 16, 4]):
        cw(ws, col, w)

    # Linha 1 — banner com logo
    rh(ws, 1, 52)
    for col in range(1, 9):
        ws.cell(row=1, column=col).fill = fill(NAVY)
    ws.merge_cells("C1:H1")
    t = ws["C1"]
    t.value     = "Cronograma de Implementação — PremieRpet"
    t.font      = fnt(bold=True, color=BRANCO, size=15)
    t.alignment = aln("left", "center")

    try:
        img = XLImage(LOGO_PATH)
        img.height = 40
        img.width  = 110
        img.anchor = "A1"
        ws.add_image(img)
    except Exception:
        ws["A1"].value = "Efcaz"
        ws["A1"].font  = fnt(bold=True, color=BRANCO, size=11)

    # Linha 2 — subtítulo
    rh(ws, 2, 16)
    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value     = "Gerado em 14/08/2026  ·  CS: Gabriel Vital  ·  Plataforma Efcaz SRM"
    s.font      = fnt(italic=True, color=TEXTO_SUB, size=9)
    s.fill      = fill(CINZA_CLARO)
    s.alignment = aln("center")

    rh(ws, 3, 10)

    # ── KPI cards (4 cards em 2 colunas de 2) ────────────────────────────────
    kpis = [
        ("=Cronograma!F22", "horas de projeto"),
        ("3",               "fases de entrega"),
        ("~8.000",          "documentos a sanear"),
        ("10 semanas",      "prazo estimado"),
    ]
    kpi_pos = [("A", "C"), ("E", "G")]   # col valor, col label — linha 4 e 5

    # 2 cards por linha, 2 linhas
    for card_idx, (val, label) in enumerate(kpis):
        lin_base = 4 + (card_idx // 2) * 3   # linhas 4,5 e 7,8
        col_val  = kpi_pos[card_idx % 2][0]
        col_lbl  = kpi_pos[card_idx % 2][1]

        rh(ws, lin_base, 30)
        rh(ws, lin_base + 1, 16)

        cv = ws[f"{col_val}{lin_base}"]
        cv.value     = val
        cv.font      = fnt(bold=True, color=BRANCO, size=20)
        cv.fill      = fill(TEAL if card_idx % 2 == 0 else NAVY)
        cv.alignment = aln("center")

        cl = ws[f"{col_lbl}{lin_base}"]
        cl.value     = label
        cl.font      = fnt(color=BRANCO, size=9)
        cl.fill      = fill(TEAL if card_idx % 2 == 0 else NAVY)
        cl.alignment = aln("left", "center")

        # linha inferior do card (texto)
        for col_l in (col_val, col_lbl):
            bc = ws[f"{col_l}{lin_base + 1}"]
            bc.fill = fill(TEAL_DARK if card_idx % 2 == 0 else "112F47")

    rh(ws, 6, 6)   # espaço entre as duas linhas de cards
    rh(ws, 9, 10)  # espaço após cards

    # ── Tabela de distribuição ────────────────────────────────────────────────
    rh(ws, 10, 20)
    ws.merge_cells("A10:H10")
    h10 = ws["A10"]
    h10.value     = "Distribuição de Horas por Fase"
    h10.font      = fnt(bold=True, color=BRANCO, size=11)
    h10.fill      = fill(NAVY)
    h10.alignment = aln("center")

    # Cabeçalhos da tabela
    rh(ws, 11, 15)
    for col_l, val in [("A", "Fase"), ("C", "Descrição"), ("E", "Horas"), ("G", "% do Total")]:
        ws.merge_cells(f"{col_l}11:{'B' if col_l=='A' else 'D' if col_l=='C' else 'F' if col_l=='E' else 'H'}11")
        c = ws[f"{col_l}11"]
        c.value     = val
        c.font      = fnt(bold=True, color=BRANCO, size=9)
        c.fill      = fill(TEAL)
        c.alignment = aln("center")
        c.border    = brd()

    # Linhas de fases (referencias ao Cronograma)
    fase_refs = [
        ("Fase 1", "Saneamento, Análise de Vencimento e Upload",  "=Cronograma!F4",  "=Cronograma!H4"),
        ("Fase 2", "Parametrização e Configuração da Base",        "=Cronograma!F11", "=Cronograma!H11"),
        ("Fase 3", "Certificação Cadastral e Enriquecimento",      "=Cronograma!F16", "=Cronograma!H16"),
    ]

    for i, (flbl, fdesc, href, pref) in enumerate(fase_refs, start=12):
        rh(ws, i, 18)
        row_bg = BRANCO if i % 2 == 0 else CINZA_CLARO

        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:D{i}")
        ws.merge_cells(f"E{i}:F{i}")
        ws.merge_cells(f"G{i}:H{i}")

        fa = ws[f"A{i}"]
        fa.value     = flbl
        fa.font      = fnt(bold=True, color=TEAL_DARK, size=10)
        fa.fill      = fill(row_bg)
        fa.alignment = aln("center")
        fa.border    = brd(left_color=TEAL)

        fd = ws[f"C{i}"]
        fd.value     = fdesc
        fd.font      = fnt(size=9, color=TEXTO_CORPO)
        fd.fill      = fill(row_bg)
        fd.alignment = aln("left")
        fd.border    = brd()

        fh = ws[f"E{i}"]
        fh.value     = href
        fh.font      = fnt(bold=True, size=10, color=TEXTO_PRINC)
        fh.fill      = fill(row_bg)
        fh.alignment = aln("center")
        fh.border    = brd()

        fp = ws[f"G{i}"]
        fp.value         = pref
        fp.font          = fnt(bold=True, size=10, color=TEAL_DARK)
        fp.fill          = fill(row_bg)
        fp.alignment     = aln("center")
        fp.number_format = "0%"
        fp.border        = brd()

    # Linha TOTAL
    total_r = 15
    rh(ws, total_r, 20)
    ws.merge_cells(f"A{total_r}:D{total_r}")
    ws.merge_cells(f"E{total_r}:F{total_r}")
    ws.merge_cells(f"G{total_r}:H{total_r}")

    tl = ws[f"A{total_r}"]
    tl.value     = "Total do Projeto"
    tl.font      = fnt(bold=True, color=BRANCO, size=10)
    tl.fill      = fill(NAVY)
    tl.alignment = aln("center")
    tl.border    = brd(NAVY)

    th = ws[f"E{total_r}"]
    th.value     = "=Cronograma!F22"
    th.font      = fnt(bold=True, color=BRANCO, size=10)
    th.fill      = fill(NAVY)
    th.alignment = aln("center")
    th.border    = brd(NAVY)

    tp = ws[f"G{total_r}"]
    tp.value         = 1
    tp.number_format = "0%"
    tp.font          = fnt(bold=True, color=BRANCO, size=10)
    tp.fill          = fill(NAVY)
    tp.alignment     = aln("center")
    tp.border        = brd(NAVY)

    rh(ws, 16, 10)

    # Nota
    rh(ws, 17, 14)
    ws.merge_cells("A17:H17")
    n = ws["A17"]
    n.value     = "★  Etapa 1.6 — Upload das documentações é responsabilidade da EFCAZ. Percentuais recalculam automaticamente ao editar horas no Cronograma."
    n.font      = fnt(italic=True, color=TEXTO_SUB, size=8)
    n.alignment = aln("left")


# ══════════════════════════════════════════════════════════════════════════════
#  ABA 2 — CRONOGRAMA
# ══════════════════════════════════════════════════════════════════════════════
def build_cronograma(wb):
    ws = wb.create_sheet("Cronograma")
    ws.sheet_view.showGridLines = False

    # Colunas: A=# | B=Atividade | C=Entregável | D=Resp.Efcaz | E=Resp.Cliente
    #          F=Horas | G=Semana | H=%Total | I=Status
    for col, w in zip("ABCDEFGHI", [5, 52, 36, 14, 20, 8, 12, 10, 18]):
        cw(ws, col, w)

    # ── Banner ────────────────────────────────────────────────────────────────
    rh(ws, 1, 52)
    for col in range(1, 10):
        ws.cell(row=1, column=col).fill = fill(NAVY)
    ws.merge_cells("B1:I1")
    t = ws["B1"]
    t.value     = "Cronograma Executivo de Projeto — PremieRpet"
    t.font      = fnt(bold=True, color=BRANCO, size=14)
    t.fill      = fill(NAVY)
    t.alignment = aln("left", "center")

    try:
        img = XLImage(LOGO_PATH)
        img.height = 40
        img.width  = 110
        img.anchor = "A1"
        ws.add_image(img)
    except Exception:
        ws["A1"].value = "Efcaz"
        ws["A1"].font  = fnt(bold=True, color=BRANCO, size=10)

    rh(ws, 2, 15)
    ws.merge_cells("A2:I2")
    s = ws["A2"]
    s.value     = "Escopo: 64 horas  ·  Consultoria & Implementação Efcaz  ·  Agosto / 2026"
    s.font      = fnt(italic=True, color=TEXTO_SUB, size=9)
    s.fill      = fill(CINZA_CLARO)
    s.alignment = aln("center")

    # ── Cabeçalho de colunas ──────────────────────────────────────────────────
    rh(ws, 3, 18)
    for ci, label in enumerate(
        ["#", "Atividade", "Entregável", "Resp. Efcaz", "Resp. Cliente",
         "Horas", "Semana", "% do Total", "Status"], start=1
    ):
        c = ws.cell(row=3, column=ci, value=label)
        c.font      = fnt(bold=True, color=BRANCO, size=9)
        c.fill      = fill(TEAL)
        c.alignment = aln("center")
        c.border    = brd(TEAL)

    ws.freeze_panes = "B4"

    # ── Pré-calcula linhas ────────────────────────────────────────────────────
    current_row = 4
    phase_meta  = []
    for fase in FASES:
        hr   = current_row
        fs   = hr + 1
        ls   = fs + len(fase["subtasks"]) - 1
        phase_meta.append((hr, fs, ls, fase))
        current_row = ls + 1
    TOTAL_ROW = current_row
    leaf_rows = [r for _, fs, ls, _ in phase_meta for r in range(fs, ls + 1)]

    # ── Escreve fases ─────────────────────────────────────────────────────────
    status_opts = '"🕐 Pendente,🔄 Em Andamento,✅ Concluído,⏸ Pausado"'

    for header_row, first_sub, last_sub, fase in phase_meta:

        # — Header da fase ————————————————————————————————————————————————————
        rh(ws, header_row, 20)
        ws.merge_cells(f"A{header_row}:E{header_row}")

        ph = ws[f"A{header_row}"]
        ph.value     = f"  {fase['label']}"
        ph.font      = fnt(bold=True, color=BRANCO, size=10)
        ph.fill      = fill(TEAL)
        ph.alignment = aln("left")
        ph.border    = brd(TEAL)

        # F — soma das subtarefas
        fh_cell = ws.cell(row=header_row, column=6,
                          value=f"=SUM(F{first_sub}:F{last_sub})")
        fh_cell.font      = fnt(bold=True, color=BRANCO, size=10)
        fh_cell.fill      = fill(TEAL)
        fh_cell.alignment = aln("center")
        fh_cell.border    = brd(TEAL)

        # G — total da fase como texto
        gh_cell = ws.cell(row=header_row, column=7)
        gh_cell.value     = f"{sum(t[5] for t in fase['subtasks'])}h"
        gh_cell.font      = fnt(bold=True, color=BRANCO, size=9)
        gh_cell.fill      = fill(TEAL)
        gh_cell.alignment = aln("center")
        gh_cell.border    = brd(TEAL)

        # H — % da fase
        hh_cell = ws.cell(row=header_row, column=8,
                          value=f"=F{header_row}/$F${TOTAL_ROW}")
        hh_cell.number_format = "0%"
        hh_cell.font      = fnt(bold=True, color=BRANCO, size=10)
        hh_cell.fill      = fill(TEAL)
        hh_cell.alignment = aln("center")
        hh_cell.border    = brd(TEAL)

        # I — vazio
        ih_cell = ws.cell(row=header_row, column=9)
        ih_cell.fill   = fill(TEAL)
        ih_cell.border = brd(TEAL)

        # — Subtarefas ————————————————————————————————————————————————————————
        for sub_idx, sub in enumerate(fase["subtasks"]):
            sub_id, ativ, entregavel, resp_e, resp_c, horas, semana = sub
            r      = first_sub + sub_idx
            row_bg = BRANCO if sub_idx % 2 == 0 else CINZA_CLARO
            is_upload = "1.6" in sub_id

            rh(ws, r, 28)

            vals = [sub_id, ativ, entregavel, resp_e, resp_c,
                    horas, semana,
                    f"=F{r}/$F${TOTAL_ROW}",
                    "🕐 Pendente"]

            for ci, val in enumerate(vals, start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.fill      = fill(row_bg)
                c.alignment = aln(
                    "center" if ci in (1, 4, 5, 6, 7, 8, 9) else "left",
                    wrap=ci in (2, 3)
                )

                # fonte: negrito apenas no id e na coluna %
                is_bold = ci in (1, 8) or is_upload
                c.font = fnt(bold=is_bold,
                             color=TEAL_DARK if ci == 8 else TEXTO_CORPO,
                             size=9,
                             italic=is_upload and ci == 2)

                # borda com acento teal na col A
                c.border = brd(left_color=TEAL_DARK) if ci == 1 else brd()

                if ci == 8:
                    c.number_format = "0.0%"

            # Dropdown status
            dv = DataValidation(type="list", formula1=status_opts,
                                allow_blank=True, showErrorMessage=False)
            ws.add_data_validation(dv)
            dv.add(ws.cell(row=r, column=9))

    # ── Linha TOTAL ───────────────────────────────────────────────────────────
    rh(ws, TOTAL_ROW, 20)
    leaf_ref = "+".join(f"F{r}" for r in leaf_rows)

    ws.merge_cells(f"A{TOTAL_ROW}:E{TOTAL_ROW}")
    tl = ws[f"A{TOTAL_ROW}"]
    tl.value     = "Total do Projeto"
    tl.font      = fnt(bold=True, color=BRANCO, size=11)
    tl.fill      = fill(NAVY)
    tl.alignment = aln("center")
    tl.border    = brd(NAVY)

    tf = ws.cell(row=TOTAL_ROW, column=6, value=f"={leaf_ref}")
    tf.font      = fnt(bold=True, color=BRANCO, size=11)
    tf.fill      = fill(NAVY)
    tf.alignment = aln("center")
    tf.border    = brd(NAVY)

    ws.cell(row=TOTAL_ROW, column=7).fill = fill(NAVY)

    tp = ws.cell(row=TOTAL_ROW, column=8, value=1)
    tp.number_format = "0%"
    tp.font      = fnt(bold=True, color=BRANCO, size=11)
    tp.fill      = fill(NAVY)
    tp.alignment = aln("center")
    tp.border    = brd(NAVY)

    ws.cell(row=TOTAL_ROW, column=9).fill = fill(NAVY)

    for col in (7, 9):
        ws.cell(row=TOTAL_ROW, column=col).border = brd(NAVY)

    # ── Notas de rodapé ───────────────────────────────────────────────────────
    rh(ws, TOTAL_ROW + 2, 14)
    ws.merge_cells(f"A{TOTAL_ROW+2}:I{TOTAL_ROW+2}")
    n = ws[f"A{TOTAL_ROW+2}"]
    n.value     = ("★  Etapa 1.6: Upload das documentações na plataforma Efcaz "
                   "é responsabilidade exclusiva da EFCAZ — o cliente não executa esta etapa.")
    n.font      = fnt(italic=True, color=TEXTO_SUB, size=8)
    n.alignment = aln("left")

    rh(ws, TOTAL_ROW + 3, 14)
    ws.merge_cells(f"A{TOTAL_ROW+3}:I{TOTAL_ROW+3}")
    n2 = ws[f"A{TOTAL_ROW+3}"]
    n2.value = (f"ℹ  Para adicionar subetapas: insira linhas dentro do bloco de cada fase "
                f"e replique a fórmula =F{{linha}}/$F${TOTAL_ROW} na coluna H.")
    n2.font      = fnt(italic=True, color=TEXTO_SUB, size=8)
    n2.alignment = aln("left")

    # ── Formatação condicional por status — coluna I e linha inteira ──────────
    # Intervalo cobre todas as linhas de subtarefa (ignora headers de fase)
    subtask_ranges = []
    for _, fs, ls, _ in phase_meta:
        subtask_ranges.append(f"A{fs}:I{ls}")
    cf_range = " ".join(subtask_ranges)

    # Verde — Concluído
    ws.conditional_formatting.add(
        cf_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Concluído",$I5))'],
            fill=PatternFill("solid", fgColor="C6EFCE"),
            font=Font(name="Arial", color="1A5C2A", size=9),
        ),
    )
    # Amarelo — Em Andamento
    ws.conditional_formatting.add(
        cf_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Em Andamento",$I5))'],
            fill=PatternFill("solid", fgColor="FFF2CC"),
            font=Font(name="Arial", color="7B4F00", size=9),
        ),
    )
    # Cinza — Pausado
    ws.conditional_formatting.add(
        cf_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Pausado",$I5))'],
            fill=PatternFill("solid", fgColor="F2F2F2"),
            font=Font(name="Arial", color="595959", size=9),
        ),
    )

    return TOTAL_ROW, phase_meta


# ══════════════════════════════════════════════════════════════════════════════
#  PROGRESSO no Resumo Executivo (referencia Cronograma)
# ══════════════════════════════════════════════════════════════════════════════
def add_progresso_resumo(wb, phase_meta):
    ws   = wb["Resumo Executivo"]
    cron = wb["Cronograma"]

    # Monta fórmula COUNTIF para contar "Concluído" nas células de status
    # Usa referência cruzada entre abas
    status_ranges = ",".join(
        f"Cronograma!I{fs}:I{ls}"
        for _, fs, ls, _ in phase_meta
    )
    total_tasks = sum(ls - fs + 1 for _, fs, ls, _ in phase_meta)

    formula_concluidas = f"=COUNTIF({status_ranges},\"*Concluído*\")"
    formula_pct        = f"=({formula_concluidas[1:]})/{total_tasks}"

    # ── Bloco de progresso (linhas 19-21) ────────────────────────────────────
    rh(ws, 19, 10)
    rh(ws, 20, 22)
    rh(ws, 21, 14)

    ws.merge_cells("A20:D20")
    lbl = ws["A20"]
    lbl.value     = "Progresso geral do projeto"
    lbl.font      = fnt(bold=True, color=NAVY, size=10)
    lbl.alignment = aln("left")

    # Etapas concluídas / total
    ws.merge_cells("E20:F20")
    cnt = ws["E20"]
    cnt.value       = formula_concluidas
    cnt.font        = fnt(bold=True, color=TEAL_DARK, size=10)
    cnt.alignment   = aln("center")
    cnt.number_format = '0" concluídas"'

    # % de progresso
    ws.merge_cells("G20:H20")
    pct = ws["G20"]
    pct.value         = formula_pct
    pct.font          = fnt(bold=True, color=NAVY, size=14)
    pct.alignment     = aln("center")
    pct.number_format = "0%"
    pct.fill          = fill(CINZA_CLARO)
    pct.border        = brd(left_color=TEAL)

    ws.merge_cells("A21:H21")
    nota = ws["A21"]
    nota.value     = "★  Muda o status de cada etapa na aba Cronograma — este indicador atualiza automaticamente."
    nota.font      = fnt(italic=True, color=TEXTO_SUB, size=8)
    nota.alignment = aln("left")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_resumo(wb)
    total_row, phase_meta = build_cronograma(wb)
    add_progresso_resumo(wb, phase_meta)
    wb.active = wb["Resumo Executivo"]

    wb.save(OUTPUT_PATH)
    print(f"✅  Salvo em:\n    {OUTPUT_PATH}")
    print(f"\n📊  Validação:")
    print(f"    Total: {TOTAL_HORAS}h  {'✅' if TOTAL_HORAS == 64 else '❌ esperado 64h'}")
    for f in FASES:
        fh = sum(t[5] for t in f["subtasks"])
        print(f"    Fase {f['num']}: {fh}h")


if __name__ == "__main__":
    main()
