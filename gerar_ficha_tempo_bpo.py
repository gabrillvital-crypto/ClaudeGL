#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = r"c:\Users\gabriel.evangelista\Documents\ClaudeGL\clientes\Lactalis\ficha_tempo_bpo_geral.xlsx"
LOGO_PATH   = r"c:\Users\gabriel.evangelista\Documents\ClaudeGL\Documentos\logos_extraidas\slide1_img4.png"

# ── Paleta Efcaz ───────────────────────────────────────────────────────────
TEAL        = "0E8FA3"
TEAL_CLARO  = "14B3CC"
TEAL_ESCURO = "0A6A7A"
VERDE       = "27AE60"
LARANJA     = "F39C12"
PRETO_SUV   = "1A2A2A"

EQUIPE = [
    "Thais Jayne Biscaia",
    "Janaina Ventura",
    "Andriela Klai Fernandes",
    "Ricardo Pedroso da Silva",
    "Gabriel Vital",
]

CLIENTES = [
    "Lactalis",
    # adicione mais clientes BPO aqui conforme necessário
]

CATEGORIAS = [
    "Reunião",
    "Análise",
    "Configuração",
    "Documentação",
    "Revisão",
    "Apresentação",
    "Suporte",
    "Treinamento",
]

EQUIPE_CORES = ["0E8FA3", "27AE60", "F39C12", "0A6A7A", "14B3CC"]

# ── Helpers ────────────────────────────────────────────────────────────────
def fill(c):    return PatternFill("solid", fgColor=c)
def fnt(bold=False, color="1F1F1F", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def border():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)
def border_accent(color=TEAL):
    return Border(left=Side(style="medium", color=color),
                  right=Side(style="thin", color="BFBFBF"),
                  top=Side(style="thin", color="BFBFBF"),
                  bottom=Side(style="thin", color="BFBFBF"))

TASK_ROW = 3
MAX_ROW  = 500   # volume maior — ficha geral, vários meses
N_COLS   = 8     # A:H

def section_header(ws, row, text, n_cols=8, h=26, bg=TEAL):
    lc = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{lc}{row}")
    ws[f"A{row}"] = text
    ws[f"A{row}"].fill = fill(bg)
    ws[f"A{row}"].font = fnt(bold=True, color="FFFFFF", size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = h

# ══════════════════════════════════════════════════════════════════════════
# ABA LISTAS (oculta)
# ══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws_l = wb.active
ws_l.title = "Listas"

listas = [("A", "Responsável", EQUIPE),
          ("C", "Cliente",     CLIENTES),
          ("E", "Categoria",   CATEGORIAS)]

for col, title, items in listas:
    ws_l[f"{col}1"] = title
    ws_l[f"{col}1"].font = fnt(bold=True, color="FFFFFF")
    ws_l[f"{col}1"].fill = fill(TEAL_ESCURO)
    for i, v in enumerate(items, 2):
        ws_l.cell(row=i, column=ord(col) - 64, value=v)
    ws_l.column_dimensions[col].width = 35

ws_l.sheet_state = "hidden"

# ══════════════════════════════════════════════════════════════════════════
# ABA REGISTRO
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Registro de Horas")
ws.sheet_view.showGridLines = False

# Colunas
COL_DATA  = 1   # A — Data
COL_MES   = 2   # B — Mês/Ano (fórmula automática)
COL_RESP  = 3   # C — Responsável
COL_CLI   = 4   # D — Cliente
COL_CAT   = 5   # E — Categoria
COL_DESC  = 6   # F — Descrição da atividade
COL_HORAS = 7   # G — Horas
COL_OBS   = 8   # H — Observação

# Linha 1: Título
ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
ws["A1"] = "FICHA DE HORAS  ·  TIME BPO  ·  EFCAZ"
ws["A1"].fill = fill(TEAL_ESCURO)
ws["A1"].font = fnt(bold=True, color="FFFFFF", size=14)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 48

if os.path.exists(LOGO_PATH):
    logo = XLImage(LOGO_PATH)
    logo.width  = 130
    logo.height = 43
    ws.add_image(logo, "A1")

# Linha 2: Cabeçalhos
HEADERS = ["Data", "Mês/Ano", "Responsável", "Cliente", "Categoria",
           "Descrição da Atividade", "Horas", "Observação"]
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.fill = fill(TEAL)
    cell.font = fnt(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()
ws.row_dimensions[2].height = 32

# Linhas de dados
for r in range(TASK_ROW, MAX_ROW + 1):
    row_bg = "FAFAFA" if r % 2 == 0 else "FFFFFF"

    # Data
    dc = ws.cell(row=r, column=COL_DATA)
    dc.number_format = "DD/MM/YYYY"
    dc.fill = fill(row_bg)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    dc.border = border_accent(TEAL)

    # Mês/Ano — fórmula automática a partir da data
    mc = ws.cell(row=r, column=COL_MES,
                 value=f'=IF(A{r}="","",TEXT(A{r},"mmm/aaaa"))')
    mc.fill = fill("EBF7F9" if r % 2 == 0 else "F0FBFC")
    mc.font = fnt(italic=True, color="595959", size=9)
    mc.alignment = Alignment(horizontal="center", vertical="center")
    mc.border = border()

    # Responsável, Cliente, Categoria
    for col in (COL_RESP, COL_CLI, COL_CAT):
        c = ws.cell(row=r, column=col)
        c.fill = fill(row_bg)
        c.font = fnt(size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border()

    # Descrição
    ac = ws.cell(row=r, column=COL_DESC)
    ac.fill = fill(row_bg)
    ac.font = fnt(size=9)
    ac.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ac.border = border()

    # Horas
    hc = ws.cell(row=r, column=COL_HORAS)
    hc.number_format = '0.0"h"'
    hc.fill = fill("EBF7F9" if r % 2 == 0 else "FFFFFF")
    hc.font = fnt(bold=True, color=TEAL_ESCURO, size=10)
    hc.alignment = Alignment(horizontal="center", vertical="center")
    hc.border = border()

    # Observação
    oc = ws.cell(row=r, column=COL_OBS)
    oc.fill = fill(row_bg)
    oc.font = fnt(size=9, italic=True, color="595959")
    oc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    oc.border = border()

    ws.row_dimensions[r].height = 22

# Linha de Total
TOTAL_ROW = MAX_ROW + 1
ws.merge_cells(f"A{TOTAL_ROW}:F{TOTAL_ROW}")
ws[f"A{TOTAL_ROW}"] = "TOTAL GERAL DE HORAS"
ws[f"A{TOTAL_ROW}"].fill = fill(TEAL_ESCURO)
ws[f"A{TOTAL_ROW}"].font = fnt(bold=True, color="FFFFFF", size=11)
ws[f"A{TOTAL_ROW}"].alignment = Alignment(horizontal="right", vertical="center")
ws[f"A{TOTAL_ROW}"].border = border()

tc = ws.cell(row=TOTAL_ROW, column=COL_HORAS,
             value=f"=SUM(G{TASK_ROW}:G{MAX_ROW})")
tc.number_format = '0.0"h"'
tc.fill = fill(TEAL_ESCURO)
tc.font = fnt(bold=True, color="FFFFFF", size=13)
tc.alignment = Alignment(horizontal="center", vertical="center")
tc.border = border()

ws.cell(row=TOTAL_ROW, column=COL_OBS).fill = fill(TEAL_ESCURO)
ws.cell(row=TOTAL_ROW, column=COL_OBS).border = border()

# CF — coluna Horas: escala de cor
ws.conditional_formatting.add(
    f"G{TASK_ROW}:G{MAX_ROW}",
    ColorScaleRule(
        start_type="num", start_value=0, start_color="F8696B",
        mid_type="num",   mid_value=4,   mid_color="FFEB84",
        end_type="num",   end_value=8,   end_color="63BE7B"
    )
)

# CF — Categoria: cores por tipo
CAT_CORES = {
    "Reunião":      ("FF9800", "FFF3E0"),
    "Análise":      ("1565C0", "E3F2FD"),
    "Configuração": ("2E7D32", "E8F5E9"),
    "Documentação": ("6A1B9A", "F3E5F5"),
    "Revisão":      ("F9A825", "FFFDE7"),
    "Apresentação": ("AD1457", "FCE4EC"),
    "Suporte":      ("00838F", "E0F7FA"),
    "Treinamento":  ("4527A0", "EDE7F6"),
}
cat_rng = f"E{TASK_ROW}:E{MAX_ROW}"
for cat_val, (fg, bg) in CAT_CORES.items():
    ws.conditional_formatting.add(cat_rng, FormulaRule(
        formula=[f'$E{TASK_ROW}="{cat_val}"'],
        fill=PatternFill("solid", fgColor=bg),
        font=Font(name="Calibri", bold=True, color=fg, size=9)
    ))

# Dropdowns
dv_resp = DataValidation(type="list",
    formula1=f"Listas!$A$2:$A${1+len(EQUIPE)}",
    allow_blank=True, showDropDown=False,
    showErrorMessage=True, errorTitle="Nome inválido",
    error="Selecione um membro da equipe.",
    showInputMessage=True, promptTitle="Responsável",
    prompt="Selecione o membro da equipe.")
ws.add_data_validation(dv_resp)
dv_resp.sqref = f"C{TASK_ROW}:C{MAX_ROW}"

dv_cli = DataValidation(type="list",
    formula1=f"Listas!$C$2:$C${1+len(CLIENTES)}",
    allow_blank=True, showDropDown=False,
    showErrorMessage=False,   # permite digitar cliente novo
    showInputMessage=True, promptTitle="Cliente",
    prompt="Selecione ou digite o cliente/projeto.")
ws.add_data_validation(dv_cli)
dv_cli.sqref = f"D{TASK_ROW}:D{MAX_ROW}"

dv_cat = DataValidation(type="list",
    formula1=f"Listas!$E$2:$E${1+len(CATEGORIAS)}",
    allow_blank=True, showDropDown=False,
    showErrorMessage=True, errorTitle="Categoria inválida",
    error="Selecione uma categoria da lista.",
    showInputMessage=True, promptTitle="Categoria",
    prompt="Tipo de atividade realizada.")
ws.add_data_validation(dv_cat)
dv_cat.sqref = f"E{TASK_ROW}:E{MAX_ROW}"

dv_data = DataValidation(type="date",
    operator="between",
    formula1="DATE(2025,1,1)", formula2="DATE(2030,12,31)",
    allow_blank=True, showDropDown=False,
    showErrorMessage=True, errorTitle="Data inválida",
    error="Digite a data no formato DD/MM/AAAA.",
    showInputMessage=True, promptTitle="Data",
    prompt="DD/MM/AAAA")
ws.add_data_validation(dv_data)
dv_data.sqref = f"A{TASK_ROW}:A{MAX_ROW}"

# Largura das colunas
col_widths = {1: 13, 2: 11, 3: 24, 4: 18, 5: 15, 6: 44, 7: 9, 8: 28}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = f"A{TASK_ROW}"

# ══════════════════════════════════════════════════════════════════════════
# ABA PAINEL GERAL
# ══════════════════════════════════════════════════════════════════════════
ws_p = wb.create_sheet("Painel Geral", 0)
ws_p.sheet_view.showGridLines = False
N_P = 8

# Logo + Título
if os.path.exists(LOGO_PATH):
    logo_p = XLImage(LOGO_PATH)
    logo_p.width  = 150
    logo_p.height = 50
    ws_p.add_image(logo_p, f"{get_column_letter(N_P - 1)}1")

ws_p.merge_cells(f"A1:{get_column_letter(N_P - 2)}1")
ws_p["A1"] = "PAINEL DE HORAS — TIME BPO EFCAZ"
ws_p["A1"].fill = fill(TEAL_ESCURO)
ws_p["A1"].font = fnt(bold=True, color="FFFFFF", size=18)
ws_p["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_p.row_dimensions[1].height = 58
for c_idx in (N_P - 1, N_P):
    ws_p.cell(row=1, column=c_idx).fill = fill("FFFFFF")

ws_p.merge_cells(f"A2:{get_column_letter(N_P)}2")
ws_p["A2"] = "Gestão de horas trabalhadas — todos os projetos e clientes BPO"
ws_p["A2"].fill = fill(TEAL)
ws_p["A2"].font = fnt(italic=True, color="FFFFFF", size=11)
ws_p["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_p.row_dimensions[2].height = 28
ws_p.row_dimensions[3].height = 10

# ── Total geral ────────────────────────────────────────────────────────────
section_header(ws_p, 4, "TOTAL GERAL", N_P, bg=TEAL_ESCURO)

def kpi(ws, row, col, label, formula, fmt='0.0"h"', span=2):
    lc = ws.cell(row=row, column=col, value=label)
    lc.fill = fill("EBF7F9")
    lc.font = fnt(bold=True, color=PRETO_SUV, size=10)
    lc.alignment = Alignment(horizontal="right", vertical="center")
    lc.border = border()
    end = get_column_letter(col + span)
    ws.merge_cells(f"{get_column_letter(col+1)}{row}:{end}{row}")
    vc = ws.cell(row=row, column=col+1, value=formula)
    vc.number_format = fmt
    vc.fill = fill("FFFFFF")
    vc.font = fnt(bold=True, color=TEAL_ESCURO, size=18)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = border()
    return vc

kpi(ws_p, 5, 1, "Total de Horas:",
    f"=IFERROR(SUM('Registro de Horas'!$G${TASK_ROW}:$G${MAX_ROW}),0)")
kpi(ws_p, 5, 5, "Nº de Sessões:",
    f"=IFERROR(COUNTA('Registro de Horas'!$A${TASK_ROW}:$A${MAX_ROW}),0)",
    fmt="0")
ws_p.row_dimensions[5].height = 38
ws_p.row_dimensions[6].height = 10

# ── Horas por responsável ──────────────────────────────────────────────────
section_header(ws_p, 7, "HORAS POR MEMBRO DA EQUIPE", N_P, bg=TEAL_ESCURO)

for c_idx, h in enumerate(["Membro", "", "Horas", "%", "Barra de Esforço", "", "", ""], 1):
    cell = ws_p.cell(row=8, column=c_idx, value=h)
    cell.fill = fill(TEAL)
    cell.font = fnt(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()
ws_p.row_dimensions[8].height = 22

TOTAL_F = f"MAX(SUM('Registro de Horas'!$G${TASK_ROW}:$G${MAX_ROW}),1)"

for idx, (membro, cor) in enumerate(zip(EQUIPE, EQUIPE_CORES)):
    r = 9 + idx

    ws_p.merge_cells(f"A{r}:B{r}")
    nc = ws_p.cell(row=r, column=1, value=membro)
    nc.font = fnt(bold=True, color=cor, size=10)
    nc.fill = fill("F8F9FA")
    nc.alignment = Alignment(horizontal="left", vertical="center")
    nc.border = border()
    ws_p.cell(row=r, column=2).fill = fill(cor)
    ws_p.cell(row=r, column=2).border = border()

    hc = ws_p.cell(row=r, column=3,
        value=f"=IFERROR(SUMIF('Registro de Horas'!$C${TASK_ROW}:$C${MAX_ROW},"
              f'"{membro}","Registro de Horas"!$G${TASK_ROW}:$G${MAX_ROW}),0)')
    hc.number_format = '0.0"h"'
    hc.font = fnt(bold=True, color=cor, size=12)
    hc.fill = fill("FFFFFF")
    hc.alignment = Alignment(horizontal="center", vertical="center")
    hc.border = border()

    pc = ws_p.cell(row=r, column=4,
        value=f"=IFERROR(C{r}/{TOTAL_F},0)")
    pc.number_format = "0%"
    pc.fill = fill("EBF7F9")
    pc.font = fnt(bold=True, color="595959", size=10)
    pc.alignment = Alignment(horizontal="center", vertical="center")
    pc.border = border()

    ws_p.merge_cells(f"E{r}:{get_column_letter(N_P)}{r}")
    bc = ws_p.cell(row=r, column=5,
        value=f'=IFERROR(REPT("█",MIN(20,ROUND(C{r}/{TOTAL_F}*20,0)))'
              f'&REPT("░",20-MIN(20,ROUND(C{r}/{TOTAL_F}*20,0))),"")')
    bc.font = Font(name="Courier New", bold=True, color=cor, size=11)
    bc.fill = fill("F8F9FA")
    bc.alignment = Alignment(horizontal="left", vertical="center")
    bc.border = border()
    ws_p.row_dimensions[r].height = 24

ws_p.row_dimensions[9 + len(EQUIPE)].height = 10

# ── Horas por cliente ──────────────────────────────────────────────────────
CLI_BASE = 9 + len(EQUIPE) + 1
section_header(ws_p, CLI_BASE, "HORAS POR CLIENTE / PROJETO", N_P, bg=TEAL_ESCURO)

for c_idx, h in enumerate(["Cliente", "", "Horas", "%", "Barra", "", "", ""], 1):
    cell = ws_p.cell(row=CLI_BASE + 1, column=c_idx, value=h)
    cell.fill = fill(TEAL)
    cell.font = fnt(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()
ws_p.row_dimensions[CLI_BASE + 1].height = 22

CLI_CORES = ["0E8FA3", "27AE60", "F39C12", "C0392B", "8E44AD", "2980B9", "16A085", "D35400"]
for idx, (cli, cor) in enumerate(zip(CLIENTES, CLI_CORES)):
    r = CLI_BASE + 2 + idx

    ws_p.merge_cells(f"A{r}:B{r}")
    cc = ws_p.cell(row=r, column=1, value=cli)
    cc.font = fnt(bold=True, color=cor, size=10)
    cc.fill = fill("F8F9FA")
    cc.alignment = Alignment(horizontal="left", vertical="center")
    cc.border = border()
    ws_p.cell(row=r, column=2).fill = fill(cor)
    ws_p.cell(row=r, column=2).border = border()

    ch = ws_p.cell(row=r, column=3,
        value=f"=IFERROR(SUMIF('Registro de Horas'!$D${TASK_ROW}:$D${MAX_ROW},"
              f'"{cli}","Registro de Horas"!$G${TASK_ROW}:$G${MAX_ROW}),0)')
    ch.number_format = '0.0"h"'
    ch.font = fnt(bold=True, color=cor, size=12)
    ch.fill = fill("FFFFFF")
    ch.alignment = Alignment(horizontal="center", vertical="center")
    ch.border = border()

    cp = ws_p.cell(row=r, column=4, value=f"=IFERROR(C{r}/{TOTAL_F},0)")
    cp.number_format = "0%"
    cp.fill = fill("EBF7F9")
    cp.font = fnt(bold=True, color="595959", size=10)
    cp.alignment = Alignment(horizontal="center", vertical="center")
    cp.border = border()

    ws_p.merge_cells(f"E{r}:{get_column_letter(N_P)}{r}")
    cb = ws_p.cell(row=r, column=5,
        value=f'=IFERROR(REPT("█",MIN(20,ROUND(C{r}/{TOTAL_F}*20,0)))'
              f'&REPT("░",20-MIN(20,ROUND(C{r}/{TOTAL_F}*20,0))),"")')
    cb.font = Font(name="Courier New", bold=True, color=cor, size=11)
    cb.fill = fill("F8F9FA")
    cb.alignment = Alignment(horizontal="left", vertical="center")
    cb.border = border()
    ws_p.row_dimensions[r].height = 24

ws_p.row_dimensions[CLI_BASE + 2 + len(CLIENTES)].height = 10

# ── Horas por categoria ────────────────────────────────────────────────────
CAT_BASE = CLI_BASE + 2 + len(CLIENTES) + 1
section_header(ws_p, CAT_BASE, "HORAS POR TIPO DE ATIVIDADE", N_P, bg=TEAL_ESCURO)

for c_idx, h in enumerate(["Categoria", "", "Horas", "%", "Barra", "", "", ""], 1):
    cell = ws_p.cell(row=CAT_BASE + 1, column=c_idx, value=h)
    cell.fill = fill(TEAL)
    cell.font = fnt(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()
ws_p.row_dimensions[CAT_BASE + 1].height = 22

CAT_CORES_P = {
    "Reunião":      ("FF9800", "FFF3E0"),
    "Análise":      ("1565C0", "E3F2FD"),
    "Configuração": ("2E7D32", "E8F5E9"),
    "Documentação": ("6A1B9A", "F3E5F5"),
    "Revisão":      ("F9A825", "FFFDE7"),
    "Apresentação": ("AD1457", "FCE4EC"),
    "Suporte":      ("00838F", "E0F7FA"),
    "Treinamento":  ("4527A0", "EDE7F6"),
}

for idx, cat in enumerate(CATEGORIAS):
    r = CAT_BASE + 2 + idx
    cor_txt, cor_bg = CAT_CORES_P[cat]

    ws_p.merge_cells(f"A{r}:B{r}")
    catc = ws_p.cell(row=r, column=1, value=cat)
    catc.font = fnt(bold=True, color=cor_txt, size=10)
    catc.fill = fill(cor_bg)
    catc.alignment = Alignment(horizontal="left", vertical="center")
    catc.border = border()
    ws_p.cell(row=r, column=2).fill = fill(cor_txt)
    ws_p.cell(row=r, column=2).border = border()

    cath = ws_p.cell(row=r, column=3,
        value=f"=IFERROR(SUMIF('Registro de Horas'!$E${TASK_ROW}:$E${MAX_ROW},"
              f'"{cat}","Registro de Horas"!$G${TASK_ROW}:$G${MAX_ROW}),0)')
    cath.number_format = '0.0"h"'
    cath.font = fnt(bold=True, color=cor_txt, size=12)
    cath.fill = fill("FFFFFF")
    cath.alignment = Alignment(horizontal="center", vertical="center")
    cath.border = border()

    catp = ws_p.cell(row=r, column=4, value=f"=IFERROR(C{r}/{TOTAL_F},0)")
    catp.number_format = "0%"
    catp.fill = fill(cor_bg)
    catp.font = fnt(bold=True, color="595959", size=10)
    catp.alignment = Alignment(horizontal="center", vertical="center")
    catp.border = border()

    ws_p.merge_cells(f"E{r}:{get_column_letter(N_P)}{r}")
    catb = ws_p.cell(row=r, column=5,
        value=f'=IFERROR(REPT("█",MIN(20,ROUND(C{r}/{TOTAL_F}*20,0)))'
              f'&REPT("░",20-MIN(20,ROUND(C{r}/{TOTAL_F}*20,0))),"")')
    catb.font = Font(name="Courier New", bold=True, color=cor_txt, size=11)
    catb.fill = fill(cor_bg)
    catb.alignment = Alignment(horizontal="left", vertical="center")
    catb.border = border()
    ws_p.row_dimensions[r].height = 24

# Nota de rodapé
note_r = CAT_BASE + 2 + len(CATEGORIAS) + 1
ws_p.row_dimensions[note_r - 1].height = 10
ws_p.merge_cells(f"A{note_r}:{get_column_letter(N_P)}{note_r}")
nc = ws_p[f"A{note_r}"]
nc.value = ("Registre as sessões na aba 'Registro de Horas'. "
            "Para adicionar um novo cliente, inclua na aba Listas (coluna C) e crie a linha correspondente no painel.")
nc.font = fnt(italic=True, color="7F7F7F", size=9)
nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
nc.border = border()
ws_p.row_dimensions[note_r].height = 28

# Largura das colunas Painel
painel_widths = {1: 28, 2: 4, 3: 10, 4: 7, 5: 18, 6: 12, 7: 12, 8: 12}
for c, w in painel_widths.items():
    ws_p.column_dimensions[get_column_letter(c)].width = w

# Painel como aba ativa
wb.active = ws_p

# ══════════════════════════════════════════════════════════════════════════
# ABA DASHBOARD POR CLIENTE
# ══════════════════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("Dashboard por Cliente", 0)
ws_d.sheet_view.showGridLines = False
N_D = 8

# Logo + Título
if os.path.exists(LOGO_PATH):
    logo_d = XLImage(LOGO_PATH)
    logo_d.width  = 150
    logo_d.height = 50
    ws_d.add_image(logo_d, f"{get_column_letter(N_D - 1)}1")

ws_d.merge_cells(f"A1:{get_column_letter(N_D - 2)}1")
ws_d["A1"] = "DASHBOARD POR CLIENTE — TIME BPO EFCAZ"
ws_d["A1"].fill = fill(TEAL_ESCURO)
ws_d["A1"].font = fnt(bold=True, color="FFFFFF", size=18)
ws_d["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_d.row_dimensions[1].height = 58
for c_idx in (N_D - 1, N_D):
    ws_d.cell(row=1, column=c_idx).fill = fill("FFFFFF")

ws_d.merge_cells(f"A2:{get_column_letter(N_D)}2")
ws_d["A2"] = "Horas trabalhadas por cliente e por tipo de atividade — atualiza automaticamente ao registrar na aba Registro de Horas"
ws_d["A2"].fill = fill(TEAL)
ws_d["A2"].font = fnt(italic=True, color="FFFFFF", size=10)
ws_d["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_d.row_dimensions[2].height = 24
ws_d.row_dimensions[3].height = 10

def cliente_block(ws, start_row, cliente, cor_header, cor_barra, reg_tab):
    """Renderiza o bloco de um cliente no dashboard."""
    N = 8
    REG = reg_tab  # nome da aba de registro

    # ── Cabeçalho do cliente ──────────────────────────────────────────────
    total_f = (f"=IFERROR(SUMIF('{REG}'!$D${TASK_ROW}:$D${MAX_ROW},"
               f'"{cliente}",\'{REG}\'!$G${TASK_ROW}:$G${MAX_ROW}),0)')
    sessoes_f = (f"=IFERROR(COUNTIF('{REG}'!$D${TASK_ROW}:$D${MAX_ROW},"
                 f'"{cliente}"),0)')

    ws.merge_cells(f"A{start_row}:E{start_row}")
    hdr = ws[f"A{start_row}"]
    hdr.value = f"  {cliente.upper()}"
    hdr.fill = fill(cor_header)
    hdr.font = fnt(bold=True, color="FFFFFF", size=14)
    hdr.alignment = Alignment(horizontal="left", vertical="center")
    hdr.border = border()

    # Total de horas — destaque à direita
    ws.merge_cells(f"F{start_row}:G{start_row}")
    th = ws.cell(row=start_row, column=6, value=total_f)
    th.number_format = '0.0"h  total"'
    th.fill = fill(cor_header)
    th.font = fnt(bold=True, color="FFFFFF", size=14)
    th.alignment = Alignment(horizontal="center", vertical="center")
    th.border = border()

    # Nº sessões
    sc = ws.cell(row=start_row, column=8, value=sessoes_f)
    sc.number_format = '0" sessões"'
    sc.fill = fill(cor_header)
    sc.font = fnt(bold=True, color="FFFFFF", size=11)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    sc.border = border()
    ws.row_dimensions[start_row].height = 36

    # ── Sub-cabeçalho de colunas ──────────────────────────────────────────
    sub_r = start_row + 1
    sub_hdrs = ["Atividade", "", "Horas", "%", "Barra de Esforço", "", "", ""]
    for c_idx, h in enumerate(sub_hdrs, 1):
        cell = ws.cell(row=sub_r, column=c_idx, value=h)
        cell.fill = fill("D0EEF2")
        cell.font = fnt(bold=True, color=cor_header, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border()
    ws.row_dimensions[sub_r].height = 18

    # ── Linhas por categoria ──────────────────────────────────────────────
    TOTAL_CLI_REF = f"MAX(SUMIF('{REG}'!$D${TASK_ROW}:$D${MAX_ROW},\"{cliente}\",'{REG}'!$G${TASK_ROW}:$G${MAX_ROW}),1)"

    for idx, cat in enumerate(CATEGORIAS):
        r = sub_r + 1 + idx
        cor_txt, cor_bg = CAT_CORES_P[cat]
        row_bg = cor_bg if idx % 2 == 0 else "FAFAFA"

        ws.merge_cells(f"A{r}:B{r}")
        catc = ws.cell(row=r, column=1, value=f"  {cat}")
        catc.font = fnt(bold=True, color=cor_txt, size=10)
        catc.fill = fill(row_bg)
        catc.alignment = Alignment(horizontal="left", vertical="center")
        catc.border = border()
        ws.cell(row=r, column=2).fill = fill(cor_txt)
        ws.cell(row=r, column=2).border = border()

        horas_f = (f"=IFERROR(SUMIFS('{REG}'!$G${TASK_ROW}:$G${MAX_ROW},"
                   f"'{REG}'!$D${TASK_ROW}:$D${MAX_ROW},\"{cliente}\","
                   f"'{REG}'!$E${TASK_ROW}:$E${MAX_ROW},\"{cat}\"),0)")
        hc = ws.cell(row=r, column=3, value=horas_f)
        hc.number_format = '0.0"h"'
        hc.font = fnt(bold=True, color=cor_txt, size=11)
        hc.fill = fill("FFFFFF")
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border = border()

        pct_c = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/{TOTAL_CLI_REF},0)")
        pct_c.number_format = "0%"
        pct_c.fill = fill(row_bg)
        pct_c.font = fnt(bold=True, color="595959", size=10)
        pct_c.alignment = Alignment(horizontal="center", vertical="center")
        pct_c.border = border()

        ws.merge_cells(f"E{r}:{get_column_letter(N)}{r}")
        bc = ws.cell(row=r, column=5,
            value=f'=IFERROR(REPT("█",MIN(20,ROUND(C{r}/{TOTAL_CLI_REF}*20,0)))'
                  f'&REPT("░",20-MIN(20,ROUND(C{r}/{TOTAL_CLI_REF}*20,0))),"")')
        bc.font = Font(name="Courier New", bold=True, color=cor_txt, size=11)
        bc.fill = fill(row_bg)
        bc.alignment = Alignment(horizontal="left", vertical="center")
        bc.border = border()
        ws.row_dimensions[r].height = 22

    # Linha de separação inferior
    sep_r = sub_r + 1 + len(CATEGORIAS)
    for c_idx in range(1, N + 1):
        cell = ws.cell(row=sep_r, column=c_idx)
        cell.fill = fill("FFFFFF")
        cell.border = border()
    ws.row_dimensions[sep_r].height = 12

    return sep_r + 1  # próxima linha disponível

# ── Renderizar bloco de cada cliente ──────────────────────────────────────
CLIENTES_DASH = [
    ("Lactalis", TEAL_ESCURO, TEAL),
    # adicione mais clientes aqui conforme o BPO crescer:
    # ("Cliente B", "27AE60", "2ECC71"),
]

cur_row = 4
for cli_nome, cor_h, cor_b in CLIENTES_DASH:
    cur_row = cliente_block(ws_d, cur_row, cli_nome, cor_h, cor_b, "Registro de Horas")
    cur_row += 1  # espaço entre clientes

# Largura das colunas Dashboard
dash_widths = {1: 24, 2: 4, 3: 10, 4: 7, 5: 18, 6: 14, 7: 12, 8: 12}
for c, w in dash_widths.items():
    ws_d.column_dimensions[get_column_letter(c)].width = w

# Dashboard como aba ativa
wb.active = ws_d

# ══════════════════════════════════════════════════════════════════════════
# SALVAR
# ══════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT_PATH)
print(f"OK — Ficha de horas BPO gerada: {OUTPUT_PATH}")
