# -*- coding: utf-8 -*-
"""
Tiers de Clientes Efcaz — Planilha Unificada (Tiers + Cadência) v2 mai/2026
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

OUTPUT = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\Tiers_Clientes_Efcaz.xlsx"
LOGO   = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\logo_efcaz_clean.png"

# ── Paleta Efcaz ───────────────────────────────────────────────────────────────
TEAL     = "0E8FA3"
TEAL_ESC = "0A6A7A"
WHITE    = "FFFFFF"
GRAY_LT  = "F5F6FA"
HDR_BG   = "EBF5FB"
TEAL_XLT = "E0F4F7"

# Cores por tier
A_BG  = "D5F5E3"; A_DK  = "1A5C2E"; A_HDR = "27AE60"
BP_BG = "D6EAF8"; BP_DK = "1A5276"; BP_HDR = "2980B9"
B_BG  = "EAF2F8"; B_DK  = "21618C"; B_HDR = "5DADE2"
C_BG  = "F8F9F9"; C_DK  = "616A6B"; C_HDR = "95A5A6"
NR_BG = "FEF9E7"; NR_DK = "9A7D0A"; NR_HDR = "F39C12"

# ── Helpers ────────────────────────────────────────────────────────────────────
def brd(color="D5D8DC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(h):
    return PatternFill("solid", fgColor=h)

def sc(ws, row, col, value="", bold=False, size=10, color=None,
       bg=None, h="left", v="center", wrap=False, fmt=None, border=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    kw = {"name": "Calibri", "bold": bold, "size": size, "italic": italic}
    if color:
        kw["color"] = color
    c.font = Font(**kw)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border:
        c.border = brd()
    if fmt:
        c.number_format = fmt
    return c

def mrr_cell(ws, row, col, value, bg, color="000000", bold=False, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, size=size, color=color)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = brd()
    c.number_format = '"R$ "#,##0.00'
    return c

def fmt_mrr(v):
    return f'R$ {v:,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")

# ── Dados dos Tiers ────────────────────────────────────────────────────────────
TIERS = [
    {
        "tier": "A", "label": "Tier A — Contas Estratégicas",
        "faixa": "MRR ≥ R$ 5.000/mês",
        "hdr": A_HDR, "bg": A_BG, "dk": A_DK,
        "clients": [
            ("Alpargatas",                                  16666.00, ""),
            ("Supera Promo Serviços Temporários (Nestlé)",  10302.00, "Abril com R$44k pontual; MRR recorrente = R$10.302"),
            ("Zurich Airport Brasil",                        8916.67, "Contrato anual R$107.000 ÷ 12"),
            ("Vinci Airports",                               7750.00, "Contrato anual R$93.000 ÷ 12"),
            ("CSU Digital",                                  7000.00, ""),
            ("Bom Futuro Agrícola",                          6500.00, ""),
            ("ECTX S/A – Eucatex",                           5509.40, ""),
            ("Unimed do Brasil (SOU)",                       5007.20, "Unimed Brasil e Unimed Sou"),
            ("PDV Marketing Promocional (Nestlé)",           1350.36, "Grupo Nestlé — tratado como Tier A junto com Supera Promo"),
        ],
    },
    {
        "tier": "B+", "label": "Tier B+ — Alto Valor",
        "faixa": "MRR R$ 3.500–R$ 4.999/mês (ou estratégico)",
        "hdr": BP_HDR, "bg": BP_BG, "dk": BP_DK,
        "clients": [
            ("Afonso Franca Engenharia",               4800.00, ""),
            ("DATA Engenharia",                        4700.00, ""),
            ("Geistlich Pharma do Brasil",             4634.00, ""),
            ("Associação Adventista Norte Brasileira", 4309.00, ""),
            ("Bunker One Combustíveis e Lubrificantes", 3839.00, ""),
            ("ISG – Instituto Sócrates Guanaes",       3784.99, "7 unidades: AME Pariquera-Açu · Amb. Med. Esp. SJC · CEAPSOL · Hosp. Doenças Tropicais · Hosp. Reg. SJC · Hosp. Reg. Litoral Norte · Hosp. Reg. Jorge Rossmann"),
            ("Dock Brasil Engenharia e Serviços",      3644.00, ""),
            ("Cielo",                                  3450.24, "Contrato anual R$41.402,88 ÷ 12 — mantido em B+ estrategicamente"),
            ("Cebrace",                                2255.87, "Conglomerado de grande porte — elevado estrategicamente para B+"),
        ],
    },
    {
        "tier": "B", "label": "Tier B — Médio Valor",
        "faixa": "MRR R$ 2.000–R$ 3.499/mês",
        "hdr": B_HDR, "bg": B_BG, "dk": B_DK,
        "clients": [
            ("Fundição Alumetaf",                      3300.00, ""),
            ("Soluções Terceirizadas",                 3200.00, ""),
            ("Norskan Offshore",                       2699.00, ""),
            ("Sabará Químicos e Ingredientes",         2504.30, ""),
            ("Premier Pet",                            2500.00, ""),
            ("Tarkett",                                2500.00, ""),
            ("Federação Paulista de Futebol",          2450.00, ""),
            ("Transportes Cavalinho",                  2385.00, ""),
            ("Pacco",                                  2309.00, ""),
            ("Engesp",                                 2250.00, "Contrato anual R$27.000 ÷ 12"),
            ("BRG Suplementos Nutricionais",           2181.92, ""),
            ("Ponsse Latin America",                   2118.62, ""),
        ],
    },
    {
        "tier": "C", "label": "Tier C — Menor Ticket",
        "faixa": "MRR < R$ 2.000/mês",
        "hdr": C_HDR, "bg": C_BG, "dk": C_DK,
        "clients": [
            ("Unimed Campo Grande",         1967.35, ""),
            ("Asso Marítima Navegação",     1899.00, ""),
            ("Amboretto Bombas",            1700.00, ""),
            ("Unimed de Dourados",          1672.00, ""),
            ("Advtec Indústria e Comércio",  700.00, ""),
        ],
    },
]

NON_REC = []

# Cadência — tier_key deve bater exatamente com o valor gravado na col D da aba Tiers
CADENCIA = [
    {"tier": "A – Estratégico",  "tier_key": "Tier A",  "faixa": "≥ R$ 5.000/mês",          "checkin": "Mensal (call / presencial)", "qbr": "Bimensal",   "canal": "WhatsApp direto + E-mail", "sla": "Mesmo dia", "prazo": "Maio/26",  "bg": A_BG,  "dk": A_DK},
    {"tier": "B+ – Alto Valor",  "tier_key": "Tier B+", "faixa": "R$ 3.500–R$ 4.999/mês *", "checkin": "Mensal (call)",              "qbr": "Trimestral", "canal": "E-mail + WhatsApp",        "sla": "24 horas",  "prazo": "Maio/26",  "bg": BP_BG, "dk": BP_DK},
    {"tier": "B – Médio Valor",  "tier_key": "Tier B",  "faixa": "R$ 2.000–R$ 3.499/mês",   "checkin": "Bimensal (call e e-mail)",   "qbr": "Trimestral", "canal": "E-mail",                   "sla": "48 horas",  "prazo": "Junho/26", "bg": B_BG,  "dk": B_DK},
    {"tier": "C – Menor Ticket", "tier_key": "Tier C",  "faixa": "< R$ 2.000/mês",          "checkin": "Trimestral (call e e-mail)", "qbr": "Semestral",  "canal": "E-mail",                   "sla": "72 horas",  "prazo": "Junho/26", "bg": C_BG,  "dk": C_DK},
]

# Faixa de fórmulas — começa em linha 9 (linha 8 = 1º cabeçalho de seção,
# evita auto-referência que gera #REF! nos cabeçalhos)
FR = "$D$9:$D$5000"    # coluna Tier (critério)
CR = "$C$9:$C$5000"    # coluna MRR  (soma)

def f_count(key):
    return f'=COUNTIF({FR},"{key}")'

def f_sum(key):
    return f'=SUMIF({FR},"{key}",{CR})'

# Pré-cálculo Python apenas para o print de validação ao rodar o script
total_clients = sum(len(t["clients"]) for t in TIERS)
total_mrr     = sum(sum(c[1] for c in t["clients"]) for t in TIERS)


# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — TIERS DE CLIENTES
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Tiers de Clientes"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = TEAL

for col, w in zip("ABCDE", [5, 44, 16, 10, 38]):
    ws1.column_dimensions[col].width = w

# ── Linha 1: cabeçalho TEAL com logo Efcaz ────────────────────────────────────
ws1.row_dimensions[1].height = 68
ws1.merge_cells("A1:E1")
c = ws1["A1"]
c.fill      = fill(TEAL_ESC)
c.border    = Border()
c.value     = "Gestão de Carteira  —  Uso Interno"
c.font      = Font(name="Calibri", size=9, italic=True, color="A8D8E0")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
img1 = XLImage(LOGO)
img1.height = 48
img1.width  = int(223 * 48 / 78)
# Posiciona logo à direita: col E row 1
ws1.add_image(img1, "D1")

# ── Linha 2: título ────────────────────────────────────────────────────────────
ws1.row_dimensions[2].height = 30
ws1.merge_cells("A2:E2")
c = ws1["A2"]
c.value = "Classificação de Clientes por Tier — Carteira SRM 2026"
c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
c.fill = fill(TEAL)
c.alignment = Alignment(horizontal="center", vertical="center")

# ── Linha 3: critérios ────────────────────────────────────────────────────────
ws1.row_dimensions[3].height = 16
ws1.merge_cells("A3:E3")
c = ws1["A3"]
c.value = ("Critérios: Tier A ≥ R$5.000/mês  |  Tier B+ R$3.500–R$4.999 (ou estratégico)  |  "
           "Tier B R$2.000–R$3.499  |  Tier C < R$2.000  |  Contratos anuais convertidos em MRR (÷ 12)")
c.font = Font(name="Calibri", size=8, italic=True, color=TEAL_ESC)
c.fill = fill(HDR_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

# ── Linha 4: resumo dos tiers (fórmula — atualiza automaticamente) ────────────
ws1.row_dimensions[4].height = 20
ws1.merge_cells("A4:E4")
c = ws1["A4"]
c.value = (
    '="Tier A: "&COUNTIF($D$8:$D$5000,"Tier A")&" cl'
    '  |  Tier B+: "&COUNTIF($D$8:$D$5000,"Tier B+")&" cl'
    '  |  Tier B: "&COUNTIF($D$8:$D$5000,"Tier B")&" cl'
    '  |  Tier C: "&COUNTIF($D$8:$D$5000,"Tier C")&" cl'
    '  |  Total: "&COUNTIF($D$8:$D$5000,"Tier*")&" clientes"'
)
c.font = Font(name="Calibri", size=9, bold=True, color=TEAL_ESC)
c.fill = fill(HDR_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

# ── Linha 5: MRR Total ────────────────────────────────────────────────────────
ws1.row_dimensions[5].height = 22
ws1.merge_cells("A5:B5")
c = ws1["A5"]
c.value = "MRR Total Carteira SRM"
c.font = Font(name="Calibri", bold=True, size=10, color=TEAL_ESC)
c.fill = fill(HDR_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("C5:E5")
c = ws1["C5"]
c.value = f'=SUMIF({FR},"Tier*",{CR})'   # soma tudo que tem "Tier" na col D
c.font = Font(name="Calibri", bold=True, size=16, color=TEAL)
c.fill = fill(HDR_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
c.number_format = '"R$ "#,##0.00'

# ── Linha 6: espaçador ────────────────────────────────────────────────────────
ws1.row_dimensions[6].height = 6

# ── Linha 7: cabeçalhos ───────────────────────────────────────────────────────
ws1.row_dimensions[7].height = 20
for ci, h in enumerate(["Nº", "Cliente", "MRR / mês", "Tier", "Observação"], 1):
    sc(ws1, 7, ci, h, bold=True, size=10, color=WHITE, bg=TEAL_ESC, h="center")

# ── Seções por tier ───────────────────────────────────────────────────────────
cur = 8
for td in TIERS:
    n        = len(td["clients"])
    mrr_tot  = sum(x[1] for x in td["clients"])

    # Cabeçalho da seção — MRR dinâmico (SUMIF), contagem estática gerada pelo script
    tier_key = f"Tier {td['tier']}"
    n        = len(td["clients"])
    ws1.merge_cells(f"A{cur}:B{cur}")
    sc(ws1, cur, 1, td["label"], bold=True, size=10, color=WHITE, bg=td["hdr"])
    c = ws1.cell(row=cur, column=3, value=f_sum(tier_key))
    c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill      = fill(td["hdr"])
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border    = brd()
    c.number_format = '"R$ "#,##0.00'
    sc(ws1, cur, 4, f"{n} clientes", bold=True, size=9, color=WHITE, bg=td["hdr"], h="center")
    sc(ws1, cur, 5, td["faixa"],     bold=False, size=9, color=WHITE, bg=td["hdr"], h="center")
    ws1.row_dimensions[cur].height = 20
    cur += 1

    # Linhas dos clientes
    for i, (name, mrr, obs) in enumerate(td["clients"], 1):
        row_bg = td["bg"] if i % 2 == 1 else GRAY_LT
        obs_color = "B71C1C" if "cancelamento" in obs.lower() else "666666"
        sc(ws1, cur, 1, i,              h="center", bg=row_bg)
        sc(ws1, cur, 2, name,                       bg=row_bg)
        mrr_cell(ws1, cur, 3, mrr, row_bg)
        sc(ws1, cur, 4, f"Tier {td['tier']}", bold=True, size=9,
           color=td["dk"], bg=row_bg, h="center")
        sc(ws1, cur, 5, obs, bg=row_bg, wrap=True, size=9, color=obs_color)
        lines = max(1, -(-len(obs) // 36)) if obs else 1
        ws1.row_dimensions[cur].height = max(18, lines * 15 + 4)
        cur += 1

# ── Seção não recorrente ──────────────────────────────────────────────────────
ws1.row_dimensions[cur].height = 8  # espaçador
cur += 1

ws1.merge_cells(f"A{cur}:E{cur}")
c = ws1[f"A{cur}"]
c.value = "⚠   NÃO RECORRENTE — verificar manualmente"
c.font  = Font(name="Calibri", bold=True, size=10, color=NR_DK)
c.fill  = fill(NR_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[cur].height = 20
cur += 1

for i, (name, mrr, obs) in enumerate(NON_REC, 1):
    sc(ws1, cur, 1, i,    h="center", bg=NR_BG)
    sc(ws1, cur, 2, name,             bg=NR_BG)
    mrr_cell(ws1, cur, 3, mrr, NR_BG, color=NR_DK)
    sc(ws1, cur, 4, "—",  h="center", bg=NR_BG, color=NR_DK, bold=True)
    sc(ws1, cur, 5, obs,  bg=NR_BG,   size=9, color=NR_DK, wrap=True)
    ws1.row_dimensions[cur].height = 20
    cur += 1

# ── Fundo externo ─────────────────────────────────────────────────────────────
ext = PatternFill("solid", fgColor=TEAL_XLT)
for r in range(1, cur + 50):
    for col in range(6, 40):
        ws1.cell(row=r, column=col).fill = ext
        ws1.cell(row=r, column=col).border = Border()
for r in range(cur, cur + 50):
    for col in range(1, 6):
        ws1.cell(row=r, column=col).fill = ext
        ws1.cell(row=r, column=col).border = Border()

ws1.freeze_panes = "A8"


# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — CADÊNCIA POR TIER
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Cadência por Tier")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = TEAL_ESC

for col, w in zip("ABCDEFGHI", [22, 27, 10, 16, 26, 13, 24, 14, 22]):
    ws2.column_dimensions[col].width = w

# ── Linha 1: cabeçalho TEAL com logo Efcaz ────────────────────────────────────
ws2.row_dimensions[1].height = 68
ws2.merge_cells("A1:I1")
c = ws2["A1"]
c.fill      = fill(TEAL_ESC)
c.border    = Border()
c.value     = "Modelo de Cadência  —  Uso Interno"
c.font      = Font(name="Calibri", size=9, italic=True, color="A8D8E0")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
img2 = XLImage(LOGO)
img2.height = 48
img2.width  = int(223 * 48 / 78)
ws2.add_image(img2, "G1")

# ── Linha 2: título ────────────────────────────────────────────────────────────
ws2.row_dimensions[2].height = 30
ws2.merge_cells("A2:I2")
c = ws2["A2"]
c.value = "Modelo de Cadência por Tier — Carteira SRM Efcaz 2026"
c.font  = Font(name="Calibri", bold=True, size=14, color=WHITE)
c.fill  = fill(TEAL)
c.alignment = Alignment(horizontal="center", vertical="center")

# ── Linha 3: critérios ────────────────────────────────────────────────────────
ws2.row_dimensions[3].height = 16
ws2.merge_cells("A3:I3")
c = ws2["A3"]
c.value = ("Tier A ≥ R$5.000  |  Tier B+ R$3.500–R$4.999 (ou estratégico)  |  "
           "Tier B R$2.000–R$3.499  |  Tier C < R$2.000")
c.font  = Font(name="Calibri", size=8, italic=True, color=TEAL_ESC)
c.fill  = fill(HDR_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

# ── Linha 4: cabeçalhos ───────────────────────────────────────────────────────
ws2.row_dimensions[4].height = 20
hdrs = ["Tier", "Faixa MRR", "Clientes", "MRR Total",
        "Check-in", "QBR", "Canal Principal", "SLA Resposta", "Prazo p/ 1º Contato"]
for ci, h in enumerate(hdrs, 1):
    sc(ws2, 4, ci, h, bold=True, size=10, color=WHITE, bg=TEAL_ESC, h="center")

# ── Linhas de cadência — Clientes e MRR via fórmula cross-sheet ───────────────
T1 = "'Tiers de Clientes'"   # nome da aba 1 para referência cruzada
for i, cad in enumerate(CADENCIA, 5):
    ws2.row_dimensions[i].height = 30
    tk = cad["tier_key"]
    sc(ws2, i, 1, cad["tier"],   bold=True, size=10, color=cad["dk"], bg=cad["bg"])
    sc(ws2, i, 2, cad["faixa"],  size=9,             color=cad["dk"], bg=cad["bg"], h="center")
    # Contagem: cross-sheet COUNTIF
    c3 = ws2.cell(row=i, column=3,
                  value=f'=COUNTIF({T1}!$D$8:$D$5000,"{tk}")')
    c3.font      = Font(name="Calibri", bold=True, size=12, color=cad["dk"])
    c3.fill      = fill(cad["bg"])
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.border    = brd()
    # MRR: cross-sheet SUMIF
    c4 = ws2.cell(row=i, column=4,
                  value=f'=SUMIF({T1}!$D$8:$D$5000,"{tk}",{T1}!$C$8:$C$5000)')
    c4.font      = Font(name="Calibri", bold=True, size=10, color=cad["dk"])
    c4.fill      = fill(cad["bg"])
    c4.alignment = Alignment(horizontal="right", vertical="center")
    c4.border    = brd()
    c4.number_format = '"R$ "#,##0.00'
    sc(ws2, i, 5, cad["checkin"], size=9,  bg=cad["bg"], h="center")
    sc(ws2, i, 6, cad["qbr"],     size=9,  bg=cad["bg"], h="center")
    sc(ws2, i, 7, cad["canal"],   size=9,  bg=cad["bg"], h="center")
    sc(ws2, i, 8, cad["sla"],     size=10, bg=cad["bg"], h="center", bold=True, color=cad["dk"])
    sc(ws2, i, 9, cad["prazo"],   size=11, bg=cad["bg"], h="center", bold=True, color=cad["dk"])

# ── Linha total — fórmulas SUM sobre as linhas 5–8 ────────────────────────────
tot_row = 5 + len(CADENCIA)
ws2.row_dimensions[tot_row].height = 26
ws2.merge_cells(f"A{tot_row}:B{tot_row}")
c = ws2[f"A{tot_row}"]
c.value = "MRR Total Ativo"
c.font  = Font(name="Calibri", bold=True, size=11, color=WHITE)
c.fill  = fill(TEAL_ESC)
c.alignment = Alignment(horizontal="left", vertical="center")
# Total clientes
ct = ws2.cell(row=tot_row, column=3, value="=SUM(C5:C8)")
ct.font      = Font(name="Calibri", bold=True, size=12, color=WHITE)
ct.fill      = fill(TEAL_ESC)
ct.alignment = Alignment(horizontal="center", vertical="center")
ct.border    = brd()
# Total MRR
cm = ws2.cell(row=tot_row, column=4, value="=SUM(D5:D8)")
cm.font      = Font(name="Calibri", bold=True, size=12, color=WHITE)
cm.fill      = fill(TEAL_ESC)
cm.alignment = Alignment(horizontal="right", vertical="center")
cm.border    = brd()
cm.number_format = '"R$ "#,##0.00'
for col in range(5, 10):
    sc(ws2, tot_row, col, "", bg=TEAL_ESC, border=True)

# ── Nota de rodapé ─────────────────────────────────────────────────────────────
note_row = tot_row + 1
ws2.row_dimensions[note_row].height = 16
ws2.merge_cells(f"A{note_row}:I{note_row}")
c = ws2[f"A{note_row}"]
c.value = ("* Tier B+ pode incluir contas classificadas estrategicamente por porte ou potencial, "
           "independente da faixa de MRR  |  Prazo p/ 1º Contato = mês limite para completar "
           "o primeiro ciclo de contato com todos os clientes do tier")
c.font  = Font(name="Calibri", size=8, italic=True, color=TEAL_ESC)
c.fill  = fill(HDR_BG)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Fundo externo ─────────────────────────────────────────────────────────────
ext2 = PatternFill("solid", fgColor=TEAL_XLT)
for r in range(1, note_row + 50):
    for col in range(10, 40):
        ws2.cell(row=r, column=col).fill = ext2
        ws2.cell(row=r, column=col).border = Border()
for r in range(note_row + 1, note_row + 50):
    for col in range(1, 10):
        ws2.cell(row=r, column=col).fill = ext2
        ws2.cell(row=r, column=col).border = Border()

ws2.freeze_panes = "A5"

# ── Salvar ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)

tier_mrr = [sum(c[1] for c in t["clients"]) for t in TIERS]
print(f"Arquivo gerado: {OUTPUT}")
print(f"  Tier A:  {len(TIERS[0]['clients'])} clientes  |  {fmt_mrr(tier_mrr[0])}")
print(f"  Tier B+: {len(TIERS[1]['clients'])} clientes  |  {fmt_mrr(tier_mrr[1])}")
print(f"  Tier B:  {len(TIERS[2]['clients'])} clientes  |  {fmt_mrr(tier_mrr[2])}")
print(f"  Tier C:  {len(TIERS[3]['clients'])} clientes  |  {fmt_mrr(tier_mrr[3])}")
print(f"  TOTAL:   {total_clients} clientes  |  {fmt_mrr(total_mrr)}")
