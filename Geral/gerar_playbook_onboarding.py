# -*- coding: utf-8 -*-
"""
Playbook de Onboarding Efcaz — Gerador de Planilha (v3)
Estrutura: 1º ao 7º Encontro (Boas-vindas = 1º Encontro oficial)
Coluna A: célula de data mesclada por fase (automação semanal a partir do 1º encontro)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image as XLImage

OUTPUT = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\Playbook_Onboarding_Efcaz.xlsx"

# Paleta Efcaz
TEAL       = "0E8FA3"
TEAL_ESC   = "0A6A7A"
WHITE      = "FFFFFF"
GRAY_LT    = "F5F6FA"
GRAY_BOX   = "D5D8DC"
GREEN_LT   = "D5F5E3"
GREEN_DK   = "1E8449"
ORANGE_LT  = "FDEBD0"
ORANGE_DK  = "B7770D"
HEADER_BG  = "EBF5FB"
TEAL_XLT   = "E0F4F7"
TEAL_LT    = "C2E9EF"
GRAY_NA_LT = "EBEBEB"
GRAY_NA_DK = "95A5A6"

# Layout de linhas
ROW_LOGO    = 1
ROW_TITLE   = 2
ROW_INFO1   = 3
ROW_INFO2   = 4
ROW_PROG    = 5
ROW_SPACER  = 6
ROW_HEADERS = 7
ROW_START   = 8

PHASES = [
    {
        "name": "1º Encontro  —  Boas-vindas",
        "encontro": "1º Enc.",
        "activities": [
            ("Licenças contratadas e objetivos do onboarding apresentados ao cliente", "CS Specialist"),
            ("Cronograma proposto apresentado e validado com o cliente", "CS Specialist"),
            ("Indicadores de sucesso mapeados (processo atual, tempo de homologação, auditorias, riscos, KPIs esperados)", "CS Specialist"),
            ("Análise inicial realizada (migração de fornecedores, integração com outros sistemas, áreas usuárias, categorização)", "CS Specialist"),
            ("Nome do ambiente e logomarca validados com o cliente", "CS Specialist"),
            ("Ponto focal do projeto definido pelo cliente", "CS + Cliente"),
            ("Grupo de WhatsApp criado com o ponto focal", "CS Specialist"),
            ("Planilha de Usuários enviada ao cliente para preenchimento até o próximo encontro", "CS Specialist"),
            ("E-mail de boas-vindas enviado com gravação, cronograma e próximos passos", "CS Specialist"),
        ]
    },
    {
        "name": "2º Encontro  —  Fluxo de Homologação e Análise de Cadastro",
        "encontro": "2º Enc.",
        "activities": [
            ("Fluxo de homologação demonstrado no Módulo Solicitação (cadastro do fornecedor, envio e aprovação de documentos)", "CS Specialist"),
            ("Categorização de fornecedores e grupos de fornecimento definidos com o cliente", "CS + Cliente"),
            ("Documentos obrigatórios por categoria/linha de fornecimento definidos", "CS + Cliente"),
            ("Necessidade de migração de fornecedores confirmada", "CS + Cliente"),
            ("Interesse em integração com outro sistema avaliado", "CS + Cliente"),
            ("Planilha de Documentos e Linhas de Fornecimento enviada ao cliente", "CS Specialist"),
            ("E-mail de resumo enviado com o que foi definido e próximos passos", "CS Specialist"),
        ]
    },
    {
        "name": "3º Encontro  —  Preenchimento de Planilhas",
        "encontro": "3º Enc.",
        "activities": [
            ("Planilha de Usuários revisada e validada junto ao cliente", "CS + Cliente"),
            ("Planilha de Documentos e Linhas de Fornecimento revisada e validada", "CS + Cliente"),
            ("Planilha de Migração de Fornecedores explicada e entregue ao cliente para preenchimento", "CS Specialist"),
            ("Planilha de Migração de Terceiros explicada e entregue ao cliente (se aplicável)", "CS Specialist"),
            ("Modelos de Avaliações Internas (PDF dos questionários) explicados e prazo de entrega acordado", "CS + Cliente"),
            ("Prazo para devolução de todas as planilhas acordado formalmente", "CS + Cliente"),
        ]
    },
    {
        "name": "4º Encontro  —  Configurações da Base, Relatórios e Outros",
        "encontro": "4º Enc.",
        "activities": [
            ("Usuários, perfis e permissões por área criados na plataforma", "CS Specialist"),
            ("Logomarca do cliente configurada na plataforma (via Suporte Efcaz)", "CS Specialist"),
            ("Tipos de documentos e linhas de fornecimento parametrizados conforme planilha validada", "CS Specialist"),
            ("Módulo Buscas Automáticas configurado com as certidões contratadas (FGTS, CND Federal, Estadual, CNDT etc.)", "CS Specialist"),
            ("Notificações e alertas automáticos de vencimento configurados", "CS Specialist"),
            ("Importação de fornecedores realizada na plataforma (se planilha entregue)", "CS Specialist"),
            ("Relatórios gerenciais apresentados ao cliente", "CS Specialist"),
        ]
    },
    {
        "name": "5º Encontro  —  Prática Assistida",
        "encontro": "5º Enc.",
        "activities": [
            ("Fluxo completo do Módulo Solicitação executado pelo cliente com acompanhamento do CS", "CS + Cliente"),
            ("Fluxo de aprovação e reprovação de documentos executado pelo cliente", "CS + Cliente"),
            ("Buscas automáticas executadas e resultado validado com o cliente", "CS + Cliente"),
            ("Cliente orientado sobre como convidar fornecedores a acessar o portal (responsabilidade do cliente e do Suporte Efcaz)", "CS Specialist"),
            ("Dúvidas da equipe operacional sanadas", "CS Specialist"),
        ]
    },
    {
        "name": "6º Encontro  —  Avaliações, Ocorrências e Gestão de Terceiros",
        "encontro": "6º Enc.",
        "activities": [
            ("Módulo Avaliação configurado com os modelos de formulário entregues pelo cliente", "CS Specialist"),
            ("Fluxo de criação, aplicação e leitura de avaliação de fornecedor demonstrado ao cliente", "CS Specialist"),
            ("Módulo Ocorrências apresentado e fluxo de notificação ao fornecedor demonstrado", "CS Specialist"),
            ("Módulo Terceiros configurado e migração realizada (se contratado e planilha entregue)", "CS Specialist"),
            ("Planilhas pendentes verificadas — prazo final de entrega reforçado", "CS + Cliente"),
        ]
    },
    {
        "name": "7º Encontro  —  Check-In Final, Dúvidas e Próximos Passos",
        "encontro": "7º Enc.",
        "activities": [
            ("Critérios de sucesso do onboarding avaliados formalmente com o cliente", "CS Specialist"),
            ("Pendências finais identificadas e plano de ação com prazo definido", "CS + Cliente"),
            ("Go-live confirmado com o cliente", "CS + Cliente"),
            ("Rotina de ongoing apresentada ao cliente (check-ins mensais, QBR trimestral)", "CS Specialist"),
            ("Gravações e materiais de apoio de todos os encontros compilados e enviados", "CS Specialist"),
            ("Certificado de Onboarding emitido (se critérios atingidos)", "CS Specialist"),
        ]
    },
]


def thin_border():
    s = Side(style="thin", color="D5D8DC")
    return Border(left=s, right=s, top=s, bottom=s)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def set_cell(ws, row, col, value="", bold=False, size=10, color=None,
             bg=None, h="left", v="center", wrap=False, fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    kw = {"name": "Calibri", "bold": bold, "size": size}
    if color:
        kw["color"] = color
    c.font = Font(**kw)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border:
        c.border = thin_border()
    if fmt:
        c.number_format = fmt
    return c


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Onboarding"
ws.sheet_view.showGridLines = False

# Larguras: A=Data (mesclada por fase)  B=Atividade  C=Responsável  D=Status  E=Prazo  F=Encontro  G=Observações
for col, w in zip("ABCDEFG", [12, 60, 24, 16, 13, 17, 28]):
    ws.column_dimensions[col].width = w

# ── ROW 1: Logos ───────────────────────────────────────────────────────────────
ws.row_dimensions[ROW_LOGO].height = 70

ws.merge_cells("A1:D1")
c = ws["A1"]
c.value = "[ Logo do Cliente ]"
c.font = Font(name="Calibri", size=9, italic=True, color=WHITE)
c.fill = fill(TEAL_ESC)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = Border(
    left=Side(style="thin", color=TEAL_ESC),
    right=Side(style="dashed", color=WHITE),
    top=Side(style="thin", color=TEAL_ESC),
    bottom=Side(style="thin", color=TEAL_ESC),
)

ws.merge_cells("E1:G1")
c = ws["E1"]
c.fill = fill(TEAL_ESC)
c.border = Border()

img_efcaz = XLImage(r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\logo_efcaz_clean.png")
img_efcaz.height = 55
img_efcaz.width  = int(223 * 55 / 78)
ws.add_image(img_efcaz, "E1")

# ── ROW 2: Título ──────────────────────────────────────────────────────────────
ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = "Playbook de Onboarding  —  Efcaz"
c.font = Font(name="Calibri", bold=True, size=15, color=WHITE)
c.fill = fill(TEAL)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[ROW_TITLE].height = 32

# ── ROW 3: Empresa + CS Specialist ────────────────────────────────────────────
ws.merge_cells("A3:D3")
c = ws["A3"]
c.value = "Empresa:  "
c.font = Font(name="Calibri", size=10, bold=True, color=TEAL_ESC)
c.fill = fill(HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("E3:G3")
c = ws["E3"]
c.value = "CS Specialist:  "
c.font = Font(name="Calibri", size=10, bold=True, color=TEAL_ESC)
c.fill = fill(HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[ROW_INFO1].height = 22

# ── ROW 4: Participantes ───────────────────────────────────────────────────────
ws.merge_cells("A4:G4")
c = ws["A4"]
c.value = "Participantes do cliente:  "
c.font = Font(name="Calibri", size=10, bold=True, color=TEAL_ESC)
c.fill = fill(HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[ROW_INFO2].height = 22

# ── ROW 5: Progresso (fórmula inserida depois) ─────────────────────────────────
ws.row_dimensions[ROW_PROG].height = 28

# ── ROW 6: Espaçador ──────────────────────────────────────────────────────────
ws.row_dimensions[ROW_SPACER].height = 6

# ── ROW 7: Cabeçalhos ─────────────────────────────────────────────────────────
for ci, h in enumerate(["Data", "Atividade", "Responsável", "Status", "Prazo", "Encontro", "Observações"], 1):
    set_cell(ws, ROW_HEADERS, ci, h, bold=True, size=10, color=WHITE, bg=TEAL_ESC, h="center")
ws.row_dimensions[ROW_HEADERS].height = 20

# ── FASES ──────────────────────────────────────────────────────────────────────
current_row = ROW_START
phase_meta  = []

for ph_idx, phase in enumerate(PHASES):
    n     = len(phase["activities"])
    hrow  = current_row
    f_act = current_row + 1
    l_act = current_row + n

    phase_meta.append({"hrow": hrow, "f_act": f_act, "l_act": l_act, "n": n})

    # ── Cabeçalho da fase (cols B-G) ──────────────────────────────────────────
    ws.merge_cells(f"B{hrow}:C{hrow}")
    set_cell(ws, hrow, 2, phase["name"], bold=True, size=11, color=WHITE, bg=TEAL)

    c = ws.cell(row=hrow, column=4,
                value=f'=IFERROR(COUNTIF(D{f_act}:D{l_act},"Concluído")/({n}-COUNTIF(D{f_act}:D{l_act},"N/A")),0)')
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    c.border = thin_border()

    ws.merge_cells(f"E{hrow}:F{hrow}")
    set_cell(ws, hrow, 5, phase["encontro"], bold=False, size=9, color=WHITE, bg=TEAL, h="center")

    set_cell(ws, hrow, 7, "", bg=TEAL)
    ws.row_dimensions[hrow].height = 22
    current_row += 1

    # ── Linhas de atividade (cols B-G; coluna A será coberta pela mescla) ─────
    for ai, (act, resp) in enumerate(phase["activities"], 1):
        row_bg = GRAY_LT if ai % 2 == 0 else WHITE
        set_cell(ws, current_row, 2, act, wrap=True, bg=row_bg)
        set_cell(ws, current_row, 3, resp, h="center", bg=row_bg)
        set_cell(ws, current_row, 4, "Pendente", h="center", bg=row_bg)
        set_cell(ws, current_row, 5, "", h="center", bg=row_bg, fmt="DD/MM/YYYY")
        set_cell(ws, current_row, 6, phase["encontro"], h="center", bg=row_bg, size=9, color="7F8C8D")
        set_cell(ws, current_row, 7, "", bg=row_bg)
        lines = max(1, -(-len(act) // 54))
        ws.row_dimensions[current_row].height = max(18, lines * 16 + 4)
        current_row += 1

    # ── Coluna A: célula mesclada de data para toda a fase ────────────────────
    # Preenche todas as células do intervalo com a cor antes de mesclar
    for r in range(hrow, l_act + 1):
        cell = ws.cell(row=r, column=1)
        cell.fill = fill(TEAL_ESC)
        cell.border = Border(
            left=Side(style="thin", color=GRAY_BOX),
            right=Side(style="thin", color=GRAY_BOX),
            top=Side(style="thin", color=GRAY_BOX),
            bottom=Side(style="thin", color=GRAY_BOX),
        )

    ws.merge_cells(f"A{hrow}:A{l_act}")
    tl = ws.cell(row=hrow, column=1)
    tl.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    tl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    tl.number_format = "DD/MM/YYYY"

    if ph_idx == 0:
        # Primeiro encontro: usuário preenche manualmente
        tl.value = None
    else:
        # Demais: fórmula automática (+7 dias por encontro) baseada no 1º
        # Para alterar uma data individualmente: apague a fórmula e digite a data
        phase1_row = phase_meta[0]["hrow"]
        offset = ph_idx * 7
        tl.value = f'=IF($A${phase1_row}="","",$A${phase1_row}+{offset})'

# ── Progresso geral ────────────────────────────────────────────────────────────
total_act      = sum(p["n"] for p in phase_meta)
parts_conc     = "+".join(f'COUNTIF(D{p["f_act"]}:D{p["l_act"]},"Concluído")' for p in phase_meta)
parts_na       = "+".join(f'COUNTIF(D{p["f_act"]}:D{p["l_act"]},"N/A")' for p in phase_meta)
overall_formula = f'=IFERROR(({parts_conc})/({total_act}-({parts_na})),0)'

ws.merge_cells("A5:B5")
c = ws["A5"]
c.value = "Progresso Geral do Onboarding"
c.font = Font(name="Calibri", bold=True, size=10, color=TEAL_ESC)
c.fill = fill(HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("C5:G5")
c = ws["C5"]
c.value = overall_formula
c.font = Font(name="Calibri", bold=True, size=18, color=TEAL)
c.fill = fill(HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
c.number_format = "0%"

# ── Data Validation ───────────────────────────────────────────────────────────
dv = DataValidation(
    type="list",
    formula1='"Pendente,Em andamento,Concluído,N/A"',
    allow_blank=False,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="Status inválido",
    error="Escolha: Pendente, Em andamento ou Concluído"
)
ws.add_data_validation(dv)
for p in phase_meta:
    dv.add(f'D{p["f_act"]}:D{p["l_act"]}')

# ── Formatação condicional (cols B-G para não interferir na espinha de datas) ─
f_row    = phase_meta[0]["f_act"]
l_row    = phase_meta[-1]["l_act"]
cf_range = f"B{f_row}:G{l_row}"

ws.conditional_formatting.add(
    cf_range,
    FormulaRule(
        formula=[f'=$D{f_row}="Concluído"'],
        fill=PatternFill("solid", fgColor=GREEN_LT),
        font=Font(name="Calibri", size=10, color=GREEN_DK)
    )
)
ws.conditional_formatting.add(
    cf_range,
    FormulaRule(
        formula=[f'=$D{f_row}="Em andamento"'],
        fill=PatternFill("solid", fgColor=ORANGE_LT),
        font=Font(name="Calibri", size=10, color=ORANGE_DK)
    )
)
ws.conditional_formatting.add(
    cf_range,
    FormulaRule(
        formula=[f'=$D{f_row}="N/A"'],
        fill=PatternFill("solid", fgColor=GRAY_NA_LT),
        font=Font(name="Calibri", size=10, color=GRAY_NA_DK, italic=True)
    )
)

# ── Fundo externo ─────────────────────────────────────────────────────────────
last_row = phase_meta[-1]["l_act"]
ext_fill = PatternFill("solid", fgColor=TEAL_XLT)

for r in range(1, last_row + 60):
    for col in range(8, 40):
        c = ws.cell(row=r, column=col)
        c.fill = ext_fill
        c.border = Border()

for r in range(last_row + 1, last_row + 60):
    for col in range(1, 8):
        c = ws.cell(row=r, column=col)
        c.fill = ext_fill
        c.border = Border()

ws.sheet_properties.tabColor = TEAL
ws.freeze_panes  = None
ws.print_title_rows = "1:7"

wb.save(OUTPUT)
print(f"Arquivo gerado: {OUTPUT}")
print(f"  Fases: {len(PHASES)} | Atividades: {total_act}")
