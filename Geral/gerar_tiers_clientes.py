# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

OUTPUT   = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\Tiers_Clientes_Efcaz.xlsx"
LOGO_EFC = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\logo_efcaz_clean.png"

TEAL     = "0E8FA3"
TEAL_ESC = "0A6A7A"
TEAL_XLT = "E0F4F7"
WHITE    = "FFFFFF"
GRAY_BOX = "D5D8DC"
GRAY_LT  = "F5F6FA"

GREEN_LT = "D5F5E3"
GREEN_DK = "1E8449"
GREEN_MD = "A9DFBF"

AMBER_LT = "FEF9E7"
AMBER_DK = "B7770D"
AMBER_MD = "F9E79F"

RED_LT   = "FDEDEC"
RED_DK   = "C0392B"
RED_MD   = "F5B7B1"

CHURN_DK  = "7B1818"
CHURN_LT  = "FDECEC"
VERIFY_DK = "A04000"
VERIFY_LT = "FDEBD0"

def fill(h): return PatternFill("solid", fgColor=h)
def thin(c=GRAY_BOX):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def side_border(left=None, right=None, top=None, bottom=None):
    return Border(
        left   = Side(style="thin", color=left)   if left   else Side(style=None),
        right  = Side(style="thin", color=right)  if right  else Side(style=None),
        top    = Side(style="thin", color=top)    if top    else Side(style=None),
        bottom = Side(style="thin", color=bottom) if bottom else Side(style=None),
    )

def cl(ws, row, col, val="", bold=False, size=10, color=None, bg=None,
       h="left", v="center", wrap=False, border=True, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    kw = {"name": "Calibri", "bold": bold, "size": size, "italic": italic}
    if color: kw["color"] = color
    c.font = Font(**kw)
    if bg: c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border: c.border = thin()
    if num_fmt: c.number_format = num_fmt
    return c

# ── DADOS (atualizado via CustomerX 11/05/2026) ──────────────────────
# Obs: "⚠ Sem contrato no CX" = cliente ativo, contrato ainda não registrado no CustomerX
dados = {
    "A": [
        ("Zurich Airport Brasil",                        8929.20, "MRR atualizado via CustomerX"),
        ("CSU Digital",                                  7000.00, "⚠ Sem contrato registrado no CustomerX"),
        ("Bom Futuro Agrícola",                          6500.00, ""),
        ("ECTX S/A — Eucatex",                           5509.40, ""),
        ("Norskan Offshore (DOF)",                       5398.00, "Promovido Tier B→A — valor atualizado via CustomerX"),
        ("Soluções Terceirizadas",                       5200.00, "Promovido Tier B→A — valor atualizado via CustomerX"),
        ("Unimed do Brasil",                             5007.20, ""),
    ],
    "B": [
        ("Afonso Franca Engenharia",                     4800.00, ""),
        ("DATA Engenharia",                              4700.00, "⚠ Sem contrato registrado no CustomerX"),
        ("Geistlich Pharma do Brasil",                   4634.00, ""),
        ("Vinci Airports",                               4500.00, "⚠ Sem contrato registrado no CustomerX"),
        ("Hospital Adventista de Belém",                 4309.00, ""),
        ("Bunker One Combustíveis e Lubrificantes",      3839.00, ""),
        ("ISG — Instituto Sócrates Guanaes",             3784.99, "7 unidades: AME Pariquera-Açu · Amb. Med. Esp. SJC · CEAPSOL · Hosp. Doenças Tropicais · Hosp. Reg. São José dos Campos · Hosp. Reg. Litoral Norte · Hosp. Reg. Jorge Rossmann"),
        ("Dock Brasil Engenharia e Serviços",            3644.00, ""),
        ("Cielo",                                        3450.24, "Contrato anual R$41.402,88 ÷ 12"),
        ("Sabarã Químicos e Ingredientes",               2504.30, ""),
        ("Agência Work On (Grupo Nestlé)",               2500.00, ""),
        ("Tarkett",                                      2500.00, ""),
        ("Premier Pet",                                  2500.00, "⚠ Sem contrato registrado no CustomerX"),
        ("Federação Paulista de Futebol",                2450.00, "Projeto encerrado — sem recorrência confirmada"),
        ("Transportes Cavalinho",                        2385.00, "⚠ Sem contrato registrado no CustomerX"),
        ("Pacco / Paccoby",                              2309.00, "⚠ Sem contrato registrado no CustomerX"),
        ("Engesp",                                       2270.00, ""),
        ("Cebrace",                                      2255.87, ""),
        ("BRG Suplementos Nutricionais",                 2181.92, ""),
        ("Ponsse Latin America",                         2118.62, ""),
        ("Killing SA Tintas e Adesivos",                 2000.00, ""),
    ],
    "C": [
        ("Unimed Campo Grande",                          1967.35, ""),
        ("Asso Marítima Navegação",                      1899.00, ""),
        ("Amboretto Bombas",                             1700.00, ""),
        ("Unimed de Dourados",                           1672.00, ""),
        ("Alumetaf — Soluções em Ferro Fundido",        1650.00, "Reclassificado Tier B→C — CustomerX: R$1.650"),
        ("Advtec Indústria e Comércio",                   700.00, ""),
    ],
}

# Clientes ativos sem cadastro confirmado no CustomerX — verificar status
a_verificar = [
    ("Alpargatas",                        16666.00, "CustomerX: Sem contrato — confirmar situação"),
    ("Supera Promo Serviços Temporários", 10302.00, "Não localizado no CustomerX"),
    ("PDV Marketing Promocional",          1350.36, "Não localizado no CustomerX"),
]

# Churn confirmado
churn = [
    ("SENAI",  799.50, "Cancelado em 30/04/2026"),
    ("SESI",   799.50, "Cancelado em 30/04/2026"),
]

especiais = [
    ("Banco Honda S/A", 8400.00, "Cobrança única em janeiro (SRM.SERV). Não é recorrente."),
]

tier_cfg = {
    "A": {"lt": GREEN_LT, "dk": GREEN_DK, "md": GREEN_MD, "label": "Tier A — Contas Estratégicas  (MRR ≥ R$ 5.000)"},
    "B": {"lt": AMBER_LT, "dk": AMBER_DK, "md": AMBER_MD, "label": "Tier B — Contas de Médio Valor  (MRR R$ 2.000 – R$ 4.999)"},
    "C": {"lt": RED_LT,   "dk": RED_DK,   "md": RED_MD,   "label": "Tier C — Contas de Menor Ticket  (MRR < R$ 2.000)"},
}

# ── WORKBOOK ──────────────────────────────────────────────────────────
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "Tiers de Clientes"
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = TEAL

for col, w in zip("ABCDE", [5, 46, 16, 10, 42]):
    ws.column_dimensions[get_column_letter(col if isinstance(col, int) else ord(col) - 64)].width = w

# ── CABEÇALHO ─────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 65
ws.merge_cells("A1:B1")
ws["A1"].fill = fill(TEAL_ESC)
ws["A1"].border = thin(TEAL_ESC)
ws.merge_cells("C1:E1")
ws["C1"].fill = fill(WHITE)
ws["C1"].border = thin(GRAY_BOX)
img = XLImage(LOGO_EFC)
img.height = 52
img.width  = int(223 * 52 / 78)
ws.add_image(img, "C1")

ws.row_dimensions[2].height = 28
ws.merge_cells("A2:E2")
t = ws["A2"]
t.value = "Classificação de Clientes por Tier — Carteira SRM 2026"
t.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
t.fill = fill(TEAL)
t.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[3].height = 16
ws.merge_cells("A3:E3")
s = ws["A3"]
s.value = "Critérios: Tier A ≥ R$5.000/mês | Tier B R$2.000–R$4.999 | Tier C < R$2.000 | Contratos anuais convertidos em MRR (÷ 12)"
s.font = Font(name="Calibri", size=9, italic=True, color="555555")
s.fill = fill(TEAL_XLT)
s.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ── RESUMO (cards por tier) ────────────────────────────────────────────
ws.row_dimensions[4].height = 8  # espaço

mrr_A = sum(v for _, v, _ in dados["A"])
mrr_B = sum(v for _, v, _ in dados["B"])
mrr_C = sum(v for _, v, _ in dados["C"])

card_data = [
    ("A", GREEN_DK, GREEN_LT, len(dados["A"]), mrr_A),
    ("B", AMBER_DK, AMBER_LT, len(dados["B"]), mrr_B),
    ("C", RED_DK,   RED_LT,   len(dados["C"]), mrr_C),
]

ws.row_dimensions[5].height = 20
ws.row_dimensions[6].height = 26
ws.row_dimensions[7].height = 20

for i, (tier, dk, lt, n, mrr) in enumerate(card_data):
    col = i + 2  # B=2, C=3, D=4
    badge = ws.cell(row=5, column=col, value=f"TIER {tier}")
    badge.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    badge.fill = fill(dk)
    badge.alignment = Alignment(horizontal="center", vertical="center")
    badge.border = thin(dk)

    count_val = ws.cell(row=6, column=col, value=f"{n} clientes")
    count_val.font = Font(name="Calibri", bold=True, size=12, color=dk)
    count_val.fill = fill(lt)
    count_val.alignment = Alignment(horizontal="center", vertical="center")
    count_val.border = thin(dk)

    mrr_val = ws.cell(row=7, column=col, value=mrr)
    mrr_val.font = Font(name="Calibri", size=9, bold=True, color=dk)
    mrr_val.fill = fill(lt)
    mrr_val.alignment = Alignment(horizontal="center", vertical="center")
    mrr_val.number_format = 'R$ #,##0.00'
    mrr_val.border = thin(dk)

# label MRR na coluna E linha 7
ws.cell(row=5, column=1).fill = fill(GRAY_LT)
ws.cell(row=5, column=5).fill = fill(GRAY_LT)
ws.cell(row=6, column=1).fill = fill(GRAY_LT)
mrr_lbl = ws.cell(row=7, column=5, value="MRR combinado")
mrr_lbl.font = Font(name="Calibri", size=8, italic=True, color="888888")
mrr_lbl.alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row=7, column=1).fill = fill(GRAY_LT)

ws.row_dimensions[8].height = 8  # espaço

# ── CABEÇALHOS DA TABELA ──────────────────────────────────────────────
ws.row_dimensions[9].height = 22
for col, label in zip(range(1, 6), ["Nº", "Cliente", "MRR/mês", "Tier", "Observação"]):
    c = ws.cell(row=9, column=col, value=label)
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = fill(TEAL_ESC)
    c.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
    c.border = thin(TEAL_ESC)

# ── LINHAS DE DADOS ────────────────────────────────────────────────────
row = 10
for tier in ["A", "B", "C"]:
    cfg = tier_cfg[tier]

    ws.merge_cells(f"A{row}:E{row}")
    hc = ws.cell(row=row, column=1, value=cfg["label"])
    hc.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    hc.fill = fill(cfg["dk"])
    hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hc.border = thin(cfg["dk"])
    ws.row_dimensions[row].height = 20
    row += 1

    for i, (nome, mrr, obs) in enumerate(dados[tier], 1):
        is_isg = nome.startswith("ISG —")
        cl(ws, row, 1, i,    h="center", bg=cfg["lt"], bold=True,  color=cfg["dk"])
        cl(ws, row, 2, nome, bg=cfg["lt"])
        mrr_cell = cl(ws, row, 3, mrr, h="right", bg=cfg["lt"], bold=True, color=cfg["dk"])
        mrr_cell.number_format = 'R$ #,##0.00'
        cl(ws, row, 4, f"Tier {tier}", h="center", bg=cfg["dk"], bold=True, color=WHITE)
        cl(ws, row, 5, obs, bg=cfg["lt"], wrap=True, size=9, italic=True, color="555555")
        ws.row_dimensions[row].height = 45 if is_isg else (28 if obs else 20)
        row += 1

# ── ESPECIAIS ─────────────────────────────────────────────────────────
row += 1
ws.merge_cells(f"A{row}:E{row}")
hc = ws.cell(row=row, column=1, value="⚠  NÃO RECORRENTE — verificar manualmente")
hc.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
hc.fill = fill("7F8C8D")
hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
hc.border = thin("7F8C8D")
ws.row_dimensions[row].height = 20
row += 1

for i, (nome, val, obs) in enumerate(especiais, 1):
    cl(ws, row, 1, i,    h="center", bg="ECF0F1", bold=True, color="7F8C8D")
    cl(ws, row, 2, nome, bg="ECF0F1", color="7F8C8D")
    v = cl(ws, row, 3, val, h="right", bg="ECF0F1", color="7F8C8D")
    v.number_format = 'R$ #,##0.00'
    cl(ws, row, 4, "—", h="center", bg="ECF0F1", color="7F8C8D")
    cl(ws, row, 5, obs, bg="ECF0F1", wrap=True, size=9, italic=True, color="7F8C8D")
    ws.row_dimensions[row].height = 28
    row += 1

# ── A VERIFICAR ───────────────────────────────────────────────────────
row += 1
ws.merge_cells(f"A{row}:E{row}")
hc = ws.cell(row=row, column=1, value="⚠  A VERIFICAR — Status não confirmado no CustomerX")
hc.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
hc.fill = fill(VERIFY_DK)
hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
hc.border = thin(VERIFY_DK)
ws.row_dimensions[row].height = 20
row += 1

for i, (nome, val, obs) in enumerate(a_verificar, 1):
    cl(ws, row, 1, i,    h="center", bg=VERIFY_LT, bold=True, color=VERIFY_DK)
    cl(ws, row, 2, nome, bg=VERIFY_LT, color=VERIFY_DK)
    v = cl(ws, row, 3, val, h="right", bg=VERIFY_LT, color=VERIFY_DK, bold=True)
    v.number_format = 'R$ #,##0.00'
    cl(ws, row, 4, "⚠ Verificar", h="center", bg=VERIFY_DK, bold=True, color=WHITE)
    cl(ws, row, 5, obs, bg=VERIFY_LT, wrap=True, size=9, italic=True, color=VERIFY_DK)
    ws.row_dimensions[row].height = 28
    row += 1

# ── CHURN ─────────────────────────────────────────────────────────────
row += 1
ws.merge_cells(f"A{row}:E{row}")
hc = ws.cell(row=row, column=1, value="✗  CHURN CONFIRMADO — Contratos Cancelados")
hc.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
hc.fill = fill(CHURN_DK)
hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
hc.border = thin(CHURN_DK)
ws.row_dimensions[row].height = 20
row += 1

for i, (nome, val, obs) in enumerate(churn, 1):
    cl(ws, row, 1, i,    h="center", bg=CHURN_LT, bold=True, color=CHURN_DK)
    cl(ws, row, 2, nome, bg=CHURN_LT, color=CHURN_DK)
    v = cl(ws, row, 3, val, h="right", bg=CHURN_LT, color=CHURN_DK)
    v.number_format = 'R$ #,##0.00'
    cl(ws, row, 4, "Churn", h="center", bg=CHURN_DK, bold=True, color=WHITE)
    cl(ws, row, 5, obs, bg=CHURN_LT, wrap=True, size=9, italic=True, color=CHURN_DK)
    ws.row_dimensions[row].height = 28
    row += 1

# ── TOTAL GERAL ────────────────────────────────────────────────────────
row += 1
ws.merge_cells(f"A{row}:B{row}")
tl = ws.cell(row=row, column=1, value="MRR Total Carteira SRM")
tl.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
tl.fill = fill(TEAL_ESC)
tl.alignment = Alignment(horizontal="right", vertical="center")
tl.border = thin(TEAL_ESC)

total_mrr = mrr_A + mrr_B + mrr_C
tv = ws.cell(row=row, column=3, value=total_mrr)
tv.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
tv.fill = fill(TEAL)
tv.alignment = Alignment(horizontal="right", vertical="center")
tv.number_format = 'R$ #,##0.00'
tv.border = thin(TEAL)

ws.cell(row=row, column=4).fill = fill(TEAL_XLT)
ws.cell(row=row, column=5).fill = fill(TEAL_XLT)
ws.row_dimensions[row].height = 24

# Fundo externo
for r in range(1, row + 15):
    for col in range(6, 20):
        ws.cell(row=r, column=col).fill = fill(TEAL_XLT)
for r in range(row + 1, row + 15):
    for col in range(1, 6):
        ws.cell(row=r, column=col).fill = fill(TEAL_XLT)

ws.freeze_panes = "A10"
ws.print_title_rows = "1:9"


# ══════════════════════════════════════════════════════════════════════════════
#  ABA 2 — TODOS OS CLIENTES (ISG consolidado como empresa única)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Todos os Clientes")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = TEAL_ESC

for col_idx, w in zip(range(1, 7), [5, 40, 22, 14, 8, 38]):
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

# ── Cabeçalho ─────────────────────────────────────────────────────────
ws2.row_dimensions[1].height = 65
ws2.merge_cells("A1:C1")
ws2["A1"].fill = fill(TEAL_ESC)
ws2["A1"].border = thin(TEAL_ESC)
ws2.merge_cells("D1:F1")
ws2["D1"].fill = fill(WHITE)
ws2["D1"].border = thin(GRAY_BOX)
img2 = XLImage(LOGO_EFC)
img2.height = 52
img2.width  = int(223 * 52 / 78)
ws2.add_image(img2, "D1")

ws2.row_dimensions[2].height = 28
ws2.merge_cells("A2:F2")
t2 = ws2["A2"]
t2.value = "Todos os Clientes — Carteira SRM 2026"
t2.font  = Font(name="Calibri", bold=True, size=13, color=WHITE)
t2.fill  = fill(TEAL)
t2.alignment = Alignment(horizontal="center", vertical="center")

ws2.row_dimensions[3].height = 16
ws2.merge_cells("A3:F3")
s2 = ws2["A3"]
s2.value = "ISG consolidado como empresa única (Tier B, R$ 3.784,99)  ·  Atualizado via CustomerX em 11/05/2026"
s2.font  = Font(name="Calibri", size=9, italic=True, color="555555")
s2.fill  = fill(TEAL_XLT)
s2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ── Totais consolidados ────────────────────────────────────────────────
todos_data = dados

mrr2_A = sum(v for _, v, _ in todos_data["A"])
mrr2_B = sum(v for _, v, _ in todos_data["B"])
mrr2_C = sum(v for _, v, _ in todos_data["C"])
n2_A   = len(todos_data["A"])
n2_B   = len(todos_data["B"])
n2_C   = len(todos_data["C"])
total2 = n2_A + n2_B + n2_C
mrr2_total = mrr2_A + mrr2_B + mrr2_C

ws2.row_dimensions[4].height = 8

card2_data = [
    ("A", GREEN_DK, GREEN_LT, n2_A, mrr2_A),
    ("B", AMBER_DK, AMBER_LT, n2_B, mrr2_B),
    ("C", RED_DK,   RED_LT,   n2_C, mrr2_C),
]
ws2.row_dimensions[5].height = 20
ws2.row_dimensions[6].height = 26
ws2.row_dimensions[7].height = 20

for i, (tier, dk, lt, n, mrr) in enumerate(card2_data):
    col = i + 2  # B, C, D
    badge2 = ws2.cell(row=5, column=col, value=f"TIER {tier}")
    badge2.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    badge2.fill = fill(dk)
    badge2.alignment = Alignment(horizontal="center", vertical="center")
    badge2.border = thin(dk)

    cnt2 = ws2.cell(row=6, column=col, value=f"{n} empresa{'s' if n != 1 else ''}")
    cnt2.font = Font(name="Calibri", bold=True, size=12, color=dk)
    cnt2.fill = fill(lt)
    cnt2.alignment = Alignment(horizontal="center", vertical="center")
    cnt2.border = thin(dk)

    mrr2 = ws2.cell(row=7, column=col, value=mrr)
    mrr2.font = Font(name="Calibri", size=9, bold=True, color=dk)
    mrr2.fill = fill(lt)
    mrr2.alignment = Alignment(horizontal="center", vertical="center")
    mrr2.number_format = 'R$ #,##0.00'
    mrr2.border = thin(dk)

for r in [5, 6, 7]:
    ws2.cell(row=r, column=1).fill = fill(GRAY_LT)
    ws2.cell(row=r, column=5).fill = fill(GRAY_LT)
    ws2.cell(row=r, column=6).fill = fill(GRAY_LT)

ws2.cell(row=5, column=5, value=f"{total2} empresas").font = Font(name="Calibri", bold=True, size=9, color=TEAL_ESC)
ws2.cell(row=5, column=5).alignment = Alignment(horizontal="center", vertical="center")
ws2.cell(row=5, column=5).fill = fill(TEAL_XLT)
ws2.cell(row=5, column=5).border = thin(TEAL)
ws2.cell(row=6, column=5, value="Total MRR").font = Font(name="Calibri", bold=True, size=11, color=TEAL_ESC)
ws2.cell(row=6, column=5).fill = fill(TEAL_XLT)
ws2.cell(row=6, column=5).alignment = Alignment(horizontal="center", vertical="center")
ws2.cell(row=6, column=5).border = thin(TEAL)
mrr_total_cell = ws2.cell(row=7, column=5, value=mrr2_total)
mrr_total_cell.font = Font(name="Calibri", bold=True, size=10, color=TEAL_ESC)
mrr_total_cell.fill = fill(TEAL_XLT)
mrr_total_cell.alignment = Alignment(horizontal="center", vertical="center")
mrr_total_cell.number_format = 'R$ #,##0.00'
mrr_total_cell.border = thin(TEAL)

ws2.row_dimensions[8].height = 8

# ── Cabeçalhos da tabela ──────────────────────────────────────────────
ws2.row_dimensions[9].height = 22
for col_idx, label in zip(range(1, 7), ["Nº", "Cliente", "Grupo / Holding", "MRR/mês", "Tier", "Observação"]):
    c2 = ws2.cell(row=9, column=col_idx, value=label)
    c2.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c2.fill = fill(TEAL_ESC)
    c2.alignment = Alignment(horizontal="center" if col_idx != 2 else "left",
                             vertical="center")
    c2.border = thin(TEAL_ESC)

# ── Dados ─────────────────────────────────────────────────────────────
def cl2(ws, row, col, val="", bold=False, size=10, color=None, bg=None,
        h="left", v="center", wrap=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    kw = {"name": "Calibri", "bold": bold, "size": size}
    if color: kw["color"] = color
    c.font = Font(**kw)
    if bg: c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    c.border = thin()
    if num_fmt: c.number_format = num_fmt
    return c

# Grupos / Holdings por cliente
GRUPOS = {
    "Unimed do Brasil":              "GRUPO UNIMED",
    "Unimed Campo Grande":           "GRUPO UNIMED",
    "Unimed de Dourados":            "GRUPO UNIMED",
    "ISG — Instituto Sócrates Guanaes": "GRUPO ISG SAÚDE",
    "Agência Work On (Grupo Nestlé)": "GRUPO NESTLÉ",
}

row2 = 10
for tier in ["A", "B", "C"]:
    cfg = tier_cfg[tier]

    ws2.merge_cells(f"A{row2}:F{row2}")
    hc2 = ws2.cell(row=row2, column=1, value=cfg["label"])
    hc2.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    hc2.fill = fill(cfg["dk"])
    hc2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hc2.border = thin(cfg["dk"])
    ws2.row_dimensions[row2].height = 20
    row2 += 1

    for i, (nome, mrr, obs) in enumerate(todos_data[tier], 1):
        is_isg = nome.startswith("ISG —")
        grupo  = GRUPOS.get(nome, "")

        cl2(ws2, row2, 1, i,     bold=True, color=cfg["dk"], bg=cfg["lt"], h="center")
        cl2(ws2, row2, 2, nome,  bg=cfg["lt"])
        cl2(ws2, row2, 3, grupo, bg=cfg["lt"], color=cfg["dk"],
            bold=is_isg, h="center")
        mrr2c = cl2(ws2, row2, 4, mrr, bold=True, color=cfg["dk"], bg=cfg["lt"],
                    h="right", num_fmt='R$ #,##0.00')
        cl2(ws2, row2, 5, f"Tier {tier}", bold=True, color=WHITE,
            bg=cfg["dk"], h="center")
        cl2(ws2, row2, 6, obs, bg=cfg["lt"], size=9, color="555555",
            wrap=True, v="top")

        ws2.row_dimensions[row2].height = 60 if is_isg else (28 if obs else 20)
        row2 += 1

# ── Total geral ───────────────────────────────────────────────────────
row2 += 1
ws2.merge_cells(f"A{row2}:C{row2}")
tl2 = ws2.cell(row=row2, column=1, value="MRR Total Carteira (empresas únicas)")
tl2.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
tl2.fill = fill(TEAL_ESC)
tl2.alignment = Alignment(horizontal="right", vertical="center")
tl2.border = thin(TEAL_ESC)

tv2 = ws2.cell(row=row2, column=4, value=mrr2_total)
tv2.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
tv2.fill = fill(TEAL)
tv2.alignment = Alignment(horizontal="right", vertical="center")
tv2.number_format = 'R$ #,##0.00'
tv2.border = thin(TEAL)

ws2.cell(row=row2, column=5).fill = fill(TEAL_XLT)
ws2.cell(row=row2, column=6).fill = fill(TEAL_XLT)
ws2.row_dimensions[row2].height = 24

# Fundo externo ws2
for r in range(1, row2 + 15):
    for col in range(7, 20):
        ws2.cell(row=r, column=col).fill = fill(TEAL_XLT)
for r in range(row2 + 1, row2 + 15):
    for col in range(1, 7):
        ws2.cell(row=r, column=col).fill = fill(TEAL_XLT)

ws2.freeze_panes = "A10"
ws2.print_title_rows = "1:9"


wb.save(OUTPUT)
print(f"Arquivo gerado: {OUTPUT}")
print(f"  Tier A: {len(dados['A'])} clientes | MRR R$ {mrr_A:,.2f}")
print(f"  Tier B: {len(dados['B'])} clientes | MRR R$ {mrr_B:,.2f}")
print(f"  Tier C: {len(dados['C'])} clientes | MRR R$ {mrr_C:,.2f}")
print(f"  Total ativo: {len(dados['A'])+len(dados['B'])+len(dados['C'])} clientes | MRR R$ {total_mrr:,.2f}")
print(f"  A Verificar: {len(a_verificar)} clientes | Churn confirmado: {len(churn)}")
