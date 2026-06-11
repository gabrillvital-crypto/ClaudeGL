from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

TIER_COLORS = {
    'A':  {'bg': 'F59E0B', 'text': '1F2937', 'light': 'FEF3C7'},
    'B+': {'bg': '3B82F6', 'text': 'FFFFFF', 'light': 'DBEAFE'},
    'B':  {'bg': '10B981', 'text': 'FFFFFF', 'light': 'D1FAE5'},
    'C':  {'bg': '6B7280', 'text': 'F9FAFB', 'light': 'F3F4F6'},
}

DARK = '1F2937'
WHITE = 'FFFFFF'
ACCENT = 'E0F2FE'

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

def font(color=DARK, bold=False, size=10, italic=False):
    return Font(name='Arial', color=color, bold=bold, size=size, italic=italic)

def align(h='center', wrap=False):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

thin = Side(style='thin', color='E5E7EB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, bg=None, fg=DARK, bold=False, size=10, h='center', wrap=False, fmt=None):
    if bg:
        cell.fill = fill(bg)
    cell.font = font(color=fg, bold=bold, size=size)
    cell.alignment = align(h=h, wrap=wrap)
    cell.border = border
    if fmt:
        cell.number_format = fmt


# ── ABA 1 — Modelo de Cadência ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Modelo de Cadência"
ws1.sheet_view.showGridLines = False

ws1.merge_cells('A1:H1')
ws1['A1'] = 'Modelo de Cadência por Tier — Carteira SRM Efcaz 2026'
style(ws1['A1'], bg='EFF6FF', fg=DARK, bold=True, size=14, h='center')
ws1.row_dimensions[1].height = 38

ws1.merge_cells('A2:H2')
ws1['A2'] = 'Tier A ≥ R$5.000  |  Tier B+ R$3.500–R$4.999  |  Tier B R$2.000–R$3.499  |  Tier C < R$2.000'
style(ws1['A2'], fg='6B7280', size=9, h='center')
ws1['A2'].font = Font(name='Arial', color='6B7280', size=9, italic=True)
ws1.row_dimensions[2].height = 18

ws1.row_dimensions[3].height = 6

headers1 = ['Tier', 'Faixa MRR', 'Clientes', 'MRR Total', 'Check-in', 'QBR', 'Canal Principal', 'SLA Resposta']
for c, h in enumerate(headers1, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    style(cell, bg=DARK, fg=WHITE, bold=True, size=10)
ws1.row_dimensions[4].height = 30

cadencia_data = [
    ('A',  'A — Estratégico',  '≥ R$ 5.000/mês',      'Mensal (call/presencial)', 'Trimestral', 'WhatsApp direto + E-mail', 'Mesmo dia'),
    ('B+', 'B+ — Alto Valor',  'R$ 3.500–R$ 4.999/mês', 'Bimestral (call)',         'Semestral',  'E-mail + WhatsApp',        '24 horas'),
    ('B',  'B — Médio Valor',  'R$ 2.000–R$ 3.499/mês', 'Trimestral (e-mail)',      'Anual',       'E-mail',                   '48 horas'),
    ('C',  'C — Menor Ticket', '< R$ 2.000/mês',        'Semestral (e-mail auto)',  'Sob demanda', 'E-mail',                   '72 horas'),
]
client_counts = {'A': 7, 'B+': 6, 'B': 15, 'C': 6}
mrr_refs = {5: 'D5', 6: 'D6', 7: 'D7', 8: 'D8'}

for i, (key, label, faixa, checkin, qbr, canal, sla) in enumerate(cadencia_data):
    r = i + 5
    clr = TIER_COLORS[key]
    row_vals = [label, faixa, client_counts[key], None, checkin, qbr, canal, sla]

    for c, v in enumerate(row_vals, 1):
        cell = ws1.cell(row=r, column=c, value=v)
        if c == 1:
            style(cell, bg=clr['bg'], fg=clr['text'], bold=True, size=10)
        else:
            style(cell, bg=clr['light'], fg=DARK, size=10, wrap=(c in [5, 7]))

    # MRR col D — hardcoded (reference values)
    mrr_vals = {'A': 43543.80, 'B+': 26782.00, 'B': 38853.94, 'C': 9588.35}
    ws1.cell(row=r, column=4).value = mrr_vals[key]
    style(ws1.cell(row=r, column=4), bg=clr['light'], fg=DARK, fmt='R$ #,##0.00')

    ws1.row_dimensions[r].height = 36

# Totais
r = 9
ws1.merge_cells(f'A{r}:C{r}')
ws1[f'A{r}'] = 'MRR Total Ativo'
style(ws1[f'A{r}'], bg=ACCENT, fg=DARK, bold=True, h='left')
ws1[f'D{r}'] = '=SUM(D5:D8)'
style(ws1[f'D{r}'], bg=ACCENT, fg=DARK, bold=True, fmt='R$ #,##0.00')
ws1.row_dimensions[r].height = 28

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 26
ws1.column_dimensions['C'].width = 11
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 24
ws1.column_dimensions['F'].width = 14
ws1.column_dimensions['G'].width = 26
ws1.column_dimensions['H'].width = 15


# ── ABA 2 — Clientes por Tier ────────────────────────────────────────────────
ws2 = wb.create_sheet("Clientes por Tier")
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:I1')
ws2['A1'] = 'Carteira SRM — Clientes por Tier e Cadência de Contato 2026'
style(ws2['A1'], bg='EFF6FF', fg=DARK, bold=True, size=14, h='center')
ws2.row_dimensions[1].height = 38

ws2.row_dimensions[2].height = 6

headers2 = ['Nº', 'Cliente', 'MRR/mês', 'Tier', 'Check-in', 'QBR', 'Canal', 'SLA', 'Observação']
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    style(cell, bg=DARK, fg=WHITE, bold=True, size=10)
ws2.row_dimensions[3].height = 28

clients_by_tier = [
    ('A', [
        (1,  'Zurich Airport Brasil',             8929.20, 'Mensal',    'Trimestral',  'WhatsApp + E-mail', 'Mesmo dia', None),
        (2,  'CSU Digital',                        7000.00, 'Mensal',    'Trimestral',  'WhatsApp + E-mail', 'Mesmo dia', '⚠ Sem contrato no CustomerX'),
        (3,  'Bom Futuro Agrícola',                6500.00, 'Mensal',    'Trimestral',  'WhatsApp + E-mail', 'Mesmo dia', None),
        (4,  'ECTX S/A — Eucatex',                5509.40, 'Mensal',    'Trimestral',  'WhatsApp + E-mail', 'Mesmo dia', None),
        (5,  'Norskan Offshore (DOF)',             5398.00, 'Mensal',    'Trimestral',  'WhatsApp + E-mail', 'Mesmo dia', 'Promovido de Tier B'),
        (6,  'Soluções Terceirizadas',             5200.00, 'Mensal',    'Trimestral',  'WhatsApp + E-mail', 'Mesmo dia', 'Promovido de Tier B'),
        (7,  'Unimed do Brasil',                   5007.20, 'Mensal',    'Trimestral',  'WhatsApp + E-mail', 'Mesmo dia', None),
    ]),
    ('B+', [
        (1,  'Afonso Franca Engenharia',           4800.00, 'Bimestral', 'Semestral',   'E-mail + WhatsApp', '24 horas',  None),
        (2,  'DATA Engenharia',                    4700.00, 'Bimestral', 'Semestral',   'E-mail + WhatsApp', '24 horas',  '⚠ Sem contrato no CustomerX'),
        (3,  'Geistlich Pharma do Brasil',         4634.00, 'Bimestral', 'Semestral',   'E-mail + WhatsApp', '24 horas',  None),
        (4,  'Vinci Airports',                     4500.00, 'Bimestral', 'Semestral',   'E-mail + WhatsApp', '24 horas',  '⚠ Sem contrato no CustomerX'),
        (5,  'Hospital Adventista de Belém',       4309.00, 'Bimestral', 'Semestral',   'E-mail + WhatsApp', '24 horas',  None),
        (6,  'Bunker One Combustíveis e Lubrificantes', 3839.00, 'Bimestral', 'Semestral', 'E-mail + WhatsApp', '24 horas', None),
    ]),
    ('B', [
        (1,  'ISG — Instituto Sócrates Guanaes',  3784.99, 'Trimestral','Anual',        'E-mail',            '48 horas',  '7 unidades consolidadas'),
        (2,  'Dock Brasil Engenharia e Serviços', 3644.00, 'Trimestral','Anual',        'E-mail',            '48 horas',  None),
        (3,  'Cielo',                             3450.24, 'Trimestral','Anual',        'E-mail',            '48 horas',  'Contrato anual'),
        (4,  'Sabarã Químicos e Ingredientes',    2504.30, 'Trimestral','Anual',        'E-mail',            '48 horas',  None),
        (5,  'Agência Work On (Grupo Nestlé)',     2500.00, 'Trimestral','Anual',        'E-mail',            '48 horas',  None),
        (6,  'Tarkett',                           2500.00, 'Trimestral','Anual',        'E-mail',            '48 horas',  None),
        (7,  'Premier Pet',                       2500.00, 'Trimestral','Anual',        'E-mail',            '48 horas',  '⚠ Sem contrato no CustomerX'),
        (8,  'Federação Paulista de Futebol',     2450.00, 'Trimestral','Anual',        'E-mail',            '48 horas',  'Projeto encerrado — verificar recorrência'),
        (9,  'Transportes Cavalinho',             2385.00, 'Trimestral','Anual',        'E-mail',            '48 horas',  '⚠ Sem contrato no CustomerX'),
        (10, 'Pacco / Paccoby',                   2309.00, 'Trimestral','Anual',        'E-mail',            '48 horas',  '⚠ Sem contrato no CustomerX'),
        (11, 'Engesp',                            2270.00, 'Trimestral','Anual',        'E-mail',            '48 horas',  'Contrato anual'),
        (12, 'Cebrace',                           2255.87, 'Trimestral','Anual',        'E-mail',            '48 horas',  None),
        (13, 'BRG Suplementos Nutricionais',      2181.92, 'Trimestral','Anual',        'E-mail',            '48 horas',  None),
        (14, 'Ponsse Latin America',              2118.62, 'Trimestral','Anual',        'E-mail',            '48 horas',  None),
        (15, 'Killing SA Tintas e Adesivos',      2000.00, 'Trimestral','Anual',        'E-mail',            '48 horas',  None),
    ]),
    ('C', [
        (1,  'Unimed Campo Grande',               1967.35, 'Semestral', 'Sob demanda', 'E-mail',            '72 horas',  None),
        (2,  'Asso Marítima Navegação',           1899.00, 'Semestral', 'Sob demanda', 'E-mail',            '72 horas',  None),
        (3,  'Amboretto Bombas',                  1700.00, 'Semestral', 'Sob demanda', 'E-mail',            '72 horas',  None),
        (4,  'Unimed de Dourados',                1672.00, 'Semestral', 'Sob demanda', 'E-mail',            '72 horas',  None),
        (5,  'Alumetaf — Soluções em Ferro Fundido', 1650.00, 'Semestral', 'Sob demanda', 'E-mail',         '72 horas',  'Reclassificado Tier B→C'),
        (6,  'Advtec Indústria e Comércio',        700.00, 'Semestral', 'Sob demanda', 'E-mail',            '72 horas',  None),
    ]),
]

tier_labels = {
    'A':  'TIER A — Estratégico  (MRR ≥ R$ 5.000)',
    'B+': 'TIER B+ — Alto Valor  (MRR R$ 3.500 – R$ 4.999)',
    'B':  'TIER B — Médio Valor  (MRR R$ 2.000 – R$ 3.499)',
    'C':  'TIER C — Menor Ticket  (MRR < R$ 2.000)',
}

current_row = 4
subtotal_rows = []

for tier_key, tier_clients in clients_by_tier:
    clr = TIER_COLORS[tier_key]

    # Section header
    ws2.merge_cells(f'A{current_row}:I{current_row}')
    ws2[f'A{current_row}'] = tier_labels[tier_key]
    style(ws2[f'A{current_row}'], bg=clr['bg'], fg=clr['text'], bold=True, size=11, h='left')
    ws2.row_dimensions[current_row].height = 28
    current_row += 1

    start_data = current_row

    for i, (num, name, mrr, checkin, qbr, canal, sla, obs) in enumerate(tier_clients):
        row_bg = clr['light'] if i % 2 == 0 else 'FFFFFF'
        values = [num, name, mrr, f'Tier {tier_key}', checkin, qbr, canal, sla, obs or '']

        for c, v in enumerate(values, 1):
            cell = ws2.cell(row=current_row, column=c, value=v)
            style(cell, bg=row_bg, fg=DARK, size=10,
                  h='left' if c in [2, 9] else 'center',
                  fmt='R$ #,##0.00' if c == 3 else None)
        ws2.row_dimensions[current_row].height = 22
        current_row += 1

    # Subtotal
    ws2.merge_cells(f'A{current_row}:B{current_row}')
    ws2[f'A{current_row}'] = f'Subtotal Tier {tier_key}'
    style(ws2[f'A{current_row}'], bg=clr['light'], fg=DARK, bold=True, h='left')
    ws2[f'C{current_row}'] = f'=SUM(C{start_data}:C{current_row - 1})'
    style(ws2[f'C{current_row}'], bg=clr['light'], fg=DARK, bold=True, fmt='R$ #,##0.00')
    subtotal_rows.append(current_row)
    ws2.row_dimensions[current_row].height = 24
    current_row += 2

# Total geral
ws2.merge_cells(f'A{current_row}:B{current_row}')
ws2[f'A{current_row}'] = 'MRR Total Ativo'
style(ws2[f'A{current_row}'], bg=ACCENT, fg=DARK, bold=True, size=11, h='left')
total_formula = '+'.join([f'C{r}' for r in subtotal_rows])
ws2[f'C{current_row}'] = f'={total_formula}'
style(ws2[f'C{current_row}'], bg=ACCENT, fg=DARK, bold=True, size=11, fmt='R$ #,##0.00')
ws2.row_dimensions[current_row].height = 32

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 38
ws2.column_dimensions['C'].width = 17
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 22
ws2.column_dimensions['H'].width = 12
ws2.column_dimensions['I'].width = 36

OUTPUT = r'C:\Users\gabriel.evangelista\Documents\ClaudeGL\Cadencia_Tiers_Efcaz.xlsx'
wb.save(OUTPUT)
print(f"Salvo: {OUTPUT}")
