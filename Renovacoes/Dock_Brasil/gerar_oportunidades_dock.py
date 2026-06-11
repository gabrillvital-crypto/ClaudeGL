# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from math import ceil

OUTPUT    = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\Oportunidades_Dock_Brasil.xlsx"
LOGO_EFC  = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\logo_efcaz_clean.png"
LOGO_DOCK = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\logo_dock_proc.png"

TEAL     = "0E8FA3"
TEAL_ESC = "0A6A7A"
TEAL_XLT = "E0F4F7"
WHITE    = "FFFFFF"
GRAY_BOX = "D5D8DC"
GREEN_LT = "D5F5E3"
GREEN_DK = "1E8449"
AMBER_LT = "FEF9E7"
AMBER_DK = "B7770D"
BLUE_LT  = "D6EAF8"
BLUE_DK  = "1A5276"

def fill(h): return PatternFill("solid", fgColor=h)
def thin(c=GRAY_BOX):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, val="", bold=False, size=10, color=None, bg=None,
         h="left", v="center", wrap=False, border=True, italic=False):
    cl = ws.cell(row=row, column=col, value=val)
    kw = {"name": "Calibri", "bold": bold, "size": size, "italic": italic}
    if color: kw["color"] = color
    cl.font = Font(**kw)
    if bg: cl.fill = fill(bg)
    cl.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border: cl.border = thin()
    return cl

def build_header(ws, title, subtitle):
    for col, w in zip("ABCDEF", [5, 50, 42, 22, 22, 24]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 65
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18

    # ── Linha 1: logo Dock (esquerda) | logo Efcaz (direita)
    ws.merge_cells("A1:C1")
    lc = ws["A1"]
    lc.fill = fill(WHITE)
    lc.border = thin(GRAY_BOX)
    img_dock = XLImage(LOGO_DOCK)
    img_dock.height = 52
    img_dock.width  = int(402 * 52 / 190)
    ws.add_image(img_dock, "A1")

    ws.merge_cells("D1:F1")
    ws["D1"].fill = fill(WHITE)
    ws["D1"].border = thin(GRAY_BOX)
    img_efc = XLImage(LOGO_EFC)
    img_efc.height = 52
    img_efc.width  = int(223 * 52 / 78)
    ws.add_image(img_efc, "D1")

    # ── Linha 2: título principal
    ws.merge_cells("A2:F2")
    t = ws["A2"]
    t.value = title
    t.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    t.fill = fill(TEAL)
    t.alignment = Alignment(horizontal="center", vertical="center")

    # ── Linha 3: subtítulo
    ws.merge_cells("A3:F3")
    s = ws["A3"]
    s.value = subtitle
    s.font = Font(name="Calibri", size=9, italic=True, color="555555")
    s.fill = fill(TEAL_XLT)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Linha 4: cabeçalhos de colunas
    ws.row_dimensions[4].height = 22
    for col, label in zip(range(1, 7),
        ["Nº", "Consulta Automática Efcaz",
         "Documento Dock Correspondente",
         "Situação na Dock", "Categoria", "Impacto"]):
        cl = ws.cell(row=4, column=col, value=label)
        cl.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cl.fill = fill(TEAL_ESC)
        cl.alignment = Alignment(horizontal="center", vertical="center")
        cl.border = thin(TEAL_ESC)

def section_header(ws, row, label):
    ws.merge_cells(f"A{row}:F{row}")
    hc = ws.cell(row=row, column=1, value=label)
    hc.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    hc.fill = fill(TEAL)
    hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hc.border = thin(TEAL_ESC)
    ws.row_dimensions[row].height = 20

def data_row(ws, row, num, busca, doc, situacao, cat, impacto, bg_lt, bg_dk, bold_status=False):
    cell(ws, row, 1, num,      h="center", bg=bg_lt, bold=True, color=bg_dk)
    cell(ws, row, 2, busca,    bg=bg_lt, wrap=True)
    cell(ws, row, 3, doc,      bg=bg_lt, wrap=True, size=9, color="333333")
    cell(ws, row, 4, situacao, bg=bg_lt, h="center", size=9, bold=bold_status, color=bg_dk)
    cell(ws, row, 5, cat,      bg=bg_lt, h="center", size=9)
    cell(ws, row, 6, impacto,  bg=bg_lt, wrap=True,  size=9, italic=True, color="444444")
    lines = max(1, ceil(max(len(busca), len(doc)) / 42))
    ws.row_dimensions[row].height = max(28, lines * 16 + 4)

def ext_fill(ws, row):
    for r in range(1, row + 10):
        for col in range(7, 25):
            ws.cell(row=r, column=col).fill = fill(TEAL_XLT)
    for r in range(row, row + 10):
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = fill(TEAL_XLT)


SUBTITLE = ("Lista completa Efcaz (400+ consultas) cruzada com documentos ativos Dock Brasil — "
            "exibindo apenas consultas com correspondência direta ao que o cliente já coleta hoje.")

# ═══════════════════════════════════════════════════════════
# ABA 1 — Oportunidades de Automação
# ═══════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Oportunidades de Automação"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = TEAL

build_header(ws1,
    "Oportunidades de Automação — Dock Brasil  ×  Efcaz",
    SUBTITLE)

row = 5

section_header(ws1, row, "JÁ ATIVOS COMO BACKGROUND CHECK")
row += 1

bg_check = [
    ("Caixa / Regularidade do Empregador (FGTS)",
     "FGTS + Relatório + Guia + Comprovante",
     "Background Check ativo", "Trabalhista",
     "Renovação automática — sem envio manual"),
    ("Receita Federal / CNPJ",
     "Cartão CNPJ",
     "Background Check ativo", "Cadastral",
     "Atualização automática de dados cadastrais"),
    ("Receita Federal / PGFN (CND Federal)",
     "Certidão Negativa de Débito Federal (CND)",
     "Background Check ativo", "Fiscal Federal",
     "Renovação automática — sem envio manual"),
    ("Receita Federal / PGFN (CND Federal) - Nova",
     "Certidão Negativa de Débito Federal (CND)",
     "Background Check ativo", "Fiscal Federal",
     "Versão atualizada da mesma certidão"),
    ("Tribunal / TST / CNDT",
     "Certidão Negativa de Débitos Trabalhistas",
     "Background Check ativo", "Trabalhista",
     "Renovação automática — sem envio manual"),
    ("Tribunal / TST / Validação de CNDT",
     "Certidão Negativa de Débitos Trabalhistas",
     "Background Check ativo", "Trabalhista",
     "Validação cruzada da CNDT emitida"),
]
for i, (b, d, s, cat, imp) in enumerate(bg_check, 1):
    data_row(ws1, row, i, b, d, s, cat, imp, GREEN_LT, GREEN_DK, bold_status=True)
    row += 1

section_header(ws1, row, "OPORTUNIDADES DE AUTOMAÇÃO — COLETA HOJE MANUAL")
row += 1

oportunidades = [
    ("IBAMA / Certificado de Regularidade",
     "Cadastro Técnico Federal (IBAMA)",
     "Coletado manualmente", "Ambiental",
     "Substitui envio do certificado pelo fornecedor"),
    ("IBAMA / Autuações Ambientais",
     "Licença de Operação Ambiental / Licença Ambiental",
     "Coletado manualmente", "Ambiental",
     "Detecta autuações que o doc não revela"),
    ("IBAMA / Certidão de Débitos",
     "Licença de Operação Ambiental / Licença Ambiental",
     "Coletado manualmente", "Ambiental",
     "Verifica pendências sem esperar renovação"),
    ("IBAMA / Certidão de Embargos (Nada Consta)",
     "Licença de Operação Ambiental / Licença Ambiental",
     "Coletado manualmente", "Ambiental",
     "Confirma ausência de embargo ativo"),
    ("ANVISA / Funcionamento de Empresa Nacional",
     "Autorização de Funcionamento de Empresa (AFE)",
     "Coletado manualmente", "Sanitário",
     "Verifica validade da AFE automaticamente"),
    ("Secretaria de Inspeção do Trabalho / CAEPI",
     "Gestão de CA - EPI e EPCs",
     "Coletado manualmente", "Segurança do Trabalho",
     "Certifica EPI/EPC usados nas NRs exigidas"),
    ("Secretaria de Inspeção do Trabalho / Trabalho Escravo",
     'Situação Cadastral (se a empresa está ativa), Consulta na "Lista Suja" do MTE.',
     "Coletado manualmente", "Segurança do Trabalho",
     "Reforço de due diligence trabalhista"),
    ("MTE / Certidão de Débitos Trabalhistas",
     "Certidão Negativa de Débitos Trabalhistas",
     "Coletado manualmente", "Trabalhista",
     "Complementa CNDT com visão MTE"),
    ("Receita Federal / Radar (Habilitação Comércio Exterior)",
     "Título de Inscrição da Embarcação / equip. importado",
     "Coletado manualmente", "Comércio Exterior",
     "Verifica habilitação para importar equip. navais"),
    ("Cadastro Ambiental Rural / Demonstrativo",
     "Outorga dos Dir. de Uso de Recursos Hídricos",
     "Coletado manualmente", "Ambiental",
     "Complementa outorga hídrica com CAR"),
]
for i, (b, d, s, cat, imp) in enumerate(oportunidades, 1):
    data_row(ws1, row, i, b, d, s, cat, imp, AMBER_LT, AMBER_DK)
    row += 1

# Legenda Aba 1
row += 1
ws1.merge_cells(f"A{row}:F{row}")
lg = ws1.cell(row=row, column=1, value="Legenda")
lg.font = Font(name="Calibri", bold=True, size=10, color=TEAL_ESC)
ws1.row_dimensions[row].height = 18
row += 1

for bg, dk, label in [
    (GREEN_LT, GREEN_DK, "Já ativo como Background Check — coleta automática já em funcionamento na plataforma"),
    (AMBER_LT, AMBER_DK, "Oportunidade de automação — Dock coleta manualmente, Efcaz já possui a consulta nativa"),
]:
    ws1.merge_cells(f"B{row}:F{row}")
    l1 = ws1.cell(row=row, column=1); l1.fill = fill(bg); l1.border = thin()
    l2 = ws1.cell(row=row, column=2, value=label)
    l2.font = Font(name="Calibri", size=9, color=dk)
    l2.fill = fill(bg); l2.border = thin()
    l2.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[row].height = 16
    row += 1

ext_fill(ws1, row)
ws1.freeze_panes = None
ws1.print_title_rows = "1:4"


# ═══════════════════════════════════════════════════════════
# ABA 2 — Para Desenvolvimento
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Para Desenvolvimento")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = BLUE_DK

build_header(ws2,
    "Oportunidades de Automação — Dock Brasil  ×  Efcaz - PARA DESENVOLVIMENTO (Pode haver investimento)",
    SUBTITLE)

row2 = 5
section_header(ws2, row2, "OPORTUNIDADES DE AUTOMAÇÃO")
row2 += 1

dev_dados = [
    ("MTE / Processos por Empregador",
     "Situação de processos e infrações trabalhistas vinculadas ao CNPJ",
     "Ainda não coletado por vocês", "Trabalhista",
     "Identifica processos sem esperar documentos",
     GREEN_LT, GREEN_DK, True),
    ("ANP / Base de Distribuição",
     "Operação em Base de Distribuição em Plataforma de Petróleo",
     "Coletado manualmente", "Regulatório O&G",
     "Confirma habilitação junto à ANP",
     AMBER_LT, AMBER_DK, False),
    ("ANP / Certificados",
     "Certificados da Unidade Plataforma de Petróleo",
     "Coletado manualmente", "Regulatório O&G",
     "Valida certificados emitidos pela ANP",
     AMBER_LT, AMBER_DK, False),
    ("ANP / Instalações do SIMP",
     "Dados do SIMP (Sistema de Informação)",
     "Coletado manualmente", "Regulatório O&G",
     "Confirma instalações no registro ANP",
     AMBER_LT, AMBER_DK, False),
    ("ANP / Postos",
     "Segurança em Postos de Serviço",
     "Coletado manualmente", "Regulatório O&G",
     "Verifica regularidade de postos/combustível",
     AMBER_LT, AMBER_DK, False),
    ("ANP / Revendas GLP",
     "Armazenamento de Gás Liquefeito",
     "Coletado manualmente", "Regulatório O&G",
     "GLP em embarcações e plataformas",
     AMBER_LT, AMBER_DK, False),
]
for i, (b, d, s, cat, imp, bg_lt, bg_dk, bs) in enumerate(dev_dados, 1):
    data_row(ws2, row2, i, b, d, s, cat, imp, bg_lt, bg_dk, bold_status=bs)
    row2 += 1

# Legenda Aba 2
row2 += 1
ws2.merge_cells(f"A{row2}:F{row2}")
lg2 = ws2.cell(row=row2, column=1, value="Legenda")
lg2.font = Font(name="Calibri", bold=True, size=10, color=TEAL_ESC)
ws2.row_dimensions[row2].height = 18
row2 += 1

# Entrada 1: "Ainda não coletado"
ws2.merge_cells(f"B{row2}:F{row2}")
e1a = ws2.cell(row=row2, column=1); e1a.fill = fill(GREEN_LT); e1a.border = thin()
e1b = ws2.cell(row=row2, column=2,
    value="Ainda não coletado por vocês, mas também é uma opção de busca automática caso interessem.")
e1b.font = Font(name="Calibri", size=9, color=GREEN_DK)
e1b.fill = fill(GREEN_LT); e1b.border = thin()
e1b.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[row2].height = 16
row2 += 1

# Entrada 2: ANP detalhado com quebras de linha
ws2.merge_cells(f"B{row2}:F{row2}")
e2a = ws2.cell(row=row2, column=1); e2a.fill = fill(GREEN_LT); e2a.border = thin()

anp_text = (
    "Linhas de 2 ate 6 - ANP:\n\n"
    "Situação Cadastral: Status da autorização (ativa, suspensa, inabilitada, cancelada).\n"
    "Dados da Empresa: Razão social, nome fantasia, CNPJ, endereço completo.\n"
    "Atividades Autorizadas: Tipo de atividade (posto revendedor, distribuidor, TRR - Transportador-Revendedor-Retalhista).\n"
    "Informações de Operação: Número do despacho de autorização, data de publicação da autorização.\n"
    "Documentação de Local: Dados sobre o alvará de funcionamento e licença ambiental.\n"
    "Dados de Tancagem: Capacidade de armazenamento de combustíveis.\n"
    "Situação de Interdição: Informações sobre interdições totais ou parciais.\n"
    "Sócios: Lista de sócios cadastrados."
)
e2b = ws2.cell(row=row2, column=2, value=anp_text)
e2b.font = Font(name="Calibri", size=9, bold=True, color=GREEN_DK)
e2b.fill = fill(GREEN_LT); e2b.border = thin()
e2b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws2.row_dimensions[row2].height = 150
row2 += 1

# Entrada 3: "Coletado manualmente"
ws2.merge_cells(f"B{row2}:F{row2}")
e3a = ws2.cell(row=row2, column=1); e3a.fill = fill(AMBER_LT); e3a.border = thin()
e3b = ws2.cell(row=row2, column=2,
    value="Oportunidade de automação — Dock coleta manualmente, Efcaz já possui a consulta nativa.")
e3b.font = Font(name="Calibri", size=9, color=AMBER_DK)
e3b.fill = fill(AMBER_LT); e3b.border = thin()
e3b.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[row2].height = 16
row2 += 1

ext_fill(ws2, row2)
ws2.freeze_panes = None
ws2.print_title_rows = "1:4"


wb.save(OUTPUT)
print(f"Arquivo gerado: {OUTPUT}")
print(f"  Aba 1: {len(bg_check)} BG Check + {len(oportunidades)} manuais")
print(f"  Aba 2: {len(dev_dados)} linhas (1 MTE + 5 ANP)")
