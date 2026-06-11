# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

OUTPUT = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\Buscas_Dock_Brasil_v2.xlsx"
LOGO   = r"C:\Users\gabriel.evangelista\Documents\ClaudeGL\logo_efcaz_clean.png"

# Paleta Efcaz
TEAL      = "0E8FA3"
TEAL_ESC  = "0A6A7A"
TEAL_XLT  = "E0F4F7"
WHITE     = "FFFFFF"
GRAY_LT   = "F5F6FA"
GRAY_BOX  = "D5D8DC"
GREEN_LT  = "D5F5E3"
GREEN_DK  = "1E8449"
AMBER_LT  = "FEF9E7"
AMBER_DK  = "B7770D"
BLUE_LT   = "EBF5FB"
BLUE_DK   = "1A5276"
ORANGE_LT = "FDEBD0"
ORANGE_DK = "A04000"
HEADER_BG = "EBF5FB"

def fill(h): return PatternFill("solid", fgColor=h)
def thin(color=GRAY_BOX):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, val="", bold=False, size=10, color=None, bg=None,
         h="left", v="center", wrap=False, border=True, fmt=None, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    kw = {"name": "Calibri", "bold": bold, "size": size, "italic": italic}
    if color: kw["color"] = color
    c.font = Font(**kw)
    if bg: c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border: c.border = thin()
    if fmt: c.number_format = fmt
    return c

def header_row(ws, row, cols_vals, bg, color=WHITE, size=10):
    for col, val in cols_vals:
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, size=size, color=color)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin(TEAL_ESC)

def add_logo_header(ws, title, ncols=6):
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 28

    ws.merge_cells(f"A1:{chr(64+ncols//2)}1")
    c = ws["A1"]
    c.value = "[ Logo do Cliente ]"
    c.font = Font(name="Calibri", size=9, italic=True, color=WHITE)
    c.fill = fill(TEAL_ESC)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"{chr(64+ncols//2+1)}1:{chr(64+ncols)}1")
    c = ws[f"{chr(64+ncols//2+1)}1"]
    c.fill = fill(WHITE)
    img = XLImage(LOGO)
    img.height = 50
    img.width  = int(223 * 50 / 78)
    ws.add_image(img, f"{chr(64+ncols//2+1)}1")

    ws.merge_cells(f"A2:{chr(64+ncols)}2")
    c = ws["A2"]
    c.value = title
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════
# ABA 1 — BUSCAS AUTOMÁTICAS
# ══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Buscas Automáticas"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = TEAL

for col, w in zip("ABCDEF", [5, 52, 22, 20, 14, 20]):
    ws1.column_dimensions[col].width = w

add_logo_header(ws1, "Consultas Automáticas — Dock Brasil", ncols=6)

header_row(ws1, 3,
    [(1,"Nº"),(2,"Fonte / Consulta"),(3,"Categoria"),(4,"Grupo de Relevância"),(5,"Tipo"),(6,"Observação")],
    bg=TEAL_ESC)
ws1.row_dimensions[3].height = 20

GRUPOS = {
    "BG Ativo": (GREEN_LT, GREEN_DK, "Já ativo na plataforma"),
    "Automação": (AMBER_LT, AMBER_DK, "Oportunidade de automação"),
    "Compliance": (BLUE_LT, BLUE_DK, "Compliance regulatório"),
    "Due Diligence": (ORANGE_LT, ORANGE_DK, "Due diligence geral"),
}

buscas = [
    # (nº, fonte, categoria, grupo, tipo, obs)
    (1,  "ANP / Base de Distribuição",                              "Regulatório O&G", "Automação",    "PJ", "Reguladora direta do setor"),
    (2,  "ANP / Certificados",                                      "Regulatório O&G", "Automação",    "PJ", ""),
    (3,  "ANP / Instalações do SIMP",                               "Regulatório O&G", "Automação",    "PJ", ""),
    (4,  "ANP / Instalações do SIMP / Distribuidores Autorizados",  "Regulatório O&G", "Automação",    "PJ", ""),
    (5,  "ANTT / Transportador",                                    "Transporte",      "Due Diligence","PJ", ""),
    (6,  "Antecedentes Criminais / Polícia Federal / Emitir",       "Antecedentes",    "Due Diligence","PF", ""),
    (7,  "Banco Central do Brasil (BCB) / Cheques sem Fundos",      "Financeiro",      "Due Diligence","PJ", ""),
    (8,  "Buscador / Google",                                       "Geral",           "Due Diligence","PJ", ""),
    (9,  "CADE / Processos",                                        "Concorrência",    "Due Diligence","PJ", ""),
    (10, "CARF / Processo",                                         "Fiscal",          "Compliance",   "PJ", ""),
    (11, "CENPROT SP / Protestos",                                  "Financeiro",      "Due Diligence","PJ", ""),
    (12, "CRCP / Central de Registros de Certificados Profissionais","Habilitação",    "Due Diligence","PJ", ""),
    (13, "CVM / Participante",                                      "Capital",         "Due Diligence","PJ", ""),
    (14, "CVM / Processos Administrativos Sancionadores",           "Capital",         "Due Diligence","PJ", ""),
    (15, "Caixa / Regularidade do Empregador (FGTS)",              "Trabalhista",     "BG Ativo",     "PJ", "Já ativo como Background Check"),
    (16, "CNJ / Improbidade Administrativa e Inelegibilidade",      "Judicial",        "Compliance",   "PJ", ""),
    (17, "CNJ / Mandados de Prisão",                                "Judicial",        "Compliance",   "PF", ""),
    (18, "CNJ / SEEU / Processos",                                  "Judicial",        "Compliance",   "PJ", ""),
    (19, "CGU / Certidão Negativa Correcional (CGU-PJ, CEIS, CNEP e CEPIM)","Integridade","Compliance","PJ",""),
    (20, "CGU / Certidão Negativa Correcional (ePAD e CGU-PAD)",   "Integridade",     "Compliance",   "PJ", ""),
    (21, "EU Financial Sanctions List (PF/PJ)",                     "Sanções Internacionais","Due Diligence","PF/PJ",""),
    (22, "FBI - Most Wanted (PF)",                                  "Sanções Internacionais","Due Diligence","PF",""),
    (23, "FINCEN - Financial Crimes Enforcement Network (PJ)",      "Sanções Internacionais","Due Diligence","PJ",""),
    (24, "IBAMA / Autuações Ambientais",                            "Ambiental",       "Automação",    "PJ", "Exigem docs IBAMA manualmente"),
    (25, "IBAMA / Certidão de Débitos",                             "Ambiental",       "Automação",    "PJ", ""),
    (26, "IBAMA / Certidão de Embargos (Nada Consta)",              "Ambiental",       "Automação",    "PJ", ""),
    (27, "IBAMA / Certificado de Regularidade",                     "Ambiental",       "Automação",    "PJ", ""),
    (28, "IEPTB (CENPROT) / Protestos",                             "Financeiro",      "Due Diligence","PJ", ""),
    (29, "INTERPOL (PF)",                                           "Sanções Internacionais","Due Diligence","PF",""),
    (30, "Junta Comercial / SP / Certidão Simplificada",            "Cadastral",       "Due Diligence","PJ", ""),
    (31, "MPF / Lava-Jato",                                         "Integridade",     "Compliance",   "PJ", "Relevante para O&G"),
    (32, "MPF / Processos",                                         "Judicial",        "Compliance",   "PJ", ""),
    (33, "MPT / RJ / Certidão Negativa de Feitos",                  "Trabalhista",     "Compliance",   "PJ", "Jurisdição RJ"),
    (34, "MPT / Unificada / Certidão Negativa de Feitos",           "Trabalhista",     "Compliance",   "PJ", ""),
    (35, "MTE / Certidão de Débitos Trabalhistas",                  "Trabalhista",     "Compliance",   "PJ", ""),
    (36, "MTE / Processos por Empregador",                          "Trabalhista",     "Compliance",   "PJ", ""),
    (37, "OFAC - Sanções (PF)",                                     "Sanções Internacionais","Due Diligence","PF",""),
    (38, "OFAC / Sanções",                                          "Sanções Internacionais","Due Diligence","PJ",""),
    (39, "ONU / Sanções",                                           "Sanções Internacionais","Due Diligence","PJ",""),
    (40, "Portal da Transparência / Acordos de Leniência",          "Integridade",     "Compliance",   "PJ", ""),
    (41, "Portal da Transparência / Busca",                         "Integridade",     "Compliance",   "PJ", ""),
    (42, "Portal da Transparência / CNEP",                          "Integridade",     "Compliance",   "PJ", ""),
    (43, "Portal da Transparência / CEIS",                          "Integridade",     "Compliance",   "PJ", ""),
    (44, "Portal da Transparência / CEPIM",                         "Integridade",     "Compliance",   "PJ", ""),
    (45, "Portal da Transparência / CEAF",                          "Integridade",     "Compliance",   "PJ", ""),
    (46, "Portal da Transparência / Convênios e Acordos",           "Integridade",     "Compliance",   "PJ", ""),
    (47, "Prefeitura / RJ / Duque de Caxias / CND",                 "Fiscal Municipal","Compliance",   "PJ", "Sede em São Gonçalo/RJ"),
    (48, "Prefeitura / RJ / Rio de Janeiro / CND",                  "Fiscal Municipal","Compliance",   "PJ", ""),
    (49, "Procuradoria Geral do Estado / RJ / Dívida Ativa",        "Fiscal Estadual", "Compliance",   "PJ", ""),
    (50, "Receita Federal / CNPJ",                                  "Cadastral",       "BG Ativo",     "PJ", "Já ativo como Background Check"),
    (51, "Receita Federal / CPF",                                   "Cadastral",       "Due Diligence","PF", ""),
    (52, "Receita Federal / PGFN (CND Federal)",                    "Fiscal Federal",  "BG Ativo",     "PJ", "Já ativo como Background Check"),
    (53, "Receita Federal / PGFN (CND Federal) - Nova",             "Fiscal Federal",  "BG Ativo",     "PJ", "Já ativo como Background Check"),
    (54, "Receita Federal / Radar (Habilitação Comércio Exterior)",  "Comércio Exterior","Automação",  "PJ", "Importação de equip. navais"),
    (55, "Receita Federal / Simples Nacional",                      "Fiscal Federal",  "Due Diligence","PJ", ""),
    (56, "SEFAZ / Certidão Negativa de Débitos Estaduais",          "Fiscal Estadual", "Compliance",   "PJ", ""),
    (57, "SEFAZ / RJ / Certidão Negativa de Débitos",               "Fiscal Estadual", "Compliance",   "PJ", ""),
    (58, "Secretaria de Inspeção do Trabalho / CAEPI",              "Segurança do Trabalho","Automação","PJ","Relacionado às NRs exigidas"),
    (59, "Secretaria de Inspeção do Trabalho / Trabalho Escravo",   "Segurança do Trabalho","Automação","PJ","Alinhado ao perfil de compliance"),
    (60, "TCU / Certidão Negativa de Inabilitado ou Inidôneo",      "Integridade",     "Compliance",   "PJ", ""),
    (61, "TCU / Consulta Consolidada de Pessoa Jurídica (APF)",     "Integridade",     "Compliance",   "PJ", ""),
    (62, "TCU / Inabilitados",                                      "Integridade",     "Compliance",   "PJ", ""),
    (63, "TCU / Relação de Inidôneos",                              "Integridade",     "Compliance",   "PJ", ""),
    (64, "Tribunal / TJRJ / Cadastro de Pedido de Certidão",        "Judicial",        "Compliance",   "PJ", ""),
    (65, "Tribunal / TJRJ / Processo",                              "Judicial",        "Compliance",   "PJ", ""),
    (66, "Tribunal / TJRJ / Visualizar Certidão",                   "Judicial",        "Compliance",   "PJ", ""),
    (67, "Tribunal / TRF2 / Certidão Negativa Cível e Criminal",    "Judicial",        "Compliance",   "PJ", "Abrange RJ"),
    (68, "Tribunal / TRT1 / Certidão Eletrônica de Ações Trabalhistas (CEAT)","Trabalhista","Compliance","PJ","Vara do Trabalho RJ"),
    (69, "Tribunal / TSE / Certidão de Quitação Eleitoral",         "Eleitoral",       "Due Diligence","PF",""),
    (70, "Tribunal / TST / Banco de Falências",                     "Judicial",        "Due Diligence","PJ",""),
    (71, "Tribunal / TST / CNDT",                                   "Trabalhista",     "BG Ativo",     "PJ", "Já ativo como Background Check"),
    (72, "Tribunal / TST / Validação de CNDT",                      "Trabalhista",     "BG Ativo",     "PJ", "Já ativo como Background Check"),
    (73, "UNSCCL - UN Security Council Consolidated List (PF)",     "Sanções Internacionais","Due Diligence","PF",""),
]

row = 4
for num, fonte, cat, grupo, tipo, obs in buscas:
    bg_lt, bg_dk, _ = GRUPOS[grupo]
    row_bg = bg_lt
    cell(ws1, row, 1, num, h="center", bg=row_bg, bold=True, color=bg_dk)
    cell(ws1, row, 2, fonte, bg=row_bg, wrap=True)
    cell(ws1, row, 3, cat, h="center", bg=row_bg, size=9)
    cell(ws1, row, 4, grupo, h="center", bg=row_bg, bold=True, color=bg_dk, size=9)
    cell(ws1, row, 5, tipo, h="center", bg=row_bg, size=9)
    cell(ws1, row, 6, obs, bg=row_bg, wrap=True, size=9, italic=bool(obs), color="555555" if obs else None)
    ws1.row_dimensions[row].height = 16
    row += 1

# Legenda
row += 1
ws1.merge_cells(f"A{row}:F{row}")
c = ws1.cell(row=row, column=1, value="Legenda de Grupos de Relevância")
c.font = Font(name="Calibri", bold=True, size=10, color=TEAL_ESC)
c.alignment = Alignment(horizontal="left")
ws1.row_dimensions[row].height = 18
row += 1

for grupo, (bg_lt, bg_dk, desc) in GRUPOS.items():
    ws1.merge_cells(f"B{row}:F{row}")
    c1 = ws1.cell(row=row, column=1)
    c1.fill = fill(bg_lt)
    c1.font = Font(name="Calibri", bold=True, size=9, color=bg_dk)
    c1.value = grupo
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = thin()
    c2 = ws1.cell(row=row, column=2)
    c2.font = Font(name="Calibri", size=9, color="333333")
    c2.value = desc
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = thin()
    ws1.row_dimensions[row].height = 16
    row += 1

# Fundo externo
for r in range(1, row + 30):
    for col in range(7, 30):
        ws1.cell(row=r, column=col).fill = fill(TEAL_XLT)
for r in range(row, row + 30):
    for col in range(1, 7):
        ws1.cell(row=r, column=col).fill = fill(TEAL_XLT)

ws1.freeze_panes = None

# ══════════════════════════════════════════════════════════════════
# ABA 2 — DOCUMENTOS ATIVOS DOCK BRASIL
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Documentos Ativos")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = TEAL_ESC

for col, w in zip("ABCDE", [5, 52, 14, 18, 28]):
    ws2.column_dimensions[col].width = w

add_logo_header(ws2, "Documentos Ativos — Dock Brasil", ncols=5)

header_row(ws2, 3,
    [(1,"Nº"),(2,"Documento"),(3,"Tem Validade?"),(4,"Background Check Ativo"),(5,"Observação")],
    bg=TEAL_ESC)
ws2.row_dimensions[3].height = 20

docs = [
    # (doc, tem_validade, bg_ativo, obs)
    # ── ADMINISTRATIVO ──
    ("── ADMINISTRATIVO ──", "", "", ""),
    ("Contrato Social (Última Alteração)", "Não", "Não", ""),
    ("Contrato Social", "Não", "Não", ""),
    ("Alvará de Localização / Funcionamento", "Não", "Não", ""),
    ("Cartão CNPJ", "Não", "Sim", "Busca automática disponível"),
    ("CNPJ — Destinação de Resíduo (Terceiros)", "Não", "Não", ""),
    ("Documento de Identificação", "Não", "Não", ""),
    ("CPF", "Não", "Não", ""),
    ("Certidão Negativa de Débito Federal (CND)", "Não", "Sim", "Busca automática disponível"),
    ("FGTS + Relatório + Guia + Comprovante", "Não", "Sim", "Busca automática disponível"),
    ("DCTFWEB", "Não", "Não", ""),
    ("Certidão Negativa de Débitos Trabalhistas (CNDT)", "Não", "Sim", "Busca automática disponível"),
    ("Folha de Pagamento + Relatório + Comprovante", "Sim", "Não", ""),
    ("Vínculo de Trabalho", "Não", "Não", ""),
    ("Seguro de Vida (Extrato — Último Mês)", "Sim", "Não", ""),
    ("Vale Alimentação (Extrato — Último Mês)", "Sim", "Não", ""),
    ("Vale Transporte (Extrato — Último Mês)", "Sim", "Não", ""),
    ("Convenção Coletiva de Trabalho (CCT)", "Não", "Não", ""),
    ("Rescisão Contratual", "Não", "Não", ""),
    ("Política de Privacidade", "Não", "Não", ""),
    ("Apresentação Institucional", "Não", "Não", ""),
    ("Comp. de Credenciamento e Aut. do MEC", "Não", "Não", ""),
    ("Projeto Pedagógico do Curso (PPC)", "Não", "Não", ""),
    ("Carteira Nacional de Vigilante", "Sim", "Não", ""),
    ("Certificado de Registro Licenciamento Veicular (CRLV)", "Sim", "Não", ""),
    ("Certificado de Segurança Veicular", "Sim", "Não", ""),
    ("Seguro de Passageiros", "Sim", "Não", ""),
    ("Briefing de Segurança — Dock Brasil", "Sim", "Não", "Documento próprio Dock Brasil"),
    # ── QSMS / SEGURANÇA ──
    ("── QSMS / SEGURANÇA DO TRABALHO ──", "", "", ""),
    ("PGR — Programa de Gerenciamento de Riscos", "Sim", "Não", ""),
    ("PCMSO — Programa de Controle Médico", "Sim", "Não", ""),
    ("PGRSS", "Não", "Não", ""),
    ("LTCAT", "Não", "Não", ""),
    ("ASO — Atestado de Saúde Ocupacional", "Sim", "Não", ""),
    ("NR 06 — EPI", "Sim", "Não", ""),
    ("NR 07 — Primeiros Socorros", "Sim", "Não", ""),
    ("NR 10 — Básico em Eletricidade", "Sim", "Não", ""),
    ("NR 10 — SEP", "Sim", "Não", ""),
    ("NR 11 — Empilhadeira", "Sim", "Não", ""),
    ("NR 11 — Auxiliar de Movimentação de Carga", "Sim", "Não", ""),
    ("NR 12 — Máquinas e Equipamentos", "Sim", "Não", ""),
    ("NR 13 — Teste Hidrostático", "Sim", "Não", ""),
    ("NR 20 — Inflamáveis e Combustíveis", "Sim", "Não", ""),
    ("NR 33 — Espaço Confinado (Vigia)", "Sim", "Não", ""),
    ("NR 33 — Espaço Confinado (Supervisor)", "Sim", "Não", ""),
    ("NR 33 — Espaço Confinado (Resgate)", "Sim", "Não", ""),
    ("NR 34 — Treinamento Admissional", "Sim", "Não", ""),
    ("NR 34 — Admissional / Periódico", "Sim", "Não", ""),
    ("NR 34 — Atividades com Solda", "Sim", "Não", ""),
    ("NR 34 — Atividades com Maçarico", "Sim", "Não", ""),
    ("NR 34 — Atividades com Hidrojateamento", "Sim", "Não", ""),
    ("NR 34 — Andaime", "Sim", "Não", ""),
    ("NR 34 — Máquinas Rotativas", "Sim", "Não", ""),
    ("NR 34 — Observador de Trabalho a Quente", "Sim", "Não", ""),
    ("NR 34 — Segurança em Trabalhos a Quente", "Sim", "Não", ""),
    ("NR 34 — Testes de Estanqueidade", "Sim", "Não", ""),
    ("NR 34 — Pintura", "Sim", "Não", ""),
    ("NR 34 — Rigger", "Sim", "Não", ""),
    ("NR 34 — Segurança Movimentação de Carga", "Sim", "Não", ""),
    ("NR 34 — Operador de Equipamento de Guindar", "Sim", "Não", ""),
    ("NR 34 — Outras Atividades", "Sim", "Não", ""),
    ("NR 35 — Trabalho em Altura", "Sim", "Não", ""),
    ("NR 35 — Resgate", "Sim", "Não", ""),
    ("NR 37 — Atividade em Plataforma de Petróleo", "Sim", "Não", "Alta relevância O&G"),
    ("Operador de Plataforma Elevatória", "Sim", "Não", ""),
    ("Formação de Brigadista", "Sim", "Não", ""),
    ("Decl. de Conhecimento e Conformidade com a NR-15", "Não", "Não", ""),
    ("Capacitação Profissional", "Sim", "Não", ""),
    ("CSSM", "Sim", "Não", ""),
    ("CAFT", "Não", "Não", ""),
    ("CIR", "Sim", "Não", ""),
    ("FCEM", "Não", "Não", ""),
    ("LRM", "Não", "Não", ""),
    ("PMP", "Não", "Não", ""),
    ("Manual de Boas Práticas", "Não", "Não", ""),
    # ── AMBIENTAL ──
    ("── AMBIENTAL ──", "", "", ""),
    ("Cadastro Técnico Federal (IBAMA)", "Sim", "Não", "Busca automática disponível"),
    ("Licença de Operação Ambiental", "Sim", "Não", ""),
    ("Licença Ambiental (Destinação de Resíduos)", "Sim", "Não", ""),
    ("Outorga dos Dir. de Uso de Recursos Hídricos", "Sim", "Não", ""),
    ("Autorização Ambiental para Transp. de Produtos", "Sim", "Não", ""),
    ("Certificado Ambiental", "Sim", "Não", ""),
    ("Laudo de Ensaio Opacímetro", "Não", "Não", ""),
    ("Controle de Pragas (Cozinha) — Trimestral", "Sim", "Não", ""),
    ("Controle de Pragas (Veículos) — Trimestral", "Sim", "Não", ""),
    ("Análise de Qualidade da Água (Potabilidade)", "Não", "Não", ""),
    # ── CERTIFICAÇÕES / HABILITAÇÕES ──
    ("── CERTIFICAÇÕES / HABILITAÇÕES ──", "", "", ""),
    ("ISO 9001:2015", "Sim", "Não", ""),
    ("Certificado IRATA", "Sim", "Não", ""),
    ("Certificado de Acreditação", "Sim", "Não", ""),
    ("CCL — Cert. de Credenciamento do Laboratório", "Sim", "Não", ""),
    ("Cert. de Manut. de Condições Operacionais", "Não", "Não", ""),
    ("Certificado do INMETRO", "Não", "Não", ""),
    ("Aut. de Funcionamento de Empresa (AFE)", "Não", "Não", ""),
    ("CNES", "Sim", "Não", ""),
    ("Cert. de Regularidade de Segurança (CRS)", "Sim", "Não", ""),
    ("Procedimento Operacional Padrão (POP)", "Não", "Não", ""),
    ("Registro do Responsável Técnico", "Não", "Não", ""),
    ("Título de Inscrição da Embarcação", "Sim", "Não", "Específico setor naval"),
    ("CRV — Cert. de Registros e Vetores", "Sim", "Não", ""),
]

row = 4
num = 1
for item in docs:
    doc, validade, bg_ativo, obs = item
    is_section = doc.startswith("──")

    if is_section:
        ws2.merge_cells(f"A{row}:E{row}")
        c = ws2.cell(row=row, column=1, value=doc)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = fill(TEAL)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin(TEAL_ESC)
        ws2.row_dimensions[row].height = 18
        row += 1
        continue

    row_bg = GRAY_LT if num % 2 == 0 else WHITE
    if bg_ativo == "Sim":
        row_bg = GREEN_LT

    cell(ws2, row, 1, num, h="center", bg=row_bg, bold=True)
    cell(ws2, row, 2, doc, bg=row_bg, wrap=True)

    c_val = ws2.cell(row=row, column=3, value=validade)
    c_val.font = Font(name="Calibri", size=10,
                      color=GREEN_DK if validade=="Sim" else "888888", bold=validade=="Sim")
    c_val.fill = fill(row_bg)
    c_val.alignment = Alignment(horizontal="center", vertical="center")
    c_val.border = thin()

    c_bg = ws2.cell(row=row, column=4, value=bg_ativo)
    c_bg.font = Font(name="Calibri", size=10, bold=bg_ativo=="Sim",
                     color=GREEN_DK if bg_ativo=="Sim" else "888888")
    c_bg.fill = fill(row_bg)
    c_bg.alignment = Alignment(horizontal="center", vertical="center")
    c_bg.border = thin()

    cell(ws2, row, 5, obs, bg=row_bg, size=9, italic=bool(obs), color="555555" if obs else None)
    ws2.row_dimensions[row].height = 16
    num += 1
    row += 1

# Fundo externo aba 2
for r in range(1, row + 20):
    for col in range(6, 25):
        ws2.cell(row=r, column=col).fill = fill(TEAL_XLT)
for r in range(row, row + 20):
    for col in range(1, 6):
        ws2.cell(row=r, column=col).fill = fill(TEAL_XLT)

ws2.freeze_panes = None

wb.save(OUTPUT)
print(f"Arquivo gerado: {OUTPUT}")
print(f"  Aba 1: {len(buscas)} buscas automáticas")
print(f"  Aba 2: {num-1} documentos ativos")
