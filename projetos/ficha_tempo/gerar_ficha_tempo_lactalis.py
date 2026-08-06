#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = r"c:\Users\gabriel.evangelista\Documents\ClaudeGL\clientes\Lactalis\ficha_tempo_lactalis.xlsx"
LOGO_PATH   = r"c:\Users\gabriel.evangelista\Documents\ClaudeGL\Documentos\logos_extraidas\slide1_img4.png"

# ── Paleta oficial Efcaz ───────────────────────────────────────────────────
TEAL        = "0E8FA3"
TEAL_CLARO  = "14B3CC"
TEAL_ESCURO = "0A6A7A"
CINZA_CLARO = "F2F4F4"
VERDE       = "27AE60"
VERMELHO    = "E74C3C"
LARANJA     = "F39C12"
PRETO_SUV   = "1A2A2A"

# ── Equipe ────────────────────────────────────────────────────────────────
EQUIPE = [
    "Thais Jayne Biscaia",
    "Janaina Ventura",
    "Andriela Klai Fernandes",
    "Ricardo Pedroso da Silva",
    "Gabriel Vital",
]

FASES = [
    "Fase 1 — Setup e Extração",
    "Fase 2 — Higienização e Categorização",
    "Fase 3 — Risco e Criticidade",
    "Fase 4 — Governança e Processos",
    "Fase 5 — Entrega e Prontidão SAP",
]

TIPOS = [
    "Reunião",
    "Análise",
    "Configuração",
    "Documentação",
    "Revisão",
    "Apresentação",
]

# ── Cores por fase ─────────────────────────────────────────────────────────
FASE_CORES = {
    "Fase 1": {"txt": TEAL,        "bg": "D5EFF2"},
    "Fase 2": {"txt": VERDE,       "bg": "D5F0E0"},
    "Fase 3": {"txt": LARANJA,     "bg": "FDE9C9"},
    "Fase 4": {"txt": TEAL_ESCURO, "bg": "C8E2E6"},
    "Fase 5": {"txt": TEAL_CLARO,  "bg": "D5F2F8"},
}

# ── Helpers ────────────────────────────────────────────────────────────────
def fill(c):    return PatternFill("solid", fgColor=c)
def fnt(bold=False, color="1F1F1F", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def border():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)
def border_accent(color="0E8FA3"):
    return Border(left=Side(style="medium", color=color),
                  right=Side(style="thin", color="BFBFBF"),
                  top=Side(style="thin", color="BFBFBF"),
                  bottom=Side(style="thin", color="BFBFBF"))

TASK_ROW = 3   # linha onde as sessões começam
MAX_ROW  = 200 # suporte a muitas sessões
N_COLS   = 7   # A:G

def section_header(ws, row, text, n_cols=7, h=26):
    lc = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{lc}{row}")
    ws[f"A{row}"] = text
    ws[f"A{row}"].fill = fill(TEAL)
    ws[f"A{row}"].font = fnt(bold=True, color="FFFFFF", size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = h

# ══════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════
# ABA LISTAS (oculta — fonte dos dropdowns)
# ══════════════════════════════════════════════════════════════════════════
ws_l = wb.active
ws_l.title = "Listas"

headers_l = [("A", "Responsável", EQUIPE),
             ("C", "Fase",        FASES),
             ("E", "Tipo",        TIPOS)]

for col, title, items in headers_l:
    ws_l[f"{col}1"] = title
    ws_l[f"{col}1"].font = fnt(bold=True, color="FFFFFF")
    ws_l[f"{col}1"].fill = fill(TEAL_ESCURO)
    for i, v in enumerate(items, 2):
        ws_l.cell(row=i, column=ord(col) - 64, value=v)
    ws_l.column_dimensions[col].width = 40

ws_l.sheet_state = "hidden"

# ══════════════════════════════════════════════════════════════════════════
# ABA FICHA DE TEMPO
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Ficha de Tempo")
ws.sheet_view.showGridLines = False

# ── Linha 1: Título ────────────────────────────────────────────────────────
ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
ws["A1"] = "FICHA DE TEMPO  ·  BPO & CONSULTORIA DE SANEAMENTO DE BASE  ·  LACTALIS"
ws["A1"].fill = fill(TEAL_ESCURO)
ws["A1"].font = fnt(bold=True, color="FFFFFF", size=14)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 48

if os.path.exists(LOGO_PATH):
    logo = XLImage(LOGO_PATH)
    logo.width  = 130
    logo.height = 43
    ws.add_image(logo, "A1")

# ── Linha 2: Cabeçalhos ────────────────────────────────────────────────────
HEADERS = ["Data", "Responsável", "Fase", "Atividade / Entregável", "Tipo", "Horas", "Observação"]
COL_DATA  = 1
COL_RESP  = 2
COL_FASE  = 3
COL_ATIV  = 4
COL_TIPO  = 5
COL_HORAS = 6
COL_OBS   = 7

for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.fill = fill(TEAL)
    cell.font = fnt(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()
ws.row_dimensions[2].height = 32

# ── Linhas de dados (TASK_ROW … MAX_ROW) ──────────────────────────────────
for r in range(TASK_ROW, MAX_ROW + 1):
    row_bg = "FAFAFA" if r % 2 == 0 else "FFFFFF"

    # Data
    dc = ws.cell(row=r, column=COL_DATA)
    dc.number_format = "DD/MM/YYYY"
    dc.fill = fill(row_bg)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    dc.border = border_accent(TEAL)

    # Responsável, Fase, Tipo — dropdowns
    for col in (COL_RESP, COL_FASE, COL_TIPO):
        c = ws.cell(row=r, column=col)
        c.fill = fill(row_bg)
        c.font = fnt(size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border()

    # Atividade / Entregável
    ac = ws.cell(row=r, column=COL_ATIV)
    ac.fill = fill(row_bg)
    ac.font = fnt(size=9)
    ac.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ac.border = border()

    # Horas — formato numérico 1 decimal
    hc = ws.cell(row=r, column=COL_HORAS)
    hc.number_format = '0.0"h"'
    hc.fill = fill("EBF7F9" if r % 2 == 0 else "FFFFFF")
    hc.font = fnt(bold=True, color=TEAL_ESCURO, size=10)
    hc.alignment = Alignment(horizontal="center", vertical="center")
    hc.border = border()

    # Observação
    oc = ws.cell(row=r, column=COL_OBS)
    oc.fill = fill(row_bg)
    oc.font = fnt(size=9, italic=True, color="595959")
    oc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    oc.border = border()

    ws.row_dimensions[r].height = 22

# ── CF — coloração automática de linha por fase ───────────────────────────
row_cf_rng = f"A{TASK_ROW}:{get_column_letter(N_COLS)}{MAX_ROW}"
for prefix, cores in reversed(list(FASE_CORES.items())):
    ws.conditional_formatting.add(row_cf_rng, FormulaRule(
        formula=[f'LEFT($C{TASK_ROW},6)="{prefix}"'],
        fill=PatternFill("solid", fgColor=cores["bg"])
    ))

# CF — coluna Horas: escala de cor (0h → vermelho; mais h → verde)
ws.conditional_formatting.add(
    f"F{TASK_ROW}:F{MAX_ROW}",
    ColorScaleRule(
        start_type="num", start_value=0,   start_color="F8696B",
        mid_type="num",   mid_value=4,     mid_color="FFEB84",
        end_type="num",   end_value=8,     end_color="63BE7B"
    )
)

# ── CF — tipo: colore célula de Tipo por categoria ────────────────────────
tipo_rng = f"E{TASK_ROW}:E{MAX_ROW}"
tipo_cores = {
    "Reunião":       ("FFE0B2", "E65100"),
    "Análise":       ("E3F2FD", "0D47A1"),
    "Configuração":  ("E8F5E9", "1B5E20"),
    "Documentação":  ("F3E5F5", "4A148C"),
    "Revisão":       ("FFF9C4", "827717"),
    "Apresentação":  ("FCE4EC", "880E4F"),
}
for tipo_val, (bg, fg) in tipo_cores.items():
    ws.conditional_formatting.add(tipo_rng, FormulaRule(
        formula=[f'$E{TASK_ROW}="{tipo_val}"'],
        fill=PatternFill("solid", fgColor=bg),
        font=Font(name="Calibri", bold=True, color=fg, size=9)
    ))

# ── Dropdowns ─────────────────────────────────────────────────────────────
dv_resp = DataValidation(
    type="list",
    formula1=f"Listas!$A$2:$A${1+len(EQUIPE)}",
    allow_blank=True, showDropDown=False,
    showErrorMessage=True, errorTitle="Nome inválido",
    error="Selecione um responsável da lista.",
    showInputMessage=True, promptTitle="Responsável",
    prompt="Selecione o membro da equipe responsável pela sessão."
)
ws.add_data_validation(dv_resp)
dv_resp.sqref = f"B{TASK_ROW}:B{MAX_ROW}"

dv_fase = DataValidation(
    type="list",
    formula1=f"Listas!$C$2:$C${1+len(FASES)}",
    allow_blank=True, showDropDown=False,
    showErrorMessage=True, errorTitle="Fase inválida",
    error="Selecione a fase do projeto.",
    showInputMessage=True, promptTitle="Fase",
    prompt="Selecione a fase a que esta sessão pertence."
)
ws.add_data_validation(dv_fase)
dv_fase.sqref = f"C{TASK_ROW}:C{MAX_ROW}"

dv_tipo = DataValidation(
    type="list",
    formula1=f"Listas!$E$2:$E${1+len(TIPOS)}",
    allow_blank=True, showDropDown=False,
    showErrorMessage=True, errorTitle="Tipo inválido",
    error="Selecione o tipo de atividade.",
    showInputMessage=True, promptTitle="Tipo de atividade",
    prompt="Reunião, Análise, Configuração, Documentação, Revisão ou Apresentação."
)
ws.add_data_validation(dv_tipo)
dv_tipo.sqref = f"E{TASK_ROW}:E{MAX_ROW}"

dv_data = DataValidation(
    type="date",
    operator="between",
    formula1="DATE(2025,1,1)", formula2="DATE(2030,12,31)",
    allow_blank=True, showDropDown=False,
    showErrorMessage=True, errorTitle="Data inválida",
    error="Digite a data no formato DD/MM/AAAA.",
    showInputMessage=True, promptTitle="Data da sessão",
    prompt="Digite a data no formato DD/MM/AAAA."
)
ws.add_data_validation(dv_data)
dv_data.sqref = f"A{TASK_ROW}:A{MAX_ROW}"

# ── Linha de Total ─────────────────────────────────────────────────────────
TOTAL_ROW = MAX_ROW + 1
ws.merge_cells(f"A{TOTAL_ROW}:E{TOTAL_ROW}")
tc_label = ws[f"A{TOTAL_ROW}"]
tc_label.value = "TOTAL GERAL DE HORAS"
tc_label.fill = fill(TEAL_ESCURO)
tc_label.font = fnt(bold=True, color="FFFFFF", size=11)
tc_label.alignment = Alignment(horizontal="right", vertical="center")
tc_label.border = border()
ws.row_dimensions[TOTAL_ROW].height = 28

tc_val = ws.cell(row=TOTAL_ROW, column=COL_HORAS,
                 value=f"=SUM(F{TASK_ROW}:F{MAX_ROW})")
tc_val.number_format = '0.0"h"'
tc_val.fill = fill(TEAL_ESCURO)
tc_val.font = fnt(bold=True, color="FFFFFF", size=13)
tc_val.alignment = Alignment(horizontal="center", vertical="center")
tc_val.border = border()

ws.cell(row=TOTAL_ROW, column=COL_OBS).fill = fill(TEAL_ESCURO)
ws.cell(row=TOTAL_ROW, column=COL_OBS).border = border()

# ── Largura das colunas ────────────────────────────────────────────────────
col_widths = {1: 13, 2: 24, 3: 34, 4: 48, 5: 15, 6: 9, 7: 30}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = f"A{TASK_ROW}"

# ══════════════════════════════════════════════════════════════════════════
# ABA PAINEL DE HORAS
# ══════════════════════════════════════════════════════════════════════════
ws_p = wb.create_sheet("Painel de Horas", 0)
ws_p.sheet_view.showGridLines = False
N_P = 7  # colunas A:G

def p_cell(ws, row, col, val=None, bg="FFFFFF", bold=False, color="1A2A2A",
           size=10, align="center", fmt=None, italic=False, wrap=False, span=1):
    if span > 1:
        end = get_column_letter(col + span - 1)
        ws.merge_cells(f"{get_column_letter(col)}{row}:{end}{row}")
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg)
    c.font = fnt(bold=bold, color=color, size=size, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = border()
    if fmt:
        c.number_format = fmt
    return c

# ── Logo + Título ──────────────────────────────────────────────────────────
if os.path.exists(LOGO_PATH):
    logo_p = XLImage(LOGO_PATH)
    logo_p.width  = 150
    logo_p.height = 50
    ws_p.add_image(logo_p, f"{get_column_letter(N_P - 1)}1")

ws_p.merge_cells(f"A1:{get_column_letter(N_P - 2)}1")
ws_p["A1"] = "PAINEL DE HORAS DO PROJETO"
ws_p["A1"].fill = fill(TEAL_ESCURO)
ws_p["A1"].font = fnt(bold=True, color="FFFFFF", size=18)
ws_p["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_p.row_dimensions[1].height = 58
for c_idx in range(N_P - 1, N_P + 1):
    ws_p.cell(row=1, column=c_idx).fill = fill("FFFFFF")

ws_p.merge_cells(f"A2:{get_column_letter(N_P)}2")
ws_p["A2"] = "BPO & Consultoria de Saneamento de Base  ·  Lactalis  ·  25.000 → ~15.000 fornecedores estratégicos"
ws_p["A2"].fill = fill(TEAL)
ws_p["A2"].font = fnt(italic=True, color="FFFFFF", size=11)
ws_p["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_p.row_dimensions[2].height = 28
ws_p.row_dimensions[3].height = 10

# ── Total geral ────────────────────────────────────────────────────────────
section_header(ws_p, 4, "TOTAL GERAL DE HORAS TRABALHADAS", N_P)

FICHA_SUM = f"='Ficha de Tempo'!F{TASK_ROW}:F{MAX_ROW}"

p_cell(ws_p, 5, 1, "Total de Horas:", bg="EBF7F9", bold=True, color=PRETO_SUV, align="right")
p_cell(ws_p, 5, 2,
       f"=IFERROR(SUM('Ficha de Tempo'!$F${TASK_ROW}:$F${MAX_ROW}),0)",
       fmt='0.0"h"', bold=True, color=TEAL_ESCURO, size=18, span=2)
p_cell(ws_p, 5, 4, "Nº de Sessões:", bg="EBF7F9", bold=True, color=PRETO_SUV, align="right")
p_cell(ws_p, 5, 5,
       f"=IFERROR(COUNTA('Ficha de Tempo'!$A${TASK_ROW}:$A${MAX_ROW}),0)",
       bold=True, color=TEAL_ESCURO, size=16, span=3)
ws_p.row_dimensions[5].height = 34
ws_p.row_dimensions[6].height = 10

# ── Horas por responsável ──────────────────────────────────────────────────
section_header(ws_p, 7, "HORAS POR RESPONSÁVEL", N_P)

hdr_cols = ["Responsável", "", "Horas", "Barra de Esforço", "", "", ""]
for c, h in enumerate(hdr_cols, 1):
    cell = ws_p.cell(row=8, column=c, value=h)
    cell.fill = fill(TEAL_ESCURO)
    cell.font = fnt(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()
ws_p.row_dimensions[8].height = 22

EQUIPE_CORES = ["0E8FA3", "27AE60", "F39C12", "0A6A7A", "14B3CC"]
for idx, (membro, cor) in enumerate(zip(EQUIPE, EQUIPE_CORES)):
    r = 9 + idx
    ws_p.merge_cells(f"A{r}:B{r}")
    nc = ws_p.cell(row=r, column=1, value=membro)
    nc.font = fnt(bold=True, color=cor, size=10)
    nc.fill = fill("F8F9FA")
    nc.alignment = Alignment(horizontal="left", vertical="center")
    nc.border = border()
    ws_p.cell(row=r, column=2).fill = fill(cor)
    ws_p.cell(row=r, column=2).border = border()

    hc = ws_p.cell(row=r, column=3,
                   value=f'=IFERROR(SUMIF(\'Ficha de Tempo\'!$B${TASK_ROW}:$B${MAX_ROW},"{membro}",\'Ficha de Tempo\'!$F${TASK_ROW}:$F${MAX_ROW}),0)')
    hc.number_format = '0.0"h"'
    hc.font = fnt(bold=True, color=cor, size=12)
    hc.fill = fill("FFFFFF")
    hc.alignment = Alignment(horizontal="center", vertical="center")
    hc.border = border()

    ws_p.merge_cells(f"D{r}:{get_column_letter(N_P)}{r}")
    bc = ws_p.cell(row=r, column=4)
    # Barra proporcional ao total geral (escala: cada bloco = 5% do total)
    bc.value = (
        f'=IFERROR(REPT("█",MIN(20,ROUND(C{r}/'
        f'MAX(SUM(\'Ficha de Tempo\'!$F${TASK_ROW}:$F${MAX_ROW}),1)*20,0)))&'
        f'REPT("░",20-MIN(20,ROUND(C{r}/'
        f'MAX(SUM(\'Ficha de Tempo\'!$F${TASK_ROW}:$F${MAX_ROW}),1)*20,0))),"")'
    )
    bc.font = Font(name="Courier New", bold=True, color=cor, size=11)
    bc.fill = fill("F8F9FA")
    bc.alignment = Alignment(horizontal="left", vertical="center")
    bc.border = border()
    ws_p.row_dimensions[r].height = 24

ws_p.row_dimensions[9 + len(EQUIPE)].height = 10

# ── Horas por fase ─────────────────────────────────────────────────────────
SEP_ROW = 9 + len(EQUIPE) + 1
section_header(ws_p, SEP_ROW, "HORAS POR FASE", N_P)

hdr2_r = SEP_ROW + 1
for c, h in enumerate(["Fase", "", "Horas", "Barra de Esforço", "", "", ""], 1):
    cell = ws_p.cell(row=hdr2_r, column=c, value=h)
    cell.fill = fill(TEAL_ESCURO)
    cell.font = fnt(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()
ws_p.row_dimensions[hdr2_r].height = 22

FASE_DETALHES = [
    ("Fase 1 — Setup e Extração",            TEAL,        "Fase 1"),
    ("Fase 2 — Higienização e Categorização", VERDE,       "Fase 2"),
    ("Fase 3 — Risco e Criticidade",          LARANJA,     "Fase 3"),
    ("Fase 4 — Governança e Processos",       TEAL_ESCURO, "Fase 4"),
    ("Fase 5 — Entrega e Prontidão SAP",      TEAL_CLARO,  "Fase 5"),
]

for idx, (fase_nome, cor, prefix) in enumerate(FASE_DETALHES):
    r = hdr2_r + 1 + idx

    ws_p.merge_cells(f"A{r}:B{r}")
    fc = ws_p.cell(row=r, column=1, value=fase_nome)
    fc.font = fnt(bold=True, color=cor, size=10)
    fc.fill = fill("F8F9FA")
    fc.alignment = Alignment(horizontal="left", vertical="center")
    fc.border = border()
    ws_p.cell(row=r, column=2).fill = fill(cor)
    ws_p.cell(row=r, column=2).border = border()

    fh = ws_p.cell(row=r, column=3,
                   value=f'=IFERROR(SUMIF(\'Ficha de Tempo\'!$C${TASK_ROW}:$C${MAX_ROW},"{prefix}*",\'Ficha de Tempo\'!$F${TASK_ROW}:$F${MAX_ROW}),0)')
    fh.number_format = '0.0"h"'
    fh.font = fnt(bold=True, color=cor, size=12)
    fh.fill = fill("FFFFFF")
    fh.alignment = Alignment(horizontal="center", vertical="center")
    fh.border = border()

    ws_p.merge_cells(f"D{r}:{get_column_letter(N_P)}{r}")
    fb = ws_p.cell(row=r, column=4)
    fb.value = (
        f'=IFERROR(REPT("█",MIN(20,ROUND(C{r}/'
        f'MAX(SUM(\'Ficha de Tempo\'!$F${TASK_ROW}:$F${MAX_ROW}),1)*20,0)))&'
        f'REPT("░",20-MIN(20,ROUND(C{r}/'
        f'MAX(SUM(\'Ficha de Tempo\'!$F${TASK_ROW}:$F${MAX_ROW}),1)*20,0))),"")'
    )
    fb.font = Font(name="Courier New", bold=True, color=cor, size=11)
    fb.fill = fill("F8F9FA")
    fb.alignment = Alignment(horizontal="left", vertical="center")
    fb.border = border()
    ws_p.row_dimensions[r].height = 24

# ── Horas por tipo ─────────────────────────────────────────────────────────
TIPO_BASE_ROW = hdr2_r + 1 + len(FASE_DETALHES) + 1
section_header(ws_p, TIPO_BASE_ROW, "HORAS POR TIPO DE ATIVIDADE", N_P)

TIPO_CORES_P = {
    "Reunião":       ("FF9800", "FFF3E0"),
    "Análise":       ("1565C0", "E3F2FD"),
    "Configuração":  ("2E7D32", "E8F5E9"),
    "Documentação":  ("6A1B9A", "F3E5F5"),
    "Revisão":       ("F9A825", "FFFDE7"),
    "Apresentação":  ("AD1457", "FCE4EC"),
}

for idx, tipo in enumerate(TIPOS):
    r = TIPO_BASE_ROW + 1 + idx
    cor_txt, cor_bg = TIPO_CORES_P[tipo]

    ws_p.merge_cells(f"A{r}:B{r}")
    tc2 = ws_p.cell(row=r, column=1, value=tipo)
    tc2.font = fnt(bold=True, color=cor_txt, size=10)
    tc2.fill = fill(cor_bg)
    tc2.alignment = Alignment(horizontal="left", vertical="center")
    tc2.border = border()
    ws_p.cell(row=r, column=2).fill = fill(cor_txt)
    ws_p.cell(row=r, column=2).border = border()

    th = ws_p.cell(row=r, column=3,
                   value=f'=IFERROR(SUMIF(\'Ficha de Tempo\'!$E${TASK_ROW}:$E${MAX_ROW},"{tipo}",\'Ficha de Tempo\'!$F${TASK_ROW}:$F${MAX_ROW}),0)')
    th.number_format = '0.0"h"'
    th.font = fnt(bold=True, color=cor_txt, size=12)
    th.fill = fill("FFFFFF")
    th.alignment = Alignment(horizontal="center", vertical="center")
    th.border = border()

    ws_p.merge_cells(f"D{r}:{get_column_letter(N_P)}{r}")
    tb = ws_p.cell(row=r, column=4)
    tb.value = (
        f'=IFERROR(REPT("█",MIN(20,ROUND(C{r}/'
        f'MAX(SUM(\'Ficha de Tempo\'!$F${TASK_ROW}:$F${MAX_ROW}),1)*20,0)))&'
        f'REPT("░",20-MIN(20,ROUND(C{r}/'
        f'MAX(SUM(\'Ficha de Tempo\'!$F${TASK_ROW}:$F${MAX_ROW}),1)*20,0))),"")'
    )
    tb.font = Font(name="Courier New", bold=True, color=cor_txt, size=11)
    tb.fill = fill(cor_bg)
    tb.alignment = Alignment(horizontal="left", vertical="center")
    tb.border = border()
    ws_p.row_dimensions[r].height = 24

# ── Nota de rodapé ─────────────────────────────────────────────────────────
note_row = TIPO_BASE_ROW + 1 + len(TIPOS) + 1
ws_p.row_dimensions[note_row - 1].height = 10
ws_p.merge_cells(f"A{note_row}:{get_column_letter(N_P)}{note_row}")
nc = ws_p[f"A{note_row}"]
nc.value = "Registre as sessões diretamente na aba 'Ficha de Tempo'. Este painel atualiza automaticamente ao salvar."
nc.font = fnt(italic=True, color="7F7F7F", size=9)
nc.alignment = Alignment(horizontal="left", vertical="center")
nc.border = border()
ws_p.row_dimensions[note_row].height = 22

# ── Largura das colunas Painel ─────────────────────────────────────────────
painel_widths = {1: 32, 2: 4, 3: 9, 4: 20, 5: 12, 6: 14, 7: 14}
for c, w in painel_widths.items():
    ws_p.column_dimensions[get_column_letter(c)].width = w

# ── Painel como aba ativa ─────────────────────────────────────────────────
wb.active = ws_p

# ══════════════════════════════════════════════════════════════════════════
# SALVAR
# ══════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT_PATH)
print(f"OK — Ficha de tempo gerada: {OUTPUT_PATH}")
