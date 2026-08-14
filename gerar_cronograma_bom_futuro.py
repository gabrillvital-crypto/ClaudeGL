#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = r"c:\Users\gabriel.evangelista\Documents\ClaudeGL\clientes\bom_futuro\cronograma_consultoria_bom_futuro_11-08-2026.xlsx"
OUTPUT_XLSM = r"c:\Users\gabriel.evangelista\Documents\ClaudeGL\clientes\bom_futuro\cronograma_consultoria_bom_futuro_11-08-2026.xlsm"
LOGO_PATH   = r"c:\Users\gabriel.evangelista\Documents\ClaudeGL\Documentos\logos_extraidas\slide1_img4.png"

# ── Paleta oficial Efcaz ───────────────────────────────────────────────────
TEAL        = "0E8FA3"
TEAL_CLARO  = "14B3CC"
TEAL_ESCURO = "0A6A7A"
VERDE       = "27AE60"
LARANJA     = "F39C12"

# ── Fases ─────────────────────────────────────────────────────────────────
PHASES = {
    "Fase 1": {"name": "Fase 1 — Diagnóstico e Alinhamento",      "color": TEAL,        "bg": "D5EFF2"},
    "Fase 2": {"name": "Fase 2 — Estruturação Documental",        "color": VERDE,       "bg": "D5F0E0"},
    "Fase 3": {"name": "Fase 3 — Parametrização na Plataforma",   "color": LARANJA,     "bg": "FDE9C9"},
    "Fase 4": {"name": "Fase 4 — Governança e Documentação",      "color": TEAL_ESCURO, "bg": "C8E2E6"},
    "Fase 5": {"name": "Fase 5 — Entrega e Consolidação",         "color": TEAL_CLARO,  "bg": "D5F2F8"},
}

TASKS = [
    # id  fase_key   atividade                                                                        d_ini d_fim tipo
    (1,  "Fase 1", "Kick-off da Consultoria — Reunião de Abertura com Cairo e Patrícia",               1,    1,   "Reunião"),
    (2,  "Fase 1", "Levantamento da estrutura atual — 115 linhas de fornecimento e volumes por ramo",  2,    5,   "Tarefa"),
    (3,  "Fase 1", "Identificação de gaps críticos: documentos ausentes, vencidos e sem periodicidade",4,    8,   "Tarefa"),
    (4,  "Fase 2", "Definição da matriz de documentos por linha de fornecimento (Crítico / Obrigatório)",9,  14,  "Tarefa"),
    (5,  "Fase 2", "Mapeamento de documentos automáticos (Efcaz) vs manuais (cliente) por categoria",  12,  16,  "Tarefa"),
    (6,  "Fase 2", "Marco 1 — Apresentação da Matriz Documental | Reunião de Validação com Cairo",     17,  17,  "Marco"),
    (7,  "Fase 3", "Configuração das linhas de fornecimento e documentos na plataforma Efcaz",         18,  22,  "Tarefa"),
    (8,  "Fase 3", "Ativação e verificação das buscas automáticas por categoria de fornecedor",        20,  24,  "Tarefa"),
    (9,  "Fase 3", "Validação operacional com Patrícia — testes, ajustes e aprovação na plataforma",   23,  25,  "Tarefa"),
    (10, "Fase 4", "Elaboração do Guia de Periodicidade e Renovação documental por linha",             26,  30,  "Tarefa"),
    (11, "Fase 4", "Planilha Mestre de Documentos — base de referência permanente para o time",        28,  32,  "Tarefa"),
    (12, "Fase 4", "Marco 2 — Revisão de Governança e Alinhamento Executivo | Reunião com Cairo",      33,  33,  "Marco"),
    (13, "Fase 5", "Ajustes finais baseados no feedback consolidado do time Bom Futuro",               34,  37,  "Tarefa"),
    (14, "Fase 5", "Manual de Orientação Estratégica — referência de homologação por ramo de atividade",35, 38,  "Tarefa"),
    (15, "Fase 5", "Marco Final — Apresentação Executiva, Aceite e Próximos Passos | Reunião com Cairo",40, 40,  "Marco"),
]

# ── Layout ─────────────────────────────────────────────────────────────────
COL_ID      = 1   # A
COL_FASE    = 2   # B
COL_ATIV    = 3   # C
COL_RESP    = 4   # D
COL_INI     = 5   # E
COL_FIM     = 6   # F
COL_DUR     = 7   # G
COL_TIPO    = 8   # H
COL_STATUS  = 9   # I
COL_PCT     = 10  # J
TASK_ROW    = 3

STATUS_OPTIONS = ["Não Iniciado", "Em Andamento", "Pausado", "Concluído"]
EQUIPE = [
    "Gabriel Vital",
    "Thais Jayne Biscaia",
    "Ricardo Pedroso da Silva",
    "Todos",
]

resp_map = {
    1:  "Gabriel Vital",
    2:  "Thais Jayne Biscaia",
    3:  "Thais Jayne Biscaia",
    4:  "Thais Jayne Biscaia",
    5:  "Gabriel Vital",
    6:  "Todos",
    7:  "Ricardo Pedroso da Silva",
    8:  "Ricardo Pedroso da Silva",
    9:  "Thais Jayne Biscaia",
    10: "Thais Jayne Biscaia",
    11: "Gabriel Vital",
    12: "Todos",
    13: "Thais Jayne Biscaia",
    14: "Gabriel Vital",
    15: "Todos",
}

# ── Helpers ────────────────────────────────────────────────────────────────
def fill(c):    return PatternFill("solid", fgColor=c)
def fnt(bold=False, color="1F1F1F", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def border():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)
def border_left_accent(color):
    return Border(left=Side(style="medium", color=color),
                  right=Side(style="thin", color="BFBFBF"),
                  top=Side(style="thin", color="BFBFBF"),
                  bottom=Side(style="thin", color="BFBFBF"))

# ══════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── ABA LISTAS (oculta) ───────────────────────────────────────────────────
ws_listas = wb.active
ws_listas.title = "Listas"

ws_listas["A1"] = "Responsável"
ws_listas["A1"].font = fnt(bold=True, color="FFFFFF")
ws_listas["A1"].fill = fill("0A6A7A")
for i, nome in enumerate(EQUIPE, 2):
    ws_listas.cell(row=i, column=1, value=nome)
ws_listas.column_dimensions["A"].width = 30

ws_listas["C1"] = "Status"
ws_listas["C1"].font = fnt(bold=True, color="FFFFFF")
ws_listas["C1"].fill = fill("0A6A7A")
for i, s in enumerate(STATUS_OPTIONS, 2):
    ws_listas.cell(row=i, column=3, value=s)
ws_listas.column_dimensions["C"].width = 18

FASE_NAMES = [p["name"] for p in PHASES.values()]
ws_listas["E1"] = "Fase"
ws_listas["E1"].font = fnt(bold=True, color="FFFFFF")
ws_listas["E1"].fill = fill("0A6A7A")
for i, fn in enumerate(FASE_NAMES, 2):
    ws_listas.cell(row=i, column=5, value=fn)
ws_listas.column_dimensions["E"].width = 44
ws_listas.sheet_state = "hidden"

# ══════════════════════════════════════════════════════════════════════════
# ABA CRONOGRAMA
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Cronograma")

# Linha 1 — Título
ws.merge_cells("A1:J1")
ws["A1"] = "CRONOGRAMA DE PROJETO  ·  CONSULTORIA DE ESTRUTURAÇÃO DOCUMENTAL  ·  BOM FUTURO AGRÍCOLA"
ws["A1"].fill = fill("0A6A7A")
ws["A1"].font = fnt(bold=True, color="FFFFFF", size=13)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 48

if os.path.exists(LOGO_PATH):
    logo_crono = XLImage(LOGO_PATH)
    logo_crono.width  = 130
    logo_crono.height = 43
    ws.add_image(logo_crono, "A1")

# Linha 2 — Cabeçalhos
headers = ["ID", "Fase", "Atividade", "Responsável",
           "Data\nInício", "Data\nFim", "Dur.\n(dias)", "Tipo", "Status", "% Concl."]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.fill = fill("0E8FA3")
    cell.font = fnt(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()
ws.row_dimensions[2].height = 32

# Tarefas
for i, (tid, fase_key, ativ, d_ini, d_fim, tipo) in enumerate(TASKS):
    r = TASK_ROW + i
    phase = PHASES[fase_key]
    is_marco = tipo in ("Marco", "Reunião")
    row_bg = "FFF3CD" if is_marco else phase["bg"]
    bar_color = "FF9900" if is_marco else phase["color"]
    resp = resp_map.get(tid, "Gabriel Vital")

    def set_cell(col, val, bold=False, color="1F1F1F", bg=row_bg, align="left", fmt=None, sz=10, wrap=False):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = fnt(bold=bold, color=color, size=sz)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cell.border = border_left_accent(phase["color"]) if col == COL_ID else border()
        if fmt:
            cell.number_format = fmt
        return cell

    # ID (fórmula auto-sequência)
    id_cell = ws.cell(row=r, column=COL_ID, value=f"=IF(C{r}<>\"\",COUNTA($C${TASK_ROW}:C{r}),\"\")")
    id_cell.font = fnt(bold=True, color="FFFFFF", size=10)
    id_cell.fill = fill(bar_color)
    id_cell.alignment = Alignment(horizontal="center", vertical="center")
    id_cell.border = border()

    set_cell(COL_FASE, phase["name"], bold=False, color=phase["color"], sz=9, wrap=True)
    set_cell(COL_ATIV, ativ, bold=is_marco, color="7F3F00" if is_marco else "1C3557", wrap=True)
    set_cell(COL_RESP, resp, align="center", sz=9)

    for col_d in (COL_INI, COL_FIM):
        dc = ws.cell(row=r, column=col_d)
        dc.number_format = "DD/MM/YYYY"
        dc.font = fnt(size=10)
        dc.fill = fill("FFFFFF")
        dc.alignment = Alignment(horizontal="center", vertical="center")
        dc.border = border()

    dur = ws.cell(row=r, column=COL_DUR, value=f"=F{r}-E{r}+1")
    dur.font = fnt(size=10)
    dur.fill = fill("F2F2F2")
    dur.alignment = Alignment(horizontal="center", vertical="center")
    dur.border = border()

    tipo_cell = ws.cell(row=r, column=COL_TIPO, value=tipo)
    if is_marco:
        tipo_cell.fill = fill("FF9900")
        tipo_cell.font = fnt(bold=True, color="FFFFFF", size=9)
    else:
        tipo_cell.fill = fill(row_bg)
        tipo_cell.font = fnt(size=9, color="595959")
    tipo_cell.alignment = Alignment(horizontal="center", vertical="center")
    tipo_cell.border = border()

    status_cell = ws.cell(row=r, column=COL_STATUS, value="Não Iniciado")
    status_cell.font = fnt(size=9)
    status_cell.fill = fill("EBEBEB")
    status_cell.alignment = Alignment(horizontal="center", vertical="center")
    status_cell.border = border()

    pct = ws.cell(row=r, column=COL_PCT,
                  value=f'=IF(I{r}="Concluído",1,IF(I{r}="Não Iniciado",0,""))')
    pct.number_format = "0%"
    pct.font = fnt(bold=True, size=10)
    pct.fill = fill("FFFFFF")
    pct.alignment = Alignment(horizontal="center", vertical="center")
    pct.border = border()

    ws.row_dimensions[r].height = 26

# Linhas extras
MAX_ROW = 50
last_task_row = TASK_ROW + len(TASKS) - 1
for extra_r in range(last_task_row + 1, MAX_ROW + 1):
    id_e = ws.cell(row=extra_r, column=COL_ID,
                   value=f"=IF(C{extra_r}<>\"\",COUNTA($C${TASK_ROW}:C{extra_r}),\"\")")
    id_e.font = fnt(bold=True, color="FFFFFF", size=10)
    id_e.fill = fill(TEAL)
    id_e.alignment = Alignment(horizontal="center", vertical="center")
    id_e.border = border()

    j_extra = ws.cell(row=extra_r, column=COL_PCT,
                      value=f'=IF(I{extra_r}="Concluído",1,IF(I{extra_r}="Não Iniciado",0,""))')
    j_extra.number_format = "0%"
    j_extra.font = fnt(bold=True, size=10)
    j_extra.fill = fill("FFFFFF")
    j_extra.alignment = Alignment(horizontal="center", vertical="center")
    j_extra.border = border()

    for col_d in (COL_INI, COL_FIM):
        dc = ws.cell(row=extra_r, column=col_d)
        dc.number_format = "DD/MM/YYYY"
        dc.alignment = Alignment(horizontal="center", vertical="center")
        dc.border = border()

    dur_e = ws.cell(row=extra_r, column=COL_DUR,
                    value=f"=IF(E{extra_r}=\"\",\"\",F{extra_r}-E{extra_r}+1)")
    dur_e.alignment = Alignment(horizontal="center", vertical="center")
    dur_e.fill = fill("F2F2F2")
    dur_e.border = border()

    ws.row_dimensions[extra_r].height = 26

# ── Dropdowns ─────────────────────────────────────────────────────────────
dv_fase = DataValidation(
    type="list",
    formula1=f"Listas!$E$2:$E${1 + len(FASE_NAMES)}",
    allow_blank=True, showDropDown=False, showErrorMessage=True,
    errorTitle="Fase inválida", error="Selecione uma das fases da lista.",
    showInputMessage=True, promptTitle="Fase do projeto",
    prompt="Selecione a fase à qual esta atividade pertence.")
ws.add_data_validation(dv_fase)
dv_fase.sqref = f"B{TASK_ROW}:B{MAX_ROW}"

dv_resp = DataValidation(
    type="list",
    formula1=f"Listas!$A$2:$A${1 + len(EQUIPE)}",
    allow_blank=True, showDropDown=False, showErrorMessage=True,
    errorTitle="Nome não encontrado", error="Selecione um responsável da lista.",
    showInputMessage=True, promptTitle="Responsável",
    prompt="Selecione o responsável pela atividade.")
ws.add_data_validation(dv_resp)
dv_resp.sqref = f"D{TASK_ROW}:D{MAX_ROW}"

dv_status = DataValidation(
    type="list",
    formula1=f"Listas!$C$2:$C${1 + len(STATUS_OPTIONS)}",
    allow_blank=False, showDropDown=False, showErrorMessage=True,
    errorTitle="Status inválido", error="Selecione: Não Iniciado, Em Andamento, Pausado ou Concluído.",
    showInputMessage=True, promptTitle="Status da atividade",
    prompt="Selecione o status e atualize o % na coluna ao lado.")
ws.add_data_validation(dv_status)
dv_status.sqref = f"I{TASK_ROW}:I{MAX_ROW}"

dv_data = DataValidation(
    type="date", operator="between",
    formula1="DATE(2025,1,1)", formula2="DATE(2030,12,31)",
    allow_blank=True, showDropDown=False, showErrorMessage=True,
    errorTitle="Data inválida", error="Digite uma data válida (DD/MM/AAAA).",
    showInputMessage=True, promptTitle="Data da atividade",
    prompt="Digite a data no formato DD/MM/AAAA.")
ws.add_data_validation(dv_data)
dv_data.sqref = f"E{TASK_ROW}:F{MAX_ROW}"

# ── Formatação Condicional ─────────────────────────────────────────────────
row_cf_rng = f"B{TASK_ROW}:J{MAX_ROW}"
for fase_key, phase in reversed(list(PHASES.items())):
    ws.conditional_formatting.add(row_cf_rng, FormulaRule(
        formula=[f'LEFT($B{TASK_ROW},6)="{fase_key}"'],
        fill=PatternFill("solid", fgColor=phase["bg"])))
ws.conditional_formatting.add(row_cf_rng, FormulaRule(
    formula=[f'OR($H{TASK_ROW}="Marco",$H{TASK_ROW}="Reunião")'],
    fill=PatternFill("solid", fgColor="FFF3CD")))

id_cf_rng = f"A{TASK_ROW}:A{MAX_ROW}"
for fase_key, phase in reversed(list(PHASES.items())):
    ws.conditional_formatting.add(id_cf_rng, FormulaRule(
        formula=[f'LEFT($B{TASK_ROW},6)="{fase_key}"'],
        fill=PatternFill("solid", fgColor=phase["color"]),
        font=Font(name="Calibri", bold=True, color="FFFFFF", size=10)))
ws.conditional_formatting.add(id_cf_rng, FormulaRule(
    formula=[f'OR($H{TASK_ROW}="Marco",$H{TASK_ROW}="Reunião")'],
    fill=PatternFill("solid", fgColor=LARANJA),
    font=Font(name="Calibri", bold=True, color="FFFFFF", size=10)))

fase_cf_rng = f"B{TASK_ROW}:B{MAX_ROW}"
for fase_key, phase in reversed(list(PHASES.items())):
    ws.conditional_formatting.add(fase_cf_rng, FormulaRule(
        formula=[f'LEFT($B{TASK_ROW},6)="{fase_key}"'],
        font=Font(name="Calibri", bold=False, color=phase["color"], size=9)))

status_rng = f"I{TASK_ROW}:I{MAX_ROW}"
ws.conditional_formatting.add(status_rng, FormulaRule(
    formula=[f'I{TASK_ROW}="Concluído"'],
    fill=PatternFill("solid", fgColor="C6EFCE"),
    font=Font(name="Calibri", bold=True, color="375623")))
ws.conditional_formatting.add(status_rng, FormulaRule(
    formula=[f'I{TASK_ROW}="Em Andamento"'],
    fill=PatternFill("solid", fgColor="FFEB9C"),
    font=Font(name="Calibri", bold=True, color="9C5700")))
ws.conditional_formatting.add(status_rng, FormulaRule(
    formula=[f'I{TASK_ROW}="Pausado"'],
    fill=PatternFill("solid", fgColor="FCE4D6"),
    font=Font(name="Calibri", bold=True, color="C55A11")))
ws.conditional_formatting.add(status_rng, FormulaRule(
    formula=[f'I{TASK_ROW}="Não Iniciado"'],
    fill=PatternFill("solid", fgColor="EBEBEB"),
    font=Font(name="Calibri", color="595959")))

ws.conditional_formatting.add(f"J{TASK_ROW}:J{MAX_ROW}", ColorScaleRule(
    start_type="num", start_value=0, start_color="F8696B",
    mid_type="num",   mid_value=0.5,  mid_color="FFEB84",
    end_type="num",   end_value=1,    end_color="63BE7B"))

# ── Larguras das colunas ──────────────────────────────────────────────────
col_widths = {1: 4, 2: 34, 3: 55, 4: 14, 5: 13, 6: 13, 7: 7, 8: 14, 9: 16, 10: 10}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = f"A{TASK_ROW}"
ws.sheet_view.showGridLines = False

# ══════════════════════════════════════════════════════════════════════════
# ABA PAINEL
# ══════════════════════════════════════════════════════════════════════════
ws_p = wb.create_sheet("Painel", 0)
ws_p.sheet_view.showGridLines = False
N_COLS = 7

def p_fill_row(row, bg, h=22):
    for c in range(1, N_COLS + 1):
        cell = ws_p.cell(row=row, column=c)
        if not cell.value:
            cell.fill = fill(bg)
        cell.border = border()
    ws_p.row_dimensions[row].height = h

def section_header(row, text, h=26):
    ws_p.merge_cells(f"A{row}:{get_column_letter(N_COLS)}{row}")
    ws_p[f"A{row}"] = text
    ws_p[f"A{row}"].fill = fill("0E8FA3")
    ws_p[f"A{row}"].font = fnt(bold=True, color="FFFFFF", size=12)
    ws_p[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws_p.row_dimensions[row].height = h

def mlabel(row, col, text):
    c = ws_p.cell(row=row, column=col, value=text)
    c.font = fnt(bold=True, color="1A2A2A", size=10)
    c.fill = fill("EBF7F9")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border()

def mvalue(row, col, val, fmt=None, color="1A2A2A", sz=11, span=1):
    if span > 1:
        end_col = get_column_letter(col + span - 1)
        ws_p.merge_cells(f"{get_column_letter(col)}{row}:{end_col}{row}")
    c = ws_p.cell(row=row, column=col, value=val)
    c.font = fnt(bold=True, color=color, size=sz)
    c.fill = fill("FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
    if fmt:
        c.number_format = fmt

# ── Título e logo ─────────────────────────────────────────────────────────
if os.path.exists(LOGO_PATH):
    logo = XLImage(LOGO_PATH)
    logo.width  = 150
    logo.height = 50
    ws_p.add_image(logo, f"{get_column_letter(N_COLS - 1)}1")

ws_p.merge_cells(f"A1:{get_column_letter(N_COLS - 2)}1")
ws_p["A1"] = "PAINEL DE CONTROLE DO PROJETO"
ws_p["A1"].fill = fill("0A6A7A")
ws_p["A1"].font = fnt(bold=True, color="FFFFFF", size=18)
ws_p["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_p.row_dimensions[1].height = 58
for c in range(N_COLS - 1, N_COLS + 1):
    ws_p.cell(row=1, column=c).fill = fill("FFFFFF")

ws_p.merge_cells(f"A2:{get_column_letter(N_COLS)}2")
ws_p["A2"] = "Consultoria de Estruturação Documental  ·  Bom Futuro Agrícola  ·  116 linhas de fornecimento  ·  816 fornecedores ativos"
ws_p["A2"].fill = fill("0E8FA3")
ws_p["A2"].font = fnt(italic=True, color="FFFFFF", size=11)
ws_p["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_p.row_dimensions[2].height = 28
ws_p.row_dimensions[3].height = 10

# ── Métricas gerais ───────────────────────────────────────────────────────
last_r = MAX_ROW
section_header(4, "MÉTRICAS GERAIS DO PROJETO")

mlabel(5, 1, "Data de Início:")
mvalue(5, 2, f"=IFERROR(MIN(Cronograma!$E${TASK_ROW}:$E${last_r}),\"\")", fmt="DD/MM/YYYY", color="0A6A7A")
mlabel(5, 4, "Data de Entrega:")
mvalue(5, 5, f"=IFERROR(MAX(Cronograma!$F${TASK_ROW}:$F${last_r}),\"\")", fmt="DD/MM/YYYY", color="0A6A7A")
ws_p.cell(5, 3).fill = fill("EBF7F9"); ws_p.cell(5, 3).border = border()
ws_p.cell(5, 7).fill = fill("FFFFFF"); ws_p.cell(5, 7).border = border()

mlabel(6, 1, "Tarefas Concluídas:")
mvalue(6, 2, f'=COUNTIF(Cronograma!I{TASK_ROW}:I{last_r},"Concluído")', color="375623")
mlabel(6, 4, "Total de Atividades:")
mvalue(6, 5, f'=COUNTA(Cronograma!$C${TASK_ROW}:$C${last_r})')
ws_p.cell(6, 3).fill = fill("EBF7F9"); ws_p.cell(6, 3).border = border()
ws_p.cell(6, 7).fill = fill("FFFFFF"); ws_p.cell(6, 7).border = border()

mlabel(7, 1, "Em Andamento:")
mvalue(7, 2, f'=COUNTIF(Cronograma!I{TASK_ROW}:I{last_r},"Em Andamento")', color="9C5700")
mlabel(7, 4, "Pausadas:")
mvalue(7, 5, f'=COUNTIF(Cronograma!I{TASK_ROW}:I{last_r},"Pausado")', color="C55A11")
ws_p.cell(7, 3).fill = fill("EBF7F9"); ws_p.cell(7, 3).border = border()
ws_p.cell(7, 7).fill = fill("FFFFFF"); ws_p.cell(7, 7).border = border()

for r in range(5, 8):
    ws_p.row_dimensions[r].height = 22

ws_p.row_dimensions[8].height = 10

# ── Avanço geral ──────────────────────────────────────────────────────────
section_header(9, "AVANÇO GERAL DO PROJETO")

pct_formula = (
    f'=IFERROR(AVERAGEIF(Cronograma!$B${TASK_ROW}:$B${last_r},"Fase*",'
    f'Cronograma!$J${TASK_ROW}:$J${last_r}),0)'
)
bar_formula = (
    f'=REPT("█",ROUND(IFERROR(AVERAGEIF(Cronograma!$B${TASK_ROW}:$B${last_r},"Fase*",'
    f'Cronograma!$J${TASK_ROW}:$J${last_r}),0)*20,0))'
    f'&REPT("░",20-ROUND(IFERROR(AVERAGEIF(Cronograma!$B${TASK_ROW}:$B${last_r},"Fase*",'
    f'Cronograma!$J${TASK_ROW}:$J${last_r}),0)*20,0))'
)
mvalue(10, 1, pct_formula, fmt="0%", color="0A6A7A", sz=22)
ws_p.merge_cells(f"B10:{get_column_letter(N_COLS)}10")
ws_p["B10"] = bar_formula
ws_p["B10"].font = Font(name="Courier New", bold=True, color="0E8FA3", size=14)
ws_p["B10"].fill = fill("EBF7F9")
ws_p["B10"].alignment = Alignment(horizontal="left", vertical="center")
ws_p["B10"].border = border()
ws_p.row_dimensions[10].height = 42
ws_p.row_dimensions[11].height = 10

# ── Avanço por fase ───────────────────────────────────────────────────────
section_header(12, "AVANÇO POR FASE")

fase_hdr = ["Fase", "", "% Avanço", "Barra de Progresso", "", "", ""]
for c, h in enumerate(fase_hdr, 1):
    cell = ws_p.cell(row=13, column=c, value=h)
    cell.fill = fill("0A6A7A")
    cell.font = fnt(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()
ws_p.row_dimensions[13].height = 22

fase_list = [
    ("Fase 1 — Diagnóstico e Alinhamento",    PHASES["Fase 1"]["color"], "Fase 1"),
    ("Fase 2 — Estruturação Documental",      PHASES["Fase 2"]["color"], "Fase 2"),
    ("Fase 3 — Parametrização na Plataforma", PHASES["Fase 3"]["color"], "Fase 3"),
    ("Fase 4 — Governança e Documentação",    PHASES["Fase 4"]["color"], "Fase 4"),
    ("Fase 5 — Entrega e Consolidação",       PHASES["Fase 5"]["color"], "Fase 5"),
]

for idx, (fase_name, fase_color, fase_prefix) in enumerate(fase_list):
    r = 14 + idx
    ws_p.cell(row=r, column=1, value=fase_name).font = fnt(bold=True, color=fase_color, size=10)
    ws_p.cell(row=r, column=1).fill = fill("F8F9FA")
    ws_p.cell(row=r, column=1).border = border()
    ws_p.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws_p.cell(row=r, column=2).fill = fill(fase_color)
    ws_p.cell(row=r, column=2).border = border()

    pct_f = (
        f'=IFERROR(AVERAGEIF(Cronograma!$B${TASK_ROW}:$B${last_r},'
        f'"{fase_prefix}*",Cronograma!$J${TASK_ROW}:$J${last_r}),0)'
    )
    pct_cell = ws_p.cell(row=r, column=3, value=pct_f)
    pct_cell.number_format = "0%"
    pct_cell.font = fnt(bold=True, color=fase_color, size=12)
    pct_cell.fill = fill("FFFFFF")
    pct_cell.alignment = Alignment(horizontal="center", vertical="center")
    pct_cell.border = border()

    ws_p.merge_cells(f"D{r}:{get_column_letter(N_COLS)}{r}")
    bar_cell = ws_p.cell(row=r, column=4)
    bar_cell.value = f'=REPT("█",ROUND(C{r}*20,0))&REPT("░",20-ROUND(C{r}*20,0))'
    bar_cell.font = Font(name="Courier New", bold=True, color=fase_color, size=11)
    bar_cell.fill = fill("F8F9FA")
    bar_cell.alignment = Alignment(horizontal="left", vertical="center")
    bar_cell.border = border()
    ws_p.row_dimensions[r].height = 24

# ── Legenda ───────────────────────────────────────────────────────────────
ws_p.row_dimensions[20].height = 10
section_header(21, "LEGENDA DE STATUS", h=22)

legend = [
    ("Concluído",     "C6EFCE", "375623"),
    ("Em Andamento",  "FFEB9C", "9C5700"),
    ("Pausado",       "FCE4D6", "C55A11"),
    ("Não Iniciado",  "EBEBEB", "595959"),
    ("Marco/Reunião", "FF9900", "FFFFFF"),
]
for li, (lbl, bg, fg) in enumerate(legend):
    cell = ws_p.cell(row=22, column=li + 1, value=lbl)
    cell.fill = fill(bg)
    cell.font = fnt(bold=True, color=fg, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()
ws_p.row_dimensions[22].height = 22

ws_p.row_dimensions[23].height = 10
ws_p.merge_cells(f"A24:{get_column_letter(N_COLS)}24")
ws_p["A24"] = "Preencha as datas diretamente na aba Cronograma. O painel atualiza automaticamente ao salvar."
ws_p["A24"].font = fnt(italic=True, color="7F7F7F", size=9)
ws_p["A24"].fill = fill("F8F9FA")
ws_p["A24"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_p["A24"].border = border()
ws_p.row_dimensions[24].height = 26

# ── Larguras do Painel ────────────────────────────────────────────────────
painel_widths = {1: 30, 2: 12, 3: 5, 4: 28, 5: 12, 6: 14, 7: 14}
for c, w in painel_widths.items():
    ws_p.column_dimensions[get_column_letter(c)].width = w

# ══════════════════════════════════════════════════════════════════════════
# FINALIZAÇÃO
# ══════════════════════════════════════════════════════════════════════════
wb.active = ws_p
wb.save(OUTPUT_PATH)
print(f"OK - Base xlsx gerada: {OUTPUT_PATH}")

# ── Macro VBA (auto-sort por fase ao editar coluna B) ─────────────────────
VBA_SORT = '''
Private Sub Workbook_SheetChange(ByVal Sh As Object, ByVal Target As Range)
    If Sh.Name <> "Cronograma" Then Exit Sub
    If Intersect(Target, Sh.Range("B3:B50")) Is Nothing Then Exit Sub
    Application.EnableEvents = False
    Application.ScreenUpdating = False
    With Sh.Sort
        .SortFields.Clear
        .SortFields.Add Key:=Sh.Range("B3:B50"), _
            SortOn:=xlSortOnValues, Order:=xlAscending, _
            DataOption:=xlSortNormal
        .SetRange Sh.Range("A3:J50")
        .Header = xlNo
        .MatchCase = False
        .Orientation = xlTopToBottom
        .Apply
    End With
    Application.EnableEvents = True
    Application.ScreenUpdating = True
End Sub
'''

try:
    import xlwings as xw
    import winreg

    _reg_key_path = None
    _old_vbom = None
    for _ver in ["16.0", "15.0", "14.0"]:
        _path = rf"Software\Microsoft\Office\{_ver}\Excel\Security"
        try:
            _rk = winreg.OpenKey(winreg.HKEY_CURRENT_USER, _path, 0, winreg.KEY_ALL_ACCESS)
            try:
                _old_vbom, _ = winreg.QueryValueEx(_rk, "AccessVBOM")
            except:
                _old_vbom = 0
            winreg.SetValueEx(_rk, "AccessVBOM", 0, winreg.REG_DWORD, 1)
            winreg.CloseKey(_rk)
            _reg_key_path = _path
            break
        except:
            pass

    app_xw = xw.App(visible=False, add_book=False)
    try:
        if os.path.exists(OUTPUT_XLSM):
            os.remove(OUTPUT_XLSM)
        wb_xw = app_xw.books.open(OUTPUT_PATH)
        wb_xw.api.SaveAs(OUTPUT_XLSM, FileFormat=52)
        wb_vb_comp = None
        sheet_names = {wb_xw.api.Sheets(i+1).CodeName
                       for i in range(wb_xw.api.Sheets.Count)}
        for c in wb_xw.api.VBProject.VBComponents:
            if c.Type == 100 and c.Name not in sheet_names:
                wb_vb_comp = c
                break
        if wb_vb_comp is None:
            raise RuntimeError("Módulo ThisWorkbook não encontrado no VBProject")
        mod = wb_vb_comp.CodeModule
        if mod.CountOfLines > 0:
            mod.DeleteLines(1, mod.CountOfLines)
        mod.InsertLines(1, VBA_SORT.strip())
        wb_xw.save()
        wb_xw.api.Close(False)
        print(f"OK - Arquivo com macro gerado: {OUTPUT_XLSM}")
    finally:
        app_xw.quit()
        if _reg_key_path and _old_vbom is not None:
            try:
                _rk2 = winreg.OpenKey(winreg.HKEY_CURRENT_USER, _reg_key_path, 0, winreg.KEY_SET_VALUE)
                winreg.SetValueEx(_rk2, "AccessVBOM", 0, winreg.REG_DWORD, _old_vbom)
                winreg.CloseKey(_rk2)
            except:
                pass
except Exception as e:
    print(f"Aviso: macro nao injetada ({e}). Use o .xlsx normalmente.")
