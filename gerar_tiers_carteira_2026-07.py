"""
Gerador de Tiers de Clientes — layout compacto e profissional
Base: FIN_Contrato_EFCAZ + Faturamento Efcaz 2026 (09/07/2026)
Saída: Documentos/Tiers_Clientes_Efcaz_2026-07.xlsx
"""

import sys, io, os, zipfile, tempfile
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as OXLImage

ORIGINAL = r"Documentos\Tiers_Clientes_Efcaz com plano de ação.xlsx"
OUTPUT   = r"Documentos\Tiers_Clientes_Efcaz_2026-07.xlsx"

_tmp_logo = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
with zipfile.ZipFile(ORIGINAL) as z:
    _tmp_logo.write(z.read("xl/media/image1.png"))
_tmp_logo.close()
LOGO_PATH = _tmp_logo.name

# ─── Paleta ──────────────────────────────────────────────────────────────────
C_DARK_TEAL = "FF0A6A7A"
C_TEAL      = "FF0E8FA3"
C_INFO      = "FFEBF5FB"
C_ALT       = "FFF5F6FA"

C_A_HDR  = "FF27AE60"
C_A_ROW  = "FFD5F5E3"
C_BP_HDR = "FF2980B9"
C_BP_ROW = "FFD6EAF8"
C_B_HDR  = "FF5DADE2"
C_B_ROW  = "FFEAF2F8"
C_C_HDR  = "FF95A5A6"
C_C_ROW  = "FFF8F9F9"
WHITE    = "FFFFFFFF"
BLACK    = "FF000000"

def _side():
    return Side(style="thin", color="FFD5D8DC")

BRD = Border(left=_side(), right=_side(), top=_side(), bottom=_side())

# ─── Dados ───────────────────────────────────────────────────────────────────
TIER_A = [
    ("Alpargatas",                               16666.00, "Tier A", ""),
    ("Nestlé (Supera Promo + PDV)",              15000.00, "Tier A", "Faturamento varia por uso — R$15k é MRR médio contratado"),
    ("Lactalis",                                 11969.00, "Tier A", "Em assinatura — onboarding a iniciar"),
    ("Zurich Airport Brasil",                    10599.00, "Tier A", "Contrato anual — faturado anualmente"),
    ("Vinci Airports",                            7250.00, "Tier A", ""),
    ("CSU Digital",                               7000.00, "Tier A", ""),
    ("Bom Futuro Agrícola",                       6500.00, "Tier A", ""),
    ("ECTX S/A – Eucatex",                        5509.40, "Tier A", "→ Wagner"),
    ("Afonso França Engenharia",                  5387.78, "Tier A", "Aditivo Serasa em processamento"),
    ("Soluções Terceirizadas",                    5200.00, "Tier A", "Aditivo em análise — faturamento ainda R$3.200"),
    ("Unimed do Brasil + Sou (estratégico)",      4987.50, "Tier A", "Dois contratos: Brasil R$2.528 + Sou R$2.458 | Mesmo CNPJ"),
]

TIER_BP = [
    ("DATA Engenharia",                           4700.00, "Tier B+", ""),
    ("Geistlich Pharma do Brasil",                4634.00, "Tier B+", ""),
    ("Associação Adventista Norte Brasileira",    4309.00, "Tier B+", ""),
    ("Bunker One Combustíveis e Lubrificantes",   3839.00, "Tier B+", ""),
    ("Dock Brasil Engenharia e Serviços",         3794.00, "Tier B+", "Aditivo jun/2025–mai/2026 ✅"),
    ("Cielo",                                     3450.24, "Tier B+", "Contrato anual R$41.402,88 ÷ 12"),
    ("Cebrace",                                   2255.87, "Tier B+", "Conglomerado Saint-Gobain — elevado estrategicamente | → Vagner"),
]

TIER_B = [
    ("Unimed Vale dos Sinos",                     3000.00, "Tier B", "Novo cliente"),
    ("Norskan Offshore (DOF)",                    2699.00, "Tier B", ""),
    ("BRG Suplementos Nutricionais (Integral Médica)", 2637.73, "Tier B", "Valor aprovado — contrato em assinatura"),
    ("Sabará Químicos e Ingredientes",            2504.30, "Tier B", ""),
    ("Premier Pet",                               2500.00, "Tier B", ""),
    ("Tarkett",                                   2500.00, "Tier B", ""),
    ("TEGRAM",                                    2500.00, "Tier B", "Novo cliente — vendido por Vagner"),
    ("Transportes Cavalinho",                     2385.00, "Tier B", ""),
    ("Pacco",                                     2309.00, "Tier B", ""),
    ("Engesp",                                    2250.00, "Tier B", "Contrato anual R$27.000 ÷ 12"),
    ("Federação Paulista de Futebol",             2158.00, "Tier B", ""),
    ("Ponsse Latin America",                      2118.62, "Tier B", ""),
]

TIER_C = [
    ("Unimed Campo Grande",                       1967.35, "Tier C", ""),
    ("Asso Marítima Navegação",                   1899.00, "Tier C", ""),
    ("Amboretto Bombas",                          1700.00, "Tier C", ""),
    ("Unimed de Dourados",                        1672.00, "Tier C", ""),
    ("Fundição Alumetaf",                         1650.00, "Tier C", "Downgrade — era Tier B (R$3.300)"),
    ("Advtec Indústria e Comércio",                700.00, "Tier C", ""),
]

# ─── Helpers ─────────────────────────────────────────────────────────────────
def f(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def ft(bold=False, size=10, color=BLACK, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ─── ABA 1 — Tiers de Clientes ───────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tiers de Clientes"

# Larguras
ws.column_dimensions["A"].width = 5.0
ws.column_dimensions["B"].width = 44.0
ws.column_dimensions["C"].width = 16.0
ws.column_dimensions["D"].width = 10.0
ws.column_dimensions["E"].width = 38.0

# ── 3 linhas de header — compacto ────────────────────────────────────────────
ws.row_dimensions[1].height = 42.0   # Logo + Título
ws.row_dimensions[2].height = 18.0   # Critérios + MRR
ws.row_dimensions[3].height = 20.0   # Cabeçalhos de coluna

ws.freeze_panes = "A4"

# Linha 1 — Título (A1:C1) + área logo (D1:E1)
ws.merge_cells("A1:C1")
ws["A1"].value     = "Tiers de Clientes  —  Carteira Gabriel Vital | Efcaz SRM"
ws["A1"].fill      = f(C_DARK_TEAL)
ws["A1"].font      = ft(bold=True, size=13, color=WHITE)
ws["A1"].alignment = al("left", "center")

ws.merge_cells("D1:E1")
ws["D1"].fill = f(C_DARK_TEAL)

_logo1 = OXLImage(LOGO_PATH)
_logo1.width  = 120
_logo1.height = 36
_logo1.anchor = "D1"
ws.add_image(_logo1)

# Linha 2 — Critérios (A2:C2) + label MRR (D2) + valor (E2)
ws.merge_cells("A2:C2")
ws["A2"].value = (
    "Tier A ≥ R$5.000  |  B+ R$3.500–4.999 (ou estratégico)  |  "
    "B R$2.000–3.499  |  C < R$2.000  —  Atualizado: 09/07/2026"
)
ws["A2"].fill      = f(C_INFO)
ws["A2"].font      = ft(size=8, italic=True)
ws["A2"].alignment = al("left", "center")

ws["D2"].value     = "MRR Total"
ws["D2"].fill      = f(C_INFO)
ws["D2"].font      = ft(bold=True, size=9, color=C_DARK_TEAL)
ws["D2"].alignment = al("right", "center")

ws["E2"].value         = '=SUMIF($D$4:$D$4999,"Tier*",$C$4:$C$4999)'
ws["E2"].fill          = f(C_INFO)
ws["E2"].font          = ft(bold=True, size=10, color=C_DARK_TEAL)
ws["E2"].number_format = 'R$ #,##0.00'
ws["E2"].alignment     = al("left", "center")

# Linha 3 — Cabeçalhos de coluna
for col, (txt, h) in enumerate(
    [("Nº","center"),("Cliente","left"),("MRR / mês","center"),("Tier","center"),("Observação","left")],
    start=1
):
    c = ws.cell(row=3, column=col, value=txt)
    c.fill      = f(C_DARK_TEAL)
    c.font      = ft(bold=True, size=10, color=WHITE)
    c.alignment = al(h, "center")
    c.border    = BRD

# ── Bloco de tier ─────────────────────────────────────────────────────────────
def write_tier(ws, start_row, tier_key, hdr_color, odd_color,
               clients, faixa_txt, label_sec):
    # Header da seção
    ws.merge_cells(f"A{start_row}:B{start_row}")
    ws[f"A{start_row}"].value     = label_sec
    ws[f"A{start_row}"].fill      = f(hdr_color)
    ws[f"A{start_row}"].font      = ft(bold=True, size=10, color=WHITE)
    ws[f"A{start_row}"].alignment = al("left", "center")
    ws[f"A{start_row}"].border    = BRD

    ws[f"C{start_row}"].value         = f'=SUMIF($D$4:$D$4999,"{tier_key}",$C$4:$C$4999)'
    ws[f"C{start_row}"].fill          = f(hdr_color)
    ws[f"C{start_row}"].font          = ft(bold=True, size=10, color=WHITE)
    ws[f"C{start_row}"].number_format = 'R$ #,##0.00'
    ws[f"C{start_row}"].alignment     = al("right", "center")
    ws[f"C{start_row}"].border        = BRD

    ws[f"D{start_row}"].value     = f"{len(clients)} clientes"
    ws[f"D{start_row}"].fill      = f(hdr_color)
    ws[f"D{start_row}"].font      = ft(size=9, color=WHITE)
    ws[f"D{start_row}"].alignment = al("center", "center")
    ws[f"D{start_row}"].border    = BRD

    ws[f"E{start_row}"].value     = faixa_txt
    ws[f"E{start_row}"].fill      = f(hdr_color)
    ws[f"E{start_row}"].font      = ft(size=9, color=WHITE)
    ws[f"E{start_row}"].alignment = al("left", "center")
    ws[f"E{start_row}"].border    = BRD

    ws.row_dimensions[start_row].height = 19.5

    for idx, (cliente, mrr, tier_lbl, obs) in enumerate(clients, start=1):
        row = start_row + idx
        bg  = odd_color if idx % 2 == 1 else C_ALT
        ws.row_dimensions[row].height = 18.75

        c1 = ws.cell(row=row, column=1, value=float(idx))
        c1.fill = f(bg); c1.font = ft(size=10); c1.alignment = al("center","center"); c1.border = BRD

        c2 = ws.cell(row=row, column=2, value=cliente)
        c2.fill = f(bg); c2.font = ft(size=10); c2.alignment = al("left","center"); c2.border = BRD

        c3 = ws.cell(row=row, column=3, value=mrr)
        c3.fill = f(bg); c3.font = ft(size=10); c3.number_format = 'R$ #,##0.00'
        c3.alignment = al("right","center"); c3.border = BRD

        c4 = ws.cell(row=row, column=4, value=tier_lbl)
        c4.fill = f(bg); c4.font = ft(size=10); c4.alignment = al("center","center"); c4.border = BRD

        c5 = ws.cell(row=row, column=5, value=obs)
        c5.fill = f(bg); c5.font = ft(size=9, italic=bool(obs))
        c5.alignment = al("left","center", wrap=True); c5.border = BRD

    return start_row + 1 + len(clients)

# Dados a partir da linha 4
next_row = write_tier(ws, 4,  "Tier A",  C_A_HDR,  C_A_ROW,  TIER_A,
                      "MRR ≥ R$ 5.000/mês", "Tier A — Contas Estratégicas")

next_row = write_tier(ws, next_row, "Tier B+", C_BP_HDR, C_BP_ROW, TIER_BP,
                      "MRR R$ 3.500–R$ 4.999/mês (ou estratégico)", "Tier B+ — Alto Valor")

next_row = write_tier(ws, next_row, "Tier B",  C_B_HDR,  C_B_ROW,  TIER_B,
                      "MRR R$ 2.000–R$ 3.499/mês", "Tier B — Médio Valor")

next_row = write_tier(ws, next_row, "Tier C",  C_C_HDR,  C_C_ROW,  TIER_C,
                      "MRR < R$ 2.000/mês", "Tier C — Menor Ticket")

# ─── ABA 2 — Cadência por Tier ───────────────────────────────────────────────
ws2 = wb.create_sheet("Cadência por Tier")

ws2.column_dimensions["A"].width = 22.0
ws2.column_dimensions["B"].width = 27.0
ws2.column_dimensions["C"].width = 10.0
ws2.column_dimensions["D"].width = 16.0
ws2.column_dimensions["E"].width = 26.0
ws2.column_dimensions["F"].width = 13.0
ws2.column_dimensions["G"].width = 24.0
ws2.column_dimensions["H"].width = 14.0

ws2.row_dimensions[1].height = 42.0
ws2.row_dimensions[2].height = 18.0

_logo2 = OXLImage(LOGO_PATH)
_logo2.width  = 120
_logo2.height = 36
_logo2.anchor = "G1"
ws2.add_image(_logo2)

# Linha 1 — Título + Logo
ws2.merge_cells("A1:F1")
ws2["A1"].value     = "Cadência por Tier  —  Gabriel Vital | Efcaz SRM"
ws2["A1"].fill      = f(C_DARK_TEAL)
ws2["A1"].font      = ft(bold=True, size=13, color=WHITE)
ws2["A1"].alignment = al("left", "center")

ws2.merge_cells("G1:H1")
ws2["G1"].fill = f(C_DARK_TEAL)

# Linha 2 — critérios
ws2.merge_cells("A2:H2")
ws2["A2"].value = (
    "Tier A ≥ R$5.000  |  B+ R$3.500–4.999  |  B R$2.000–3.499  |  C < R$2.000  "
    "—  Contagem dinâmica via Tiers de Clientes"
)
ws2["A2"].fill      = f(C_INFO)
ws2["A2"].font      = ft(size=8, italic=True)
ws2["A2"].alignment = al("left", "center")

# Linha 3 — cabeçalhos
cad_hdrs = ["Tier","Faixa MRR","Clientes","MRR Total","Check-in","QBR","Canal Principal","SLA Resposta"]
for col, h in enumerate(cad_hdrs, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.fill      = f(C_DARK_TEAL)
    c.font      = ft(bold=True, size=10, color=WHITE)
    c.alignment = al("center", "center")
    c.border    = BRD

ws2.row_dimensions[3].height = 20.0
ws2.freeze_panes = "A4"

# Linhas 4–7 — dados por tier
cad_rows = [
    ("A – Estratégico",  "≥ R$ 5.000/mês",        "Tier A",  C_A_ROW,
     "Mensal (call / presencial)", "Bimensal",   "WhatsApp direto + E-mail", "Mesmo dia"),
    ("B+ – Alto Valor",  "R$ 3.500–R$ 4.999/mês *","Tier B+", C_BP_ROW,
     "Mensal (call)",              "Trimestral", "E-mail + WhatsApp",        "24 horas"),
    ("B – Médio Valor",  "R$ 2.000–R$ 3.499/mês",  "Tier B",  C_B_ROW,
     "Bimensal (call e e-mail)",   "Trimestral", "E-mail",                   "48 horas"),
    ("C – Menor Ticket", "< R$ 2.000/mês",          "Tier C",  C_C_ROW,
     "Trimestral (call e e-mail)", "Semestral",  "E-mail",                   "72 horas"),
]

for r, (lbl, faixa, tier_key, row_bg, checkin, qbr, canal, sla) in enumerate(cad_rows, start=4):
    cnt_f = f"=COUNTIF('Tiers de Clientes'!$D$4:$D$4999,\"{tier_key}\")"
    mrr_f = f"=SUMIF('Tiers de Clientes'!$D$4:$D$4999,\"{tier_key}\",'Tiers de Clientes'!$C$4:$C$4999)"

    vals = [lbl, faixa, cnt_f, mrr_f, checkin, qbr, canal, sla]
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.fill      = f(row_bg)
        c.font      = ft(size=10)
        c.alignment = al("center" if col in (1,3,4,6,7,8) else "left", "center")
        c.border    = BRD
        if col == 4:
            c.number_format = 'R$ #,##0.00'
    ws2.row_dimensions[r].height = 22.5

# Linha 8 — total
ws2.merge_cells("A8:B8")
ws2["A8"].value     = "MRR Total Ativo"
ws2["A8"].fill      = f(C_DARK_TEAL)
ws2["A8"].font      = ft(bold=True, size=10, color=WHITE)
ws2["A8"].alignment = al("left", "center")
ws2["A8"].border    = BRD

ws2["C8"].value     = "=SUM(C4:C7)"
ws2["C8"].fill      = f(C_DARK_TEAL)
ws2["C8"].font      = ft(bold=True, size=10, color=WHITE)
ws2["C8"].alignment = al("center", "center")
ws2["C8"].border    = BRD

ws2["D8"].value         = "=SUM(D4:D7)"
ws2["D8"].fill          = f(C_DARK_TEAL)
ws2["D8"].font          = ft(bold=True, size=10, color=WHITE)
ws2["D8"].number_format = 'R$ #,##0.00'
ws2["D8"].alignment     = al("center", "center")
ws2["D8"].border        = BRD

for col in range(5, 9):
    c = ws2.cell(row=8, column=col)
    c.fill = f(C_DARK_TEAL)
    c.border = BRD

ws2.row_dimensions[8].height = 19.5

# Linha 9 — nota de rodapé
ws2.merge_cells("A9:H9")
ws2["A9"].value     = "* Tier B+ pode incluir contas abaixo de R$3.500 por decisão estratégica (ex: conglomerados de grande porte)"
ws2["A9"].fill      = f(C_INFO)
ws2["A9"].font      = ft(size=8, italic=True)
ws2["A9"].alignment = al("left", "center")
ws2.row_dimensions[9].height = 15.75

# ─── Salvar ───────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
os.unlink(LOGO_PATH)
print(f"Salvo: {OUTPUT}")

total_a  = sum(m for _, m, _, _ in TIER_A)
total_bp = sum(m for _, m, _, _ in TIER_BP)
total_b  = sum(m for _, m, _, _ in TIER_B)
total_c  = sum(m for _, m, _, _ in TIER_C)
print(f"  Tier A  ({len(TIER_A):2d} clientes): R$ {total_a:,.2f}")
print(f"  Tier B+ ({len(TIER_BP):2d} clientes): R$ {total_bp:,.2f}")
print(f"  Tier B  ({len(TIER_B):2d} clientes): R$ {total_b:,.2f}")
print(f"  Tier C  ({len(TIER_C):2d} clientes): R$ {total_c:,.2f}")
print(f"  TOTAL: R$ {total_a + total_bp + total_b + total_c:,.2f}")
